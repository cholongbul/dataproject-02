import excellstyle
import shutil
import re
from urllib import request
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
import openpyxl
import sys, traceback, os
from PyQt5.QtWidgets import *
import pandas as pd
import module.module as md
class MyApp(QWidget):



    def __init__(self):
        super().__init__()
        self.initUI()

    ##GUI그리기
    def initUI(self):
        ##레이아웃 설정
        self.grid = QGridLayout()
        self.setLayout(self.grid)

        ##진단시작버튼
        self.pushButton = QPushButton("진단 시작")
        ##클릭 이벤트
        self.pushButton.clicked.connect(self.pushButtonClicked)
        ##파일명 라벨
        self.label = QLabel()
        ##상태라벨
        self.label2 = QLabel()

        ##레이아웃 좌표 설정
        self.grid.addWidget(self.pushButton, 0,0)
        self.grid.addWidget(self.label, 1, 0)
        self.grid.addWidget(self.label2, 3, 0)

        ##윈도우 타이틀
        self.setWindowTitle('오픈API 보고서 작성')
        ##윈도우 크기
        self.setGeometry(300, 100, 600, 300)
        ##종료버튼 활성화
        self.setAcceptDrops(True)
        self.show()
        ##자동진단프로그램이기에 자동 클릭
        self.pushButton.click()

    ##클릭 이벤트 함수
    def pushButtonClicked(self):
        ##리소스 폴더가 비었다면
        if len(os.listdir('./resource/')) == 0:
            self.messagebox_open('리소스 폴더가 비어있습니다.')

        servicekey_tail = 'WBaXX3pce9C9AKfYTQc5%2FXVYPXYJWfHVzWNaird%2Fv0f8C0zKhPFhjY10Tuf2QuiA83hfkGLzHknlOz5FWPbaDQ%3D%3D'
        code_df = pd.read_csv('document/codelist.csv')
        fname_list = os.listdir('./resource/')


        try:
            ##db커서 생성
            try:
                cursor = md.dbcursor_module(self)
            except:
                traceback.print_exc()

            ##리소스폴더속 파일 반복
            for fname in fname_list:
                print(fname)
                self.label.setText(fname)
                self.label.repaint()
                self.label2.setText("보고서 생성 중")
                self.label2.repaint()

                ##변수 선언
                result_path = './report2/' #출력 보고서 경로
                path = './resource/' + fname  #원본 리소스 경로
                templet = './보고서_템플릿.xlsx'

                code_cnt = 0
                s6cnt = 0
                s7cnt = 0
                num1_cnt1 = 0

                ##결과값 리스트
                result_list1 = []
                result_list2 = []
                result_list3 = []
                result_list4 = []
                result_list5 = []
                result_list6 = []
                result_list7 = []
                result_list8 = []

                opername_list =[]

                ##엑셀 열기
                wb_data = openpyxl.load_workbook(path)
                sheet_name_list = wb_data.sheetnames
                operation_len = len(sheet_name_list)

                ##파일아이디
                apiid = fname.rstrip('.xlsx')
                if apiid.endswith('_문서미제공'):
                    apiid = apiid.rstrip('_문서미제공')
                result_file = apiid + '_오픈API_진단결과보고서_.xlsx'

                try:
                    ##sql문 api목록 요청, 판다스에 담기
                    apilist_df = md.apilist_df_module(cursor,apiid)
                    dataformmat = apilist_df.iloc[0]['DATAFORMMAT']
                    docxname = apilist_df.iloc[0]['DOCXNAME']
                    api_name = apilist_df['APINM'][0]
                    api_type = apilist_df['APITYPE'][0]
                    organ_name = apilist_df['ORGANNM'][0]
                    api_portal_url = 'https://www.data.go.kr/data/' +  apiid +'/openapi.do'
                    print(apiid)
                    print(api_name)
                except:
                    traceback.print_exc()
                    self.pushButton.click()



                ##오퍼레이션 별로 시트를 나누기에 시트 수로 오퍼레이션 나누기
                for i in range(0, len(sheet_name_list)):

                    ##변수선언
                    req_name_index = []
                    res_name_index = []
                    portal_res_name_index = []
                    res_eng_name_list = []

                    ##오퍼 요청변수 판다스에 담기
                    oper_req_df = md.oper_req_df_A(self, cursor, apiid)

                    ##시트 정하기
                    sheet = sheet_name_list[i]
                    df_data = pd.read_excel(path, sheet_name=sheet)
                    
                    ##이름 코드 체크 함수(코드, 이름 순 리스트 리턴)
                    oper_code_name = md.oper_code_name_setter(self,df_data,opername_list,oper_req_df)
                    opercode = oper_code_name[0]
                    opername = oper_code_name[1]

                    ##오퍼 요청변수 판다스에 담기
                    oper_req_df = md.oper_req_df_B(self,cursor, apiid, opercode)

                    ##오퍼 응답변수 판다스에 담기
                    oper_res_df = md.oper_res_df_module(self,cursor, apiid, opercode)

                    ##api링크 조립
                    api_req_link = md.mkApilink(self,df_data,servicekey_tail)
                    headers = {
                        'User-Agent': ' Mozilla/5.0 (Windows NT 6.1; WOW64; rv:12.0) Gecko/20100101 Firefox/12.0'}
                    try:
                        api_req = request.Request(api_req_link, headers=headers)
                        ##api응답
                        req_html = md.req_html_module(self,api_req)
                        req_html2 = ''
                        req_html3 = ''
                    except Exception as e:
                        traceback.print_exc()
                        if str(e).startswith("'utf-8' codec can't decode"):
                            req_html = '이미지'
                        else:
                            req_html = "응답없음"
                    
                    ##JSON+XML일 경우 추가 요청,응답
                    if dataformmat == 'JSON+XML':
                        req_html_list = md.json_xml(self,api_req_link,oper_req_df, req_html)
                        req_html2 = req_html_list[0]
                        req_html3 = req_html_list[1]

                    ##첫번째 오퍼레이션이면 보고서 파일 생성
                    if i == 0:
                        result = shutil.copy(templet, result_path  + result_file)
                        wb_result = openpyxl.load_workbook(result)

                    ##요청항목
                    for req_l in range(0, len(df_data['항목명(영문)'])):  # 요청항목 영문명 길이만큼 반복
                        if str(df_data['항목명(영문)'].iloc[req_l]).lower() == 'nan':  # 요청항목이 NaN값이면 중단
                            break
                        req_name_index.append(req_l)  # 요청항목 인덱스를 담음

                    ##응답항목
                    for res_l in range(0, len(df_data['응답항목명(영문)'])):  # 응답항목 영문명 길이만큼 반복
                        if str(df_data['응답항목명(영문)'].iloc[res_l]).lower() == 'nan':  # 응답항목이 NaN값이면 중단
                            break
                        res_name_index.append(res_l)  # 응답항목 인덱스를 담음
                        res_eng_name_list.append(str(df_data['응답항목명(영문)'].iloc[res_l]).lower)  # 응답항목 이름을 리스트에 담음

                    ##포털
                    for res_r in range(0, len(oper_res_df['OPERCD'])):  # 응답항목 영문명 길이만큼 반복
                        if str(oper_res_df['COLNMEN'].iloc[res_r]).lower() == 'nan':  # 응답항목이 NaN값이면 중단
                            break
                        portal_res_name_index.append(res_r)  # 응답항목 인덱스를 담음
                    if len(portal_res_name_index) == 0:
                        portal_res_name_index.append(1)
                    
                    ##api호출
                    real_tagnames = []
                    if req_html.startswith('<') or req_html.startswith('This XML'):
                        tagnams = re.findall('<.*?>', req_html)
                        for tagnam in tagnams:
                            if not str(tagnam).startswith('</') and not str(tagnam).startswith('<?') and not str(tagnam).startswith(
                                    '<!'):
                                real_tagnames.append(tagnam.lstrip('<').rstrip('>').lower())
                    elif '{' in req_html:
                        real_tagnames = re.findall('"([^"]*)"', req_html)

                    real_tagnames = set(real_tagnames)
                    real_tagnames = sorted(real_tagnames)

                    ## 1번은 요청항목 기준
                    s3 = wb_result['10.S-1']
                    s3m_list = md.s3m(self, s3, oper_req_df, df_data, opername, result_list1, num1_cnt1)
                    s3 = s3m_list[0]
                    result_list1 = s3m_list[1]
                    num1_cnt1 = s3m_list[2]

                    ## 2번항목
                    s4 = wb_result['20.S-2']
                    s4m_list = md.s4m(self,s4, i, opername, api_req_link, df_data, result_list2, req_html)
                    result_list2 = s4m_list[1]

                    ## 3번 항목
                    s5 = wb_result['30.S-3']
                    s5m_list = md.s5m(self,s5, num1_cnt1, s3, result_list3)
                    result_list3 = s5m_list[1]

                    ## 4번 항목
                    ## 응답항목 기준
                    s6 = wb_result['40.S-4']
                    s6m_list = md.s6m(self, s6, res_name_index, df_data, s6cnt, opername, req_html, res_eng_name_list, real_tagnames,
                        result_list4, i)
                    result_list4 = s6m_list[1]
                    s6cnt = s6m_list[2]

                    ## 5번 항목
                    ## 응답항목 기준
                    s7 = wb_result['50.S-5']
                    s7m_list = md.s7m(self, s7, oper_res_df, s7cnt, opername, req_html, real_tagnames, portal_res_name_index,
                        result_list5, i)
                    result_list5 = s7m_list[1]
                    s7cnt = s7m_list[2]

                    ## 6번 항목
                    s8 = wb_result['60.S-6']
                    s8m_list = md.s8m(self,df_data, req_name_index,code_df, opername, s8, result_list6, code_cnt)
                    result_list6 = s8m_list[1]

                    ## 7번 항목
                    s9 = wb_result['70.S-7']
                    s9m_list = md.s9m(self, dataformmat, i, s9, opername, result_list7, req_html, req_html2, req_html3)
                    result_list7 = s9m_list[1]

                ##8번 항목
                s10 = wb_result['80.S-8']
                s10m_list = md.s10m(self, s10, operation_len, opername_list, result_list1, result_list2, result_list4, result_list8)
                result_list8 = s10m_list

                ##수준평가
                s1 = wb_result['00.수준평가']
                md.s1m(self, s1, organ_name, api_name, api_type, dataformmat, operation_len, docxname, api_portal_url,
                    result_list1, result_list2, result_list3, result_list4, result_list5, result_list6)
                ##추가진단
                s2 = wb_result['02.추가 진단']
                md.s2m(self, s2, organ_name, api_name, api_type, dataformmat, operation_len, docxname, api_portal_url,
                    result_list7, result_list8)

                wb_result.save(result)
                self.label2.setText("보고서 생성 완료")
                self.label2.repaint()
                print(api_name)

        except Exception as e:
            traceback.print_exc()
            if opername.strip() == '문서오류':
                self.pushButton.click()
                pass
            elif str(e).startswith('Message: Could not locate element with visible text:'):
                shutil.move(result_path  + result_file,'./error/' + result_file.rstrip('.xlsx') + '_오퍼명오류.xlsx')
                shutil.move(path, './리소스오류/' + fname.rstrip('.xlsx') + '_오퍼명오류.xlsx')
                self.pushButton.click()
                pass
            elif str(e).startswith("Invalid URL 'nan':"):
                shutil.move(result_path  + result_file,'./error/' + result_file.rstrip('.xlsx') + '_URL오류.xlsx')
                shutil.move(path, './리소스오류/' + fname.rstrip('.xlsx') + '_URL오류.xlsx')
                self.pushButton.click()
                pass
            elif str(e).startswith("'numpy.float64' object has no"):
                shutil.move(result_path  + result_file,'./error/' + result_file.rstrip('.xlsx') + '_빈칸오류.xlsx')
                shutil.move(path, './리소스오류/' + fname.rstrip('.xlsx') + '_빈칸오류.xlsx')
                self.pushButton.click()
                pass
            elif str(e).startswith('Message: no such element:'):
                shutil.move(result_path  + result_file,'./error/' + result_file.rstrip('.xlsx') + '_연결오류.xlsx')
                shutil.move(path, './리소스오류/' + fname.rstrip('.xlsx') + '_연결오류.xlsx')
                self.pushButton.click()
                pass
            else:
                shutil.move(result_path  + result_file,'./error/' + result_file.rstrip('.xlsx') + '_보고서생성에러.xlsx')
                shutil.move(path, './리소스오류/' + fname.rstrip('.xlsx') + '_보고서생성에러.xlsx')
                self.pushButton.click()
            self.label2.setText("보고서 생성 에러")
            self.label2.repaint()

            pass

    def messagebox_open(self,e):
        msgBox = QMessageBox()
        msgBox.setIcon(QMessageBox.Information)
        msgBox.setText(str(e))
        msgBox.setWindowTitle("파일문제")
        msgBox.setStandardButtons(QMessageBox.Ok)
        returnValue = msgBox.exec()



if __name__ == '__main__':
    app = QApplication(sys.argv)
    ex = MyApp()
    sys.exit(app.exec_())




