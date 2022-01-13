
import shutil

import time
import lxml
from selenium.webdriver.support.ui import Select
import requests
from openpyxl.styles import Font, Border, Side, Alignment
from selenium import webdriver
import re
from difflib import SequenceMatcher

import openpyxl

import sys, traceback, os

from PyQt5.QtWidgets import *


import pandas as pd

class MyApp(QWidget):

    def __init__(self):
        super().__init__()
        self.initUI()

    def initUI(self):
        self.grid = QGridLayout()
        self.setLayout(self.grid)

        self.pushButton = QPushButton("진단 시작")
        self.pushButton.clicked.connect(self.pushButtonClicked)
        self.label = QLabel()
        self.label2 = QLabel()

        self.grid.addWidget(self.pushButton, 0,0)
        self.grid.addWidget(self.label, 1, 0)
        self.grid.addWidget(self.label2, 3, 0)


        self.setWindowTitle('오픈API 보고서 작성')
        self.setGeometry(300, 100, 600, 300)

        self.setAcceptDrops(True)
        self.show()

    def pushButtonClicked(self):
            if len(os.listdir('./resource/')) == 0:
                self.messagebox_open('리소스 폴더가 비었습니다.')

            def style_range(ws, cell_range, border=Border(), fill=None, font=None,
                            alignment=None):
                """
                Apply styles to a range of cells as if they were a single cell.

                :param ws:  Excel worksheet instance
                :param range: An excel range to style (e.g. A1:F20)
                :param border: An openpyxl Border
                :param fill: An openpyxl PatternFill or GradientFill
                :param font: An openpyxl Font object
                """

                top = Border(top=border.top)
                left = Border(left=border.left)
                right = Border(right=border.right)
                bottom = Border(bottom=border.bottom)

                first_cell = ws[cell_range.split(":")[0]]
                if alignment:
                    ws.merge_cells(cell_range)
                    first_cell.alignment = alignment

                rows = ws[cell_range]
                if font:
                    first_cell.font = font

                for cell in rows[0]:
                    cell.border = cell.border + top
                for cell in rows[-1]:
                    cell.border = cell.border + bottom

                for row in rows:
                    l = row[0]
                    r = row[-1]
                    l.border = l.border + left
                    r.border = r.border + right
                    if fill:
                        for c in row:
                            c.fill = fill

            def recursive_items(dictionary):
                for key, value in dictionary.items():
                    if type(value) is dict:
                        yield (key, value)
                        yield from recursive_items(value)
                    else:
                        yield (key, value)

            def exel_font_set(sheet):
                sheet.font = Font(name="맑은 고딕")
                sheet.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                sheet.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'),
                                      bottom=Side(style='thin'))
            try:

                fname_list = os.listdir('./resource/')

                for fname in fname_list:
                    self.label.setText(fname)
                    self.label.repaint()
                    self.label2.setText("보고서 생성 중")
                    self.label2.repaint()

                    path = './resource/' + fname

                    # 크롤링을 위한 드라이버
                    options = webdriver.ChromeOptions()
                    options.add_argument("headless")
                    options.add_argument('--window-size=1024,768')
                    options.add_argument("--disable-gpu")
                    options.add_experimental_option('excludeSwitches', ['enable-logging'])
                    templet = '보고서_템플릿.xlsx'
                    result_path = './report/'

                    ##사용자가 엑셀에 입력한 값(참고 문서 데이터)
                    data = path

                    wb_data = openpyxl.load_workbook(data)
                    sheet_name_list = wb_data.sheetnames

                    temp_req = []
                    temp_res = []
                    s6cnt = 0
                    s7cnt = 0

                    num1_cnt1 = 0
                    num1_cnt2 = 0
                    num3_cnt2 = 0
                    portal_temp_res = []
                    result_list1 = []
                    result_list2 = []
                    result_list3 = []
                    result_list4 = []
                    result_list5 = []
                    result_list8 = []
                    chromedriver_path = './chromedriver.exe'
                    operation_len = len(sheet_name_list)
                    opername_list = []
                    #chromedriver_path = os.path.join(sys._MEIPASS,'./chromedriver.exe')


                    ##오퍼레이션 별로 시트를 나누기에 시트 수로 오퍼레이션 나누기
                    for i in range(0, len(sheet_name_list)):
                        driver = webdriver.Chrome(executable_path=chromedriver_path, options=options)
                        sheet = sheet_name_list[i]
                        ws_data = wb_data[sheet]
                        df_data = pd.read_excel(data, sheet_name=sheet)
                        opername = df_data.iloc[0, 0]
                        api_req_link = str(df_data.iloc[0, 12])
                        opername_list.append(opername)

                        ##문서 오류 데이터
                        if df_data.iloc[0, 0].strip() == '문서오류':

                            ##사용자가 준 링크를 포털에서 크롤링하여 불러오기
                            potal_link = str(ws_data['L2'].value).replace(' ','')
                            driver.get(url=potal_link)
                            driver.implicitly_wait(3)
                            api_name = driver.find_element_by_css_selector(
                                '#contents > div.data-search-view > div.data-set-title.open-api > div.tit-area > p').text


                            portal_html = driver.page_source
                            time.sleep(1)

                            driver.close()

                            html_df_list = pd.read_html(portal_html, encoding='utf8')
                            html_meta_df = html_df_list[0]

                            ##모인 데이터로 엑셀 조립하기
                            if i == 0:
                                result = shutil.copy(templet, result_path + api_name + '_오픈API_문서오류_보고서.xlsx')
                                wb_result = openpyxl.load_workbook(result)

                            s1 = wb_result['00.수준평가']
                            s1['C5'].value = html_meta_df.iloc[0, 3]
                            s1['C6'].value = api_name
                            s1['C7'].value = html_meta_df.iloc[2, 1]
                            s1['C8'].value = html_meta_df.iloc[2, 3]
                            s1['C9'].value = len(sheet_name_list)
                            s1['C10'].value = html_meta_df.iloc[8, 1]
                            s1['C11'].value = str(ws_data['L2'].value).replace(' ','')

                            s1.merge_cells("D14:G19")
                            style_range(s1, "D14:G19",
                                        border=Border(left=Side(style='thin'), right=Side(style='thin'),
                                                      top=Side(style='thin'),
                                                      bottom=Side(style='thin')))

                            s1["D14"].value = "참고 문서 불완전으로 인한 진단불가"
                            s1["G20"].value = ""
                            exel_font_set(s1['D14'])
                            break



                        ##사용자가 준 링크를 포털에서 크롤링하여 불러오기
                        potal_link = str(ws_data['L2'].value).replace(' ','')
                        driver.get(url=potal_link)
                        driver.implicitly_wait(3)
                        api_name = driver.find_element_by_css_selector(
                            '#contents > div.data-search-view > div.data-set-title.open-api > div.tit-area > p').text



                        selectbox = Select(driver.find_element_by_css_selector('#open_api_detail_select'))
                        selectbox.select_by_visible_text(df_data.iloc[0, 0].strip())
                        button = driver.find_element_by_css_selector('#apiDetailFunctionDiv > div.open-api-detail-select > button')
                        button.click()
                        time.sleep(1)
                        portal_html = driver.page_source
                        time.sleep(1)


                        html_df_list = pd.read_html(portal_html, encoding='utf8')
                        html_req_data = html_df_list[1]
                        html_meta_df = html_df_list[0]
                        portal_res_df = html_df_list[2]

                        try:
                            driver.set_page_load_timeout(30)
                            driver.get(url=df_data.iloc[0, 12])
                            time.sleep(1)
                            req_html = str(driver.find_element_by_tag_name("body").text)
                            driver.close()
                        except:
                            print("응답없음")
                            req_html = "응답없음"

                        ##모인 데이터로 엑셀 조립하기
                        if i == 0:
                            file_name = html_meta_df[1][0]
                            result = shutil.copy(templet, result_path  + api_name.replace('/', '').replace('\\', '').replace(':', '').replace('*', '').replace('?', '').replace('|','').replace('<','').replace('>','') + '_오픈API_진단결과보고서_.xlsx')
                            wb_result = openpyxl.load_workbook(result)

                        s3 = wb_result['10.S-1']
                        if str(df_data['항목명(영문)'][0]) != 'nan':
                            df_data_sorted = df_data.iloc[df_data['항목명(영문)'].str.lower().argsort()]
                        else:
                            df_data_sorted = df_data
                        if len(html_req_data['항목명(영문)'].index) != 0 and str(html_req_data['항목명(영문)'][0]) != 'nan':
                            html_req_data_sorted = html_req_data.iloc[html_req_data['항목명(영문)'].str.lower().argsort()]
                        else:
                            html_req_data_sorted = html_req_data
                        req_name_index = []
                        res_name_index = []
                        portal_res_name_index = []
                        portal_eng_name_list = []
                        res_eng_name_list = []
                        ##요청항목 리스트 기준
                        for req_l in range(0, len(df_data['항목명(영문)'])):  # 요청항목 영문명 길이만큼 반복
                            if str(df_data['항목명(영문)'].iloc[req_l]).lower() == 'nan':  # 요청항목이 NaN값이면 중단
                                break
                            req_name_index.append(req_l)  # 요청항목 인덱스를 담음

                        temp_req.append(len(req_name_index))  # 인덱스 길이를 임시 리스트에 저장

                        ## 출력항목 리스트 기준
                        for res_l in range(0, len(df_data['응답항목명(영문)'])):  # 응답항목 영문명 길이만큼 반복
                            if str(df_data['응답항목명(영문)'].iloc[res_l]).lower() == 'nan':  # 응답항목이 NaN값이면 중단
                                break
                            res_name_index.append(res_l)  # 응답항목 인덱스를 담음
                            res_eng_name_list.append(df_data['응답항목명(영문)'].iloc[res_l])  # 응답항목 이름을 리스트에 담음
                        temp_res.append(len(res_name_index))  # 인덱스 길이를 임시 리스트에 저장
                        ##포털
                        for res_r in range(0, len(portal_res_df['항목명(영문)'])):  # 응답항목 영문명 길이만큼 반복
                            if str(portal_res_df['항목명(영문)'].iloc[res_r]).lower() == 'nan':  # 응답항목이 NaN값이면 중단
                                break
                            portal_res_name_index.append(res_r)  # 응답항목 인덱스를 담음
                            portal_eng_name_list.append(portal_res_df['항목명(영문)'].iloc[res_r])  # 응답항목 이름을 리스트에 담음
                        portal_temp_res.append(len(portal_res_name_index))
                        ##오퍼레이션 추가 고려 요청항목 길이
                        req_length = 0
                        if i == 0:
                            req_length = 0  ##처음은 0부터
                        else:
                            for muimi in range(0, i):
                                req_length = req_length + temp_req[muimi]  ##임시 리스트에 저장한 값을 더해나감
                        ##오퍼레이션 추가 고려 응답항목 길이
                        res_length = 0
                        if i == 0:
                            res_length = 0
                        else:
                            for muimi in range(0, i):
                                res_length = res_length + temp_res[muimi]
                        ##xml 태그 이름 추출
                        portal_res_lenth = 0
                        if i == 0:
                            portal_res_lenth = 0
                        else:
                            for muimi in range(0, i):
                                portal_res_lenth = portal_res_lenth + portal_temp_res[muimi]


                        real_tagnames = []
                        if req_html.startswith('<') or req_html.startswith('This XML'):
                            tagnams = re.findall('<.*?>', req_html)
                            for tagnam in tagnams:
                                if not str(tagnam).startswith('</') and not str(tagnam).startswith('<?') and not str(tagnam).startswith(
                                        '<!'):
                                    real_tagnames.append(tagnam.lstrip('<').rstrip('>'))
                        elif '{' in req_html:
                            real_tagnames = re.findall('"([^"]*)"', req_html)
                        real_tagnames = set(real_tagnames)
                        real_tagnames = sorted(real_tagnames)
                        num1htmllist = html_req_data['항목명(영문)'].values.tolist()
                        num1reqlist = df_data['항목명(영문)'].values.tolist()
                        ## 1번항목 *iloc(y,x)
                        ## 1번은 요청항목 기준
                        resultlist1_temp = []
                        num1reqlist_temp = []
                        num1htmllist_temp = []
                        for j in num1reqlist:  ##리소스에 적혀있는 요청변수만큼 반복한다
                            if str(j) == 'nan':  ##만약 리소스에 적혀 있는 요청변수가 비어있다면 반복을 멈춘다
                                break
                            if len(num1htmllist) == 0:  ##만약 포털이 비어있다면 멈춘다
                                break
                            if num1htmllist[0] == '검색 결과가 없습니다.':
                                break
                            for jk in num1htmllist:
                                if str(j).lower().strip() == str(jk).lower().strip():
                                    num1reqlist_temp.append(j)
                                    num1htmllist_temp.append(jk)

                                    num1_cnt1 = num1_cnt1 + 1
                                    s3['A' + str(8 + num1_cnt1 - 1)].value = str(num1_cnt1)
                                    exel_font_set(s3['A' + str(8 + num1_cnt1 - 1)])

                                    s3['B' + str(8 + num1_cnt1 - 1)].value = opername
                                    exel_font_set(s3['B' + str(8 + num1_cnt1 - 1)])

                                    s3['C' + str(8 + num1_cnt1 - 1)].value = j
                                    exel_font_set(s3['C' + str(8 + num1_cnt1 - 1)])
                                    try:
                                        s3['D' + str(8 + num1_cnt1 - 1)].value = \
                                            df_data[df_data['항목명(영문)'] == j]['항목명(국문)'].values[0]
                                        exel_font_set(s3['D' + str(8 + num1_cnt1 - 1)])
                                    except IndexError:
                                        s3['D' + str(8 + num1_cnt1 - 1)].value = ''
                                        exel_font_set(s3['D' + str(8 + num1_cnt1 - 1)])
                                    try:
                                        s3['E' + str(8 + num1_cnt1 - 1)].value = \
                                            df_data[df_data['항목명(영문)'] == j]['항목크기'].values[0]
                                        exel_font_set(s3['E' + str(8 + num1_cnt1 - 1)])
                                    except IndexError:
                                        s3['E' + str(8 + num1_cnt1 - 1)].value = ''
                                        exel_font_set(s3['E' + str(8 + num1_cnt1 - 1)])
                                    try:
                                        s3['F' + str(8 + num1_cnt1 - 1)].value = \
                                            df_data[df_data['항목명(영문)'] == j]['항목구분'].values[0]
                                        exel_font_set(s3['F' + str(8 + num1_cnt1 - 1)])
                                    except IndexError:
                                        s3['F' + str(8 + num1_cnt1 - 1)].value = ''
                                        exel_font_set(s3['F' + str(8 + num1_cnt1 - 1)])
                                    s3['G' + str(8 + num1_cnt1 - 1)].value = str(num1_cnt1)
                                    exel_font_set(s3['G' + str(8 + num1_cnt1 - 1)])
                                    s3['H' + str(8 + num1_cnt1 - 1)].value = opername
                                    exel_font_set(s3['H' + str(8 + num1_cnt1 - 1)])
                                    s3['I' + str(8 + num1_cnt1 - 1)].value = jk
                                    exel_font_set(s3['I' + str(8 + num1_cnt1 - 1)])
                                    s3['J' + str(8 + num1_cnt1 - 1)].value = \
                                        html_req_data[html_req_data['항목명(영문)'] == jk]['항목명(국문)'].values[0]
                                    exel_font_set(s3['J' + str(8 + num1_cnt1 - 1)])
                                    s3['K' + str(8 + num1_cnt1 - 1)].value = \
                                        html_req_data[html_req_data['항목명(영문)'] == jk]['항목크기'].values[0]
                                    exel_font_set(s3['K' + str(8 + num1_cnt1 - 1)])
                                    s3['L' + str(8 + num1_cnt1 - 1)].value = \
                                        html_req_data[html_req_data['항목명(영문)'] == jk]['항목구분'].values[0]
                                    exel_font_set(s3['L' + str(8 + num1_cnt1 - 1)])
                                    s3['M' + str(8 + num1_cnt1 - 1)].value = '정상'
                                    resultlist1_temp.append(s3['M' + str(8 + num1_cnt1 - 1)].value)
                                    exel_font_set(s3['M' + str(8 + num1_cnt1 - 1)])

                        num1reqlist = [x for x in num1reqlist if x not in num1reqlist_temp]
                        print(num1reqlist)
                        num1htmllist = [x for x in num1htmllist if x not in num1htmllist_temp]
                        print(num1htmllist)

                        if len(num1reqlist) > 0:
                            for j in num1reqlist:
                                if str(j) == 'nan':
                                    continue
                                num1_cnt1 = num1_cnt1 + 1
                                s3['A' + str(8 + num1_cnt1 - 1)].value = str(num1_cnt1)
                                exel_font_set(s3['A' + str(8 + num1_cnt1 - 1)])
                                s3['G' + str(8 + num1_cnt1 - 1)].value = str(num1_cnt1)
                                exel_font_set(s3['G' + str(8 + num1_cnt1 - 1)])
                                s3['B' + str(8 + num1_cnt1 - 1)].value = opername
                                exel_font_set(s3['B' + str(8 + num1_cnt1 - 1)])
                                s3['H' + str(8 + num1_cnt1 - 1)].value = opername
                                exel_font_set(s3['H' + str(8 + num1_cnt1 - 1)])
                                s3['C' + str(8 + num1_cnt1 - 1)].value = j
                                exel_font_set(s3['C' + str(8 + num1_cnt1 - 1)])
                                s3['D' + str(8 + num1_cnt1 - 1)].value = \
                                    df_data[df_data['항목명(영문)'] == j]['항목명(국문)'].values[0]
                                exel_font_set(s3['D' + str(8 + num1_cnt1 - 1)])
                                s3['E' + str(8 + num1_cnt1 - 1)].value = \
                                    df_data[df_data['항목명(영문)'] == j]['항목크기'].values[0]
                                exel_font_set(s3['E' + str(8 + num1_cnt1 - 1)])
                                s3['F' + str(8 + num1_cnt1 - 1)].value = \
                                    df_data[df_data['항목명(영문)'] == j]['항목구분'].values[0]
                                exel_font_set(s3['F' + str(8 + num1_cnt1 - 1)])
                                s3['I' + str(8 + num1_cnt1 - 1)].value = '동일항목없음'
                                exel_font_set(s3['I' + str(8 + num1_cnt1 - 1)])
                                s3['J' + str(8 + num1_cnt1 - 1)].value = '동일항목없음'
                                exel_font_set(s3['J' + str(8 + num1_cnt1 - 1)])
                                exel_font_set(s3['K' + str(8 + num1_cnt1 - 1)])
                                exel_font_set(s3['L' + str(8 + num1_cnt1 - 1)])
                                s3['M' + str(8 + num1_cnt1 - 1)].value = '오류'
                                resultlist1_temp.append(s3['M' + str(8 + num1_cnt1 - 1)].value)
                                exel_font_set(s3['M' + str(8 + num1_cnt1 - 1)])

                        if len(num1htmllist) > 0:
                            for j in num1htmllist:
                                if str(j) == 'nan':
                                    continue
                                elif str(j) == '검색 결과가 없습니다.':
                                    continue
                                num1_cnt1 = num1_cnt1 + 1
                                s3['G' + str(8 + num1_cnt1 - 1)].value = str(num1_cnt1 - 1 + 1)
                                exel_font_set(s3['G' + str(8 + num1_cnt1 - 1)])
                                s3['A' + str(8 + num1_cnt1 - 1)].value = str(num1_cnt1 - 1 + 1)
                                exel_font_set(s3['A' + str(8 + num1_cnt1 - 1)])
                                s3['H' + str(8 + num1_cnt1 - 1)].value = opername
                                exel_font_set(s3['H' + str(8 + num1_cnt1 - 1)])
                                s3['B' + str(8 + num1_cnt1 - 1)].value = opername
                                exel_font_set(s3['B' + str(8 + num1_cnt1 - 1)])
                                s3['I' + str(8 + num1_cnt1 - 1)].value = j
                                exel_font_set(s3['I' + str(8 + num1_cnt1 - 1)])
                                s3['J' + str(8 + num1_cnt1 - 1)].value = \
                                    html_req_data[html_req_data['항목명(영문)'] == j]['항목명(국문)'].values[0]
                                exel_font_set(s3['J' + str(8 + num1_cnt1 - 1)])
                                s3['K' + str(8 + num1_cnt1 - 1)].value = \
                                    html_req_data[html_req_data['항목명(영문)'] == j]['항목크기'].values[0]
                                exel_font_set(s3['K' + str(8 + num1_cnt1 - 1)])
                                s3['L' + str(8 + num1_cnt1 - 1)].value = \
                                    html_req_data[html_req_data['항목명(영문)'] == j]['항목구분'].values[0]
                                exel_font_set(s3['L' + str(8 + num1_cnt1 - 1)])
                                s3['M' + str(8 + num1_cnt1 - 1)].value = '오류'
                                resultlist1_temp.append(s3['M' + str(8 + num1_cnt1 - 1)].value)
                                exel_font_set(s3['M' + str(8 + num1_cnt1 - 1)])
                                exel_font_set(s3['B' + str(8 + num1_cnt1 - 1)])
                                s3['C' + str(8 + num1_cnt1 - 1)].value = '동일항목없음'
                                s3['D' + str(8 + num1_cnt1 - 1)].value = '동일항목없음'
                                exel_font_set(s3['C' + str(8 + num1_cnt1 - 1)])
                                exel_font_set(s3['D' + str(8 + num1_cnt1 - 1)])
                                exel_font_set(s3['E' + str(8 + num1_cnt1 - 1)])
                                exel_font_set(s3['F' + str(8 + num1_cnt1 - 1)])

                        if '오류' in resultlist1_temp:
                            result_list1.append('오류')
                        else:
                            result_list1.append('정상')

                        s4 = wb_result['20.S-2']
                        ## 2번항목
                        s4['A' + str(8 + i)].value = str(i + 1)
                        exel_font_set(s4['A' + str(8 + i)])
                        s4['B' + str(8 + i)].value = opername
                        exel_font_set(s4['B' + str(8 + i)])
                        s4['C' + str(8 + i)].value = api_req_link
                        exel_font_set(s4['C' + str(8 + i)])
                        s4['D' + str(8 + i)].value = df_data.iloc[0, 13]
                        exel_font_set(s4['D' + str(8 + i)])
                        s4['E' + str(8 + i)].value = str(i + 1)
                        exel_font_set(s4['E' + str(8 + i)])
                        s4['F' + str(8 + i)].value = opername
                        exel_font_set(s4['F' + str(8 + i)])
                        s4['G' + str(8 + i)].value = api_req_link
                        exel_font_set(s4['G' + str(8 + i)])
                        s4['H' + str(8 + i)].value = req_html[:1000]
                        exel_font_set(s4['H' + str(8 + i)])
                        matchraio = int(
                            SequenceMatcher(None, str(req_html.replace(
                                'This XML file does not appear to have any style information associated with it. The document tree is shown below.',
                                '')).strip().replace('\n', '').replace('\t', '').replace(' ', '')[:200],
                                            str(df_data.iloc[0, 13]).strip().replace('\n', '').replace('\t',
                                                                                                       '').replace(' ',
                                                                                                                   '')[
                                            :200]).ratio() * 100)
                        if 'normal service' in req_html.lower():
                            s4['I' + str(8 + i)].value = "정상"
                        elif 'normal_service' in req_html.lower():
                            s4['I' + str(8 + i)].value = "정상"
                        elif req_html == '응답없음':
                            s4['I' + str(8 + i)].value = "오류"
                        elif 'success' in req_html.lower():
                            s4['I' + str(8 + i)].value = "정상"
                        elif '정상' in req_html:
                            s4['I' + str(8 + i)].value = "정상"
                        elif '이미지' in req_html:
                            s4['I' + str(8 + i)].value = "정상"
                        elif 'success_info' in req_html.lower():
                            s4['I' + str(8 + i)].value = "정상"
                        elif matchraio > 70:
                            s4['I' + str(8 + i)].value = "정상"
                        elif '<successYN>N</successYN>' in req_html.lower():
                            s4['I' + str(8 + i)].value = "오류"
                        elif 'normal' in req_html.lower():
                            s4['I' + str(8 + i)].value = '정상'
                        elif 'nomal' in req_html.lower():
                            s4['I' + str(8 + i)].value = '정상'
                        elif '<successYN>Y</successYN>' in req_html.lower():
                            s4['I' + str(8 + i)].value = '정상'
                        elif '<resultcode>00' in req_html.lower():
                            s4['I' + str(8 + i)].value = '정상'
                        elif 'wfs' in req_html.lower():
                            s4['I' + str(8 + i)].value = '정상'
                        elif 'soapenv' in req_html.lower():
                            s4['I' + str(8 + i)].value = "오류"
                        elif 'errorcode' in req_html.lower():
                            s4['I' + str(8 + i)].value = "오류"
                        elif 'service error' in req_html.lower():
                            s4['I' + str(8 + i)].value = "오류"
                        elif 'SERVICE KEY IS NOT' in req_html.upper():
                            s4['I' + str(8 + i)].value = "오류"
                        elif 'APPLICATION ERROR' in req_html.upper():
                            s4['I' + str(8 + i)].value = "오류"
                        elif 'DB ERROR' in req_html.upper():
                            s4['I' + str(8 + i)].value = "오류"
                        elif 'NODATA ERROR' in req_html.upper():
                            s4['I' + str(8 + i)].value = "오류"
                        elif 'HTTP ERROR' in req_html.upper():
                            s4['I' + str(8 + i)].value = "오류"
                        elif 'SERVICETIMEOUT ERROR' in req_html.upper():
                            s4['I' + str(8 + i)].value = "오류"
                        elif 'INVALID_REQUEST PARAMETER_ERROR' in req_html.upper():
                            s4['I' + str(8 + i)].value = "오류"
                        elif 'NO MANDATORY REQUEST PARAMETERS ERROR' in req_html.upper():
                            s4['I' + str(8 + i)].value = "오류"
                        elif 'NO OPENAPI SERVICE ERROR' in req_html.upper():
                            s4['I' + str(8 + i)].value = "오류"
                        elif 'SERVICE ACCESS DENIED ERROR' in req_html.upper():
                            s4['I' + str(8 + i)].value = "오류"
                        elif 'LIMITED NUMBER OF SERVICE REQUESTS EXCEEDS ERROR' in req_html.upper():
                            s4['I' + str(8 + i)].value = "오류"
                        elif 'SERVICE KEY IS NOT REGISTERED ERROR' in req_html.upper():
                            s4['I' + str(8 + i)].value = "오류"
                        elif 'DEADLINE HAS EXPIRED ERROR' in req_html.upper():
                            s4['I' + str(8 + i)].value = "오류"
                        elif 'UNREGISTERED IP ERROR' in req_html.upper():
                            s4['I' + str(8 + i)].value = "오류"
                        elif 'INVALID REQUEST PARAMETER ERROR' in req_html.upper():
                            s4['I' + str(8 + i)].value = "오류"
                        else:
                            s4['I' + str(8 + i)].value = "오류"

                        exel_font_set(s4['I' + str(8 + i)])
                        result_list2.append(s4['I' + str(8 + i)].value)

                        ## 3번 항목
                        s5 = wb_result['30.S-3']
                        result3_temp = []

                        for j in range(0, num1_cnt1):
                            s5['A' + str(8 + j)].value = s3['A' + str(8 + j)].value
                            exel_font_set(s5['A' + str(8 + j)])
                            s5['B' + str(8 + j)].value = s3['B' + str(8 + j)].value
                            exel_font_set(s5['B' + str(8 + j)])
                            s5['C' + str(8 + j)].value = s3['C' + str(8 + j)].value
                            exel_font_set(s5['C' + str(8 + j)])
                            s5['D' + str(8 + j)].value = s3['D' + str(8 + j)].value
                            exel_font_set(s5['D' + str(8 + j)])
                            s5['E' + str(8 + j)].value = s3['E' + str(8 + j)].value
                            exel_font_set(s5['E' + str(8 + j)])
                            s5['F' + str(8 + j)].value = s3['F' + str(8 + j)].value
                            exel_font_set(s5['F' + str(8 + j)])
                            s5['G' + str(8 + j)].value = s3['G' + str(8 + j)].value
                            exel_font_set(s5['G' + str(8 + j)])
                            s5['H' + str(8 + j)].value = s3['H' + str(8 + j)].value
                            exel_font_set(s5['H' + str(8 + j)])
                            s5['I' + str(8 + j)].value = s3['I' + str(8 + j)].value
                            exel_font_set(s5['I' + str(8 + j)])
                            s5['J' + str(8 + j)].value = s3['J' + str(8 + j)].value
                            exel_font_set(s5['J' + str(8 + j)])
                            s5['K' + str(8 + j)].value = s3['K' + str(8 + j)].value
                            exel_font_set(s5['K' + str(8 + j)])
                            s5['L' + str(8 + j)].value = s3['L' + str(8 + j)].value
                            exel_font_set(s5['L' + str(8 + j)])
                            if str(s5['I' + str(8 + j)].value) == '동일항목없음' or str(
                                    s5['C' + str(8 + j)].value) == '동일항목없음':
                                s5['M' + str(8 + j)].value = '오류'
                                result3_temp.append("오류")
                            elif str(s5['F' + str(8 + j)].value).startswith('0') and str(
                                    s5['L' + str(8 + j)].value).startswith('옵'):
                                s5['M' + str(8 + j)].value = '정상'
                            elif str(s5['F' + str(8 + j)].value).startswith('0') and str(
                                    s5['L' + str(8 + j)].value) == '':
                                s5['M' + str(8 + j)].value = '정상'
                            elif str(s5['F' + str(8 + j)].value) == '선택' and str(
                                    s5['L' + str(8 + j)].value).startswith('옵'):
                                s5['M' + str(8 + j)].value = '정상'
                            elif '0' in str(s5['F' + str(8 + j)].value) and str(
                                    s5['L' + str(8 + j)].value) == '':
                                s5['M' + str(8 + j)].value = '정상'
                            elif str(s5['F' + str(8 + j)].value) == '선택' and str(
                                    s5['L' + str(8 + j)].value) == '':
                                s5['M' + str(8 + j)].value = '정상'
                            elif '0' in str(s5['F' + str(8 + j)].value) and str(
                                    s5['L' + str(8 + j)].value) == 'nan':
                                s5['M' + str(8 + j)].value = '정상'
                            elif str(s5['F' + str(8 + j)].value) == 'nan' and str(
                                    s5['L' + str(8 + j)].value).startswith('옵'):
                                s5['M' + str(8 + j)].value = '정상'
                            elif str(s5['F' + str(8 + j)].value) == '선택' and str(
                                    s5['L' + str(8 + j)].value) == 'nan':
                                s5['M' + str(8 + j)].value = '정상'
                            elif str(s5['F' + str(8 + j)].value) == str(
                                    s5['L' + str(8 + j)].value):
                                s5['M' + str(8 + j)].value = '정상'
                            elif str(s5['F' + str(8 + j)].value).startswith('1') and str(
                                    s5['L' + str(8 + j)].value).startswith('1'):
                                s5['M' + str(8 + j)].value = '정상'
                            elif str(s5['F' + str(8 + j)].value).startswith('0') and str(
                                    s5['L' + str(8 + j)].value).startswith('0'):
                                s5['M' + str(8 + j)].value = '정상'
                            elif str(s5['F' + str(8 + j)].value).startswith('0') and str(
                                    s5['L' + str(8 + j)].value).startswith('필'):
                                s5['M' + str(8 + j)].value = '오류'
                                result3_temp.append("오류")
                            elif str(s5['F' + str(8 + j)].value).startswith('0') and str(
                                    s5['L' + str(8 + j)].value) == 'nan':
                                s5['M' + str(8 + j)].value = '정상'
                            elif str(s5['F' + str(8 + j)].value).startswith('1') and str(
                                    s5['L' + str(8 + j)].value).startswith('필'):
                                s5['M' + str(8 + j)].value = '정상'
                            else:
                                s5['M' + str(8 + j)].value = '오류'
                                result3_temp.append("오류")
                            exel_font_set(s5['M' + str(8 + j)])

                        if '오류' in result3_temp:
                            result_list3.append('오류')

                        ## 4번 항목
                        ## 응답항목 기준
                        s6 = wb_result['40.S-4']
                        if len(res_name_index) == 0 or str(df_data.iloc[0, 2]) == 'nan':
                            s6['A' + str(8 + s6cnt)].value = str(8 + s6cnt - 7)
                            exel_font_set(s6['A' + str(8 + s6cnt)])
                            s6['B' + str(8 + s6cnt)].value = opername
                            exel_font_set(s6['B' + str(8 + s6cnt)])
                            s6['C' + str(8 + s6cnt)].value = ''
                            exel_font_set(s6['C' + str(8 + s6cnt)])
                            s6['D' + str(8 + s6cnt)].value = str(8 + s6cnt - 7)
                            exel_font_set(s6['D' + str(8 + s6cnt)])
                            s6['E' + str(8 + s6cnt)].value = opername
                            exel_font_set(s6['E' + str(8 + s6cnt)])
                            s6['F' + str(8 + s6cnt)].value = req_html[:1000]
                            if req_html == "응답없음":
                                s6['G' + str(8 + s6cnt)].value = '오류'
                            elif '이미지' in req_html:
                                s6['G' + str(8 + s6cnt)].value = '정상'
                            elif 'normal' in req_html.lower():
                                s6['G' + str(8 + s6cnt)].value = '정상'
                            elif 'nomal' in req_html.lower():
                                s6['G' + str(8 + s6cnt)].value = '정상'
                            elif '<successYN>Y</successYN>' in req_html.lower():
                                s6['G' + str(8 + s6cnt)].value = '정상'
                            elif 'wfs' in req_html.lower():
                                s6['G' + str(8 + s6cnt)].value = '정상'
                            elif '<resultcode>00' in req_html.lower():
                                s6['G' + str(8 + s6cnt)].value = '정상'
                            elif 'soapenv' in req_html.lower():
                                s6['G' + str(8 + s6cnt)].value = "오류"
                            elif '<successYN>N</successYN>' in req_html.lower():
                                s6['G' + str(8 + s6cnt)].value = "오류"
                            elif 'errorcode' in req_html.lower():
                                s6['G' + str(8 + s6cnt)].value = "오류"
                            elif 'service error' in req_html.lower():
                                s6['G' + str(8 + s6cnt)].value = "오류"
                            elif 'SERVICE KEY IS NOT' in req_html.upper():
                                s6['G' + str(8 + s6cnt)].value = "오류"
                            elif 'APPLICATION ERROR' in req_html.upper():
                                s6['G' + str(8 + s6cnt)].value = "오류"
                            elif 'DB ERROR' in req_html.upper():
                                s6['G' + str(8 + s6cnt)].value = "오류"
                            elif 'NODATA ERROR' in req_html.upper():
                                s6['G' + str(8 + s6cnt)].value = "오류"
                            elif 'HTTP ERROR' in req_html.upper():
                                s6['G' + str(8 + s6cnt)].value = "오류"
                            elif 'SERVICETIMEOUT ERROR' in req_html.upper():
                                s6['G' + str(8 + s6cnt)].value = "오류"
                            elif 'INVALID_REQUEST PARAMETER_ERROR' in req_html.upper():
                                s6['G' + str(8 + s6cnt)].value = "오류"
                            elif 'NO MANDATORY REQUEST PARAMETERS ERROR' in req_html.upper():
                                s6['G' + str(8 + s6cnt)].value = "오류"
                            elif 'NO OPENAPI SERVICE ERROR' in req_html.upper():
                                s6['G' + str(8 + s6cnt)].value = "오류"
                            elif 'SERVICE ACCESS DENIED ERROR' in req_html.upper():
                                s6['G' + str(8 + s6cnt)].value = "오류"
                            elif 'LIMITED NUMBER OF SERVICE REQUESTS EXCEEDS ERROR' in req_html.upper():
                                s6['G' + str(8 + s6cnt)].value = "오류"
                            elif 'SERVICE KEY IS NOT REGISTERED ERROR' in req_html.upper():
                                s6['G' + str(8 + s6cnt)].value = "오류"
                            elif 'DEADLINE HAS EXPIRED ERROR' in req_html.upper():
                                s6['G' + str(8 + s6cnt)].value = "오류"
                            elif 'UNREGISTERED IP ERROR' in req_html.upper():
                                s6['G' + str(8 + s6cnt)].value = "오류"
                            elif 'INVALID REQUEST PARAMETER ERROR' in req_html.upper():
                                s6['G' + str(8 + s6cnt)].value = "오류"
                            elif len(set(res_eng_name_list)) == 0:
                                pass
                            elif len(set(res_eng_name_list) - set(real_tagnames)) / len(set(res_eng_name_list)) < 0.7:
                                s6['G' + str(8 + s6cnt)].value = '정상'
                            else:
                                s6['G' + str(8 + s6cnt)].value = '오류'
                            exel_font_set(s6['F' + str(8 + s6cnt)])
                            exel_font_set(s6['G' + str(8 + s6cnt)])
                            result_list4.append(s6['G' + str(8 + s6cnt)].value)
                            s6cnt = s6cnt + 1
                        else:
                            s6.merge_cells("D" + str(8 + s6cnt) + ":" + 'D' + str(
                                8 + s6cnt + len(res_name_index) - 1))
                            s6.merge_cells("E" + str(8 + s6cnt) + ":" + 'E' + str(
                                8 + len(res_name_index) + s6cnt - 1))
                            s6.merge_cells("F" + str(8 + s6cnt) + ":" + 'F' + str(
                                8 + len(res_name_index) + s6cnt - 1))
                            s6.merge_cells("G" + str(8 + s6cnt) + ":" + 'G' + str(
                                8 + len(res_name_index) + s6cnt - 1))
                            style_range(s6, "D" + str(8 + s6cnt) + ":" + 'D' + str(
                                8 + s6cnt + len(res_name_index) - 1),
                                                    border=Border(left=Side(style='thin'), right=Side(style='thin'),
                                                                  top=Side(style='thin'),
                                                                  bottom=Side(style='thin')))
                            style_range(s6, "E" + str(8 + s6cnt) + ":" + 'E' + str(
                                8 + s6cnt + len(res_name_index) - 1),
                                                    border=Border(left=Side(style='thin'), right=Side(style='thin'),
                                                                  top=Side(style='thin'),
                                                                  bottom=Side(style='thin')))
                            style_range(s6, "F" + str(8 + s6cnt) + ":" + 'F' + str(
                                8 + s6cnt + len(res_name_index) - 1),
                                                    border=Border(left=Side(style='thin'), right=Side(style='thin'),
                                                                  top=Side(style='thin'),
                                                                  bottom=Side(style='thin')))
                            style_range(s6, "G" + str(8 + s6cnt) + ":" + 'G' + str(
                                8 + s6cnt + len(res_name_index) - 1),
                                                    border=Border(left=Side(style='thin'), right=Side(style='thin'),
                                                                  top=Side(style='thin'),
                                                                  bottom=Side(style='thin')))
                            s6['D' + str(8 + s6cnt)].value = str(i + 1)
                            exel_font_set(s6['D' + str(8 + s6cnt)])
                            s6['E' + str(8 + s6cnt)].value = opername
                            exel_font_set(s6['E' + str(8 + s6cnt)])
                            s6['F' + str(8 + s6cnt)].value = req_html[:1000]
                            exel_font_set(s6['F' + str(8 + s6cnt)])
                            res_eng_name_list = []
                            s6cnt2 = s6cnt
                            for j in range(0, len(res_name_index)):
                                s6['A' + str(8 + s6cnt)].value = str(j + 1 + s6cnt)
                                exel_font_set(s6['A' + str(8 + s6cnt)])
                                s6['B' + str(8 + s6cnt)].value = opername
                                exel_font_set(s6['B' + str(8 + s6cnt)])
                                s6['C' + str(8 + s6cnt)].value = df_data.iloc[j, 9]
                                exel_font_set(s6['C' + str(8 + s6cnt)])
                                res_eng_name_list.append(str(df_data.iloc[j, 9]))
                                s6cnt = s6cnt + 1

                            if req_html == "응답없음":
                                s6['G' + str(8 + s6cnt2)].value = '오류'
                            elif '이미지' in req_html:
                                s6['G' + str(8 + s6cnt2)].value = '정상'
                            elif 'normal' in req_html.lower():
                                s6['G' + str(8 + s6cnt2)].value = '정상'
                            elif 'nomal' in req_html.lower():
                                s6['G' + str(8 + s6cnt2)].value = '정상'
                            elif '<successYN>Y</successYN>' in req_html.lower():
                                s6['G' + str(8 + s6cnt2)].value = '정상'
                            elif '<resultcode>00' in req_html.lower():
                                s6['G' + str(8 + s6cnt2)].value = '정상'
                            elif 'wfs' in req_html.lower():
                                s6['G' + str(8 + s6cnt2)].value = '정상'
                            elif 'soapenv' in req_html.lower():
                                s6['G' + str(8 + s6cnt2)].value = "오류"
                            elif '<successYN>N</successYN>' in req_html.lower():
                                s6['G' + str(8 + s6cnt2)].value = "오류"
                            elif 'errorcode' in req_html.lower():
                                s6['G' + str(8 + s6cnt2)].value = "오류"
                            elif 'service error' in req_html.lower():
                                s6['G' + str(8 + s6cnt2)].value = "오류"
                            elif 'SERVICE KEY IS NOT' in req_html.upper():
                                s6['G' + str(8 + s6cnt2)].value = "오류"
                            elif 'APPLICATION ERROR' in req_html.upper():
                                s6['G' + str(8 + s6cnt2)].value = "오류"
                            elif 'DB ERROR' in req_html.upper():
                                s6['G' + str(8 + s6cnt2)].value = "오류"
                            elif 'NODATA ERROR' in req_html.upper():
                                s6['G' + str(8 + s6cnt2)].value = "오류"
                            elif 'HTTP ERROR' in req_html.upper():
                                s6['G' + str(8 + s6cnt2)].value = "오류"
                            elif 'SERVICETIMEOUT ERROR' in req_html.upper():
                                s6['G' + str(8 + s6cnt2)].value = "오류"
                            elif 'INVALID_REQUEST PARAMETER_ERROR' in req_html.upper():
                                s6['G' + str(8 + s6cnt2)].value = "오류"
                            elif 'NO MANDATORY REQUEST PARAMETERS ERROR' in req_html.upper():
                                s6['G' + str(8 + s6cnt2)].value = "오류"
                            elif 'NO OPENAPI SERVICE ERROR' in req_html.upper():
                                s6['G' + str(8 + s6cnt2)].value = "오류"
                            elif 'SERVICE ACCESS DENIED ERROR' in req_html.upper():
                                s6['G' + str(8 + s6cnt2)].value = "오류"
                            elif 'LIMITED NUMBER OF SERVICE REQUESTS EXCEEDS ERROR' in req_html.upper():
                                s6['G' + str(8 + s6cnt2)].value = "오류"
                            elif 'SERVICE KEY IS NOT REGISTERED ERROR' in req_html.upper():
                                s6['G' + str(8 + s6cnt2)].value = "오류"
                            elif 'DEADLINE HAS EXPIRED ERROR' in req_html.upper():
                                s6['G' + str(8 + s6cnt2)].value = "오류"
                            elif 'UNREGISTERED IP ERROR' in req_html.upper():
                                s6['G' + str(8 + s6cnt2)].value = "오류"
                            elif 'INVALID REQUEST PARAMETER ERROR' in req_html.upper():
                                s6['G' + str(8 + s6cnt2)].value = "오류"
                            elif len(set(res_eng_name_list)) == 0:
                                pass
                            elif len(set(res_eng_name_list) - set(real_tagnames)) / len(set(res_eng_name_list)) < 0.7:
                                s6['G' + str(8 + s6cnt2)].value = '정상'
                            else:
                                s6['G' + str(8 + s6cnt2)].value = '오류'
                            exel_font_set(s6['G' + str(8 + s6cnt2)])
                            result_list4.append(s6['G' + str(8 + s6cnt2)].value)

                        ## 5번 항목
                        ## 응답항목 기준
                        s7 = wb_result['50.S-5']
                        if len(portal_res_name_index) == 0 or str(portal_res_df.iloc[0, 1]) == 'nan':
                            s7['A' + str(8 + s7cnt)].value = str(8 + s7cnt - 7)
                            exel_font_set(s7['A' + str(8 + s7cnt)])
                            s7['B' + str(8 + s7cnt)].value = opername
                            exel_font_set(s7['B' + str(8 + s7cnt)])
                            s7['C' + str(8 + s7cnt)].value = ''
                            exel_font_set(s7['C' + str(8 + s7cnt)])
                            s7['D' + str(8 + s7cnt)].value = str(8 + s7cnt - 7)
                            exel_font_set(s7['D' + str(8 + s7cnt)])
                            s7['E' + str(8 + s7cnt)].value = opername
                            exel_font_set(s7['E' + str(8 + s7cnt)])
                            s7['F' + str(8 + s7cnt)].value = ''
                            exel_font_set(s7['F' + str(8 + s7cnt)])
                            portal_engname_list = []

                            for portal_engname in portal_res_df['항목명(영문)']:
                                portal_engname_list.append(portal_engname)
                            if req_html == "응답없음":
                                s7['G' + str(8 + s7cnt)].value = '오류'
                            elif '이미지' in req_html:
                                s7['G' + str(8 + s7cnt)].value = '정상'
                            elif 'normal' in req_html.lower():
                                s7['G' + str(8 + s7cnt)].value = '정상'
                            elif 'nomal' in req_html.lower():
                                s7['G' + str(8 + s7cnt)].value = '정상'
                            elif '<successYN>Y</successYN>' in req_html.lower():
                                s7['G' + str(8 + s7cnt)].value = '정상'
                            elif '<resultcode>00' in req_html.lower():
                                s7['G' + str(8 + s7cnt)].value = '정상'
                            elif 'wfs' in req_html.lower():
                                s7['G' + str(8 + s7cnt)].value = '정상'
                            elif 'soapenv' in req_html.lower():
                                s7['G' + str(8 + s7cnt)].value = "오류"
                            elif 'errorcode' in req_html.lower():
                                s7['G' + str(8 + s7cnt)].value = "오류"
                            elif 'service error' in req_html.lower():
                                s7['G' + str(8 + s7cnt)].value = "오류"
                            elif 'SERVICE KEY IS NOT' in req_html.upper():
                                s7['G' + str(8 + s7cnt)].value = "오류"
                            elif '<successYN>N</successYN>' in req_html.lower():
                                s7['G' + str(8 + s7cnt)].value = "오류"
                            elif 'APPLICATION ERROR' in req_html.upper():
                                s7['G' + str(8 + s7cnt)].value = "오류"
                            elif 'DB ERROR' in req_html.upper():
                                s7['G' + str(8 + s7cnt)].value = "오류"
                            elif 'NODATA ERROR' in req_html.upper():
                                s7['G' + str(8 + s7cnt)].value = "오류"
                            elif 'HTTP ERROR' in req_html.upper():
                                s7['G' + str(8 + s7cnt)].value = "오류"
                            elif 'SERVICETIMEOUT ERROR' in req_html.upper():
                                s7['G' + str(8 + s7cnt)].value = "오류"
                            elif 'INVALID_REQUEST PARAMETER_ERROR' in req_html.upper():
                                s7['G' + str(8 + s7cnt)].value = "오류"
                            elif 'NO MANDATORY REQUEST PARAMETERS ERROR' in req_html.upper():
                                s7['G' + str(8 + s7cnt)].value = "오류"
                            elif 'NO OPENAPI SERVICE ERROR' in req_html.upper():
                                s7['G' + str(8 + s7cnt)].value = "오류"
                            elif 'SERVICE ACCESS DENIED ERROR' in req_html.upper():
                                s7['G' + str(8 + s7cnt)].value = "오류"
                            elif 'LIMITED NUMBER OF SERVICE REQUESTS EXCEEDS ERROR' in req_html.upper():
                                s7['G' + str(8 + s7cnt)].value = "오류"
                            elif 'SERVICE KEY IS NOT REGISTERED ERROR' in req_html.upper():
                                s7['G' + str(8 + s7cnt)].value = "오류"
                            elif 'DEADLINE HAS EXPIRED ERROR' in req_html.upper():
                                s7['G' + str(8 + s7cnt)].value = "오류"
                            elif 'UNREGISTERED IP ERROR' in req_html.upper():
                                s7['G' + str(8 + s7cnt)].value = "오류"
                            elif 'INVALID REQUEST PARAMETER ERROR' in req_html.upper():
                                s7['G' + str(8 + s7cnt)].value = "오류"
                            elif len(set(portal_engname_list)) == 0:
                                pass
                            elif len(set(portal_engname_list) - set(real_tagnames)) / len(
                                    set(portal_engname_list)) < 0.7:
                                s7['G' + str(8 + s7cnt)].value = '정상'
                            else:
                                s7['G' + str(8 + s7cnt)].value = '오류'
                            exel_font_set(s7['G' + str(8 + s7cnt)])
                            result_list5.append(s7['G' + str(8 + s7cnt)].value)
                            s7cnt = s7cnt + 1
                        else:
                            s7.merge_cells("D" + str(8 + s7cnt) + ":" + 'D' + str(
                                8 + len(portal_res_name_index) + s7cnt - 1))
                            s7.merge_cells("E" + str(8 + s7cnt) + ":" + 'E' + str(
                                8 + len(portal_res_name_index) + s7cnt - 1))
                            s7.merge_cells("F" + str(8 + s7cnt) + ":" + 'F' + str(
                                8 + len(portal_res_name_index) + s7cnt - 1))
                            s7.merge_cells("G" + str(8 + s7cnt) + ":" + 'G' + str(
                                8 + len(portal_res_name_index) + s7cnt - 1))
                            style_range(s7, "D" + str(8 + s7cnt) + ":" + 'D' + str(
                                8 + len(portal_res_name_index) + s7cnt - 1),
                                                    border=Border(left=Side(style='thin'), right=Side(style='thin'),
                                                                  top=Side(style='thin'),
                                                                  bottom=Side(style='thin')))
                            style_range(s7, "E" + str(8 + s7cnt) + ":" + 'E' + str(
                                8 + len(portal_res_name_index) + s7cnt - 1),
                                                    border=Border(left=Side(style='thin'), right=Side(style='thin'),
                                                                  top=Side(style='thin'),
                                                                  bottom=Side(style='thin')))
                            style_range(s7, "F" + str(8 + s7cnt) + ":" + 'F' + str(
                                8 + len(portal_res_name_index) + s7cnt - 1),
                                                    border=Border(left=Side(style='thin'), right=Side(style='thin'),
                                                                  top=Side(style='thin'),
                                                                  bottom=Side(style='thin')))
                            style_range(s7, "G" + str(8 + s7cnt) + ":" + 'G' + str(
                                8 + len(portal_res_name_index) + s7cnt - 1),
                                                    border=Border(left=Side(style='thin'), right=Side(style='thin'),
                                                                  top=Side(style='thin'),
                                                                  bottom=Side(style='thin')))
                            s7cnt2 = s7cnt
                            for j in portal_res_df.index:
                                s7['A' + str(8 + s7cnt)].value = str(
                                    + s7cnt + 1)
                                exel_font_set(s7['A' + str(8 + s7cnt)])
                                s7['B' + str(8 + s7cnt)].value = opername
                                exel_font_set(s7['B' + str(8 + s7cnt)])
                                s7['C' + str(8 + s7cnt)].value = portal_res_df.iloc[j, 1]
                                exel_font_set(s7['C' + str(8 + s7cnt)])
                                s7cnt = s7cnt + 1

                            s7['D' + str(8 + s7cnt2)].value = str(i + 1)
                            exel_font_set(s7['D' + str(8 + s7cnt2)])
                            s7['E' + str(8 + s7cnt2)].value = opername
                            exel_font_set(s7['E' + str(8 + s7cnt2)])
                            s7['F' + str(8 + s7cnt2)].value = req_html[:1000].replace('\\n', '')
                            exel_font_set(s7['F' + str(8 + s7cnt2)])
                            ##포털 항목명 영문 리스트 작성
                            portal_engname_list = []

                            for portal_engname in portal_res_df['항목명(영문)']:
                                portal_engname_list.append(portal_engname)
                            if req_html == "응답없음":
                                s7['G' + str(8 + s7cnt2)].value = '오류'
                            elif '이미지' in req_html:
                                s7['G' + str(8 + s7cnt2)].value = '정상'
                            elif 'normal' in req_html.lower():
                                s7['G' + str(8 + s7cnt2)].value = '정상'
                            elif 'nomal' in req_html.lower():
                                s7['G' + str(8 + s7cnt2)].value = '정상'
                            elif '<successYN>Y</successYN>' in req_html.lower():
                                s7['G' + str(8 + s7cnt2)].value = '정상'
                            elif '<resultcode>00' in req_html.lower():
                                s7['G' + str(8 + s7cnt2)].value = '정상'
                            elif 'wfs' in req_html.lower():
                                s7['G' + str(8 + s7cnt2)].value = '정상'
                            elif 'soapenv' in req_html.lower():
                                s7['G' + str(8 + s7cnt2)].value = "오류"
                            elif 'errorcode' in req_html.lower():
                                s7['G' + str(8 + s7cnt2)].value = "오류"
                            elif 'service error' in req_html.lower():
                                s7['G' + str(8 + s7cnt2)].value = "오류"
                            elif '<successYN>N</successYN>' in req_html.lower():
                                s7['G' + str(8 + s7cnt2)].value = "오류"
                            elif 'SERVICE KEY IS NOT' in req_html.upper():
                                s7['G' + str(8 + s7cnt2)].value = "오류"
                            elif 'APPLICATION ERROR' in req_html.upper():
                                s7['G' + str(8 + s7cnt2)].value = "오류"
                            elif 'DB ERROR' in req_html.upper():
                                s7['G' + str(8 + s7cnt2)].value = "오류"
                            elif 'NODATA ERROR' in req_html.upper():
                                s7['G' + str(8 + s7cnt2)].value = "오류"
                            elif 'HTTP ERROR' in req_html.upper():
                                s7['G' + str(8 + s7cnt2)].value = "오류"
                            elif 'SERVICETIMEOUT ERROR' in req_html.upper():
                                s7['G' + str(8 + s7cnt2)].value = "오류"
                            elif 'INVALID_REQUEST PARAMETER_ERROR' in req_html.upper():
                                s7['G' + str(8 + s7cnt2)].value = "오류"
                            elif 'NO MANDATORY REQUEST PARAMETERS ERROR' in req_html.upper():
                                s7['G' + str(8 + s7cnt2)].value = "오류"
                            elif 'NO OPENAPI SERVICE ERROR' in req_html.upper():
                                s7['G' + str(8 + s7cnt2)].value = "오류"
                            elif 'SERVICE ACCESS DENIED ERROR' in req_html.upper():
                                s7['G' + str(8 + s7cnt2)].value = "오류"
                            elif 'LIMITED NUMBER OF SERVICE REQUESTS EXCEEDS ERROR' in req_html.upper():
                                s7['G' + str(8 + s7cnt2)].value = "오류"
                            elif 'SERVICE KEY IS NOT REGISTERED ERROR' in req_html.upper():
                                s7['G' + str(8 + s7cnt2)].value = "오류"
                            elif 'DEADLINE HAS EXPIRED ERROR' in req_html.upper():
                                s7['G' + str(8 + s7cnt2)].value = "오류"
                            elif 'UNREGISTERED IP ERROR' in req_html.upper():
                                s7['G' + str(8 + s7cnt2)].value = "오류"
                            elif 'INVALID REQUEST PARAMETER ERROR' in req_html.upper():
                                s7['G' + str(8 + s7cnt2)].value = "오류"
                            elif len(set(portal_engname_list)) == 0:
                                pass
                            elif len(set(portal_engname_list) - set(real_tagnames)) / len(
                                    set(portal_engname_list)) < 0.7:
                                s7['G' + str(8 + s7cnt2)].value = '정상'
                            else:
                                s7['G' + str(8 + s7cnt2)].value = '오류'
                            exel_font_set(s7['G' + str(8 + s7cnt2)])
                            result_list5.append(s7['G' + str(8 + s7cnt2)].value)


                    if not df_data.iloc[0, 0].strip() == '문서오류':

                        s10cnt = 0
                        ##8번 항목
                        s10 = wb_result['80.S-8']
                        for chek in range(0, operation_len):
                            s10cnt = s10cnt + 1
                            s10['A' + str(9 + chek)].value = s10cnt
                            s10['B' + str(9 + chek)].value = opername_list[chek]
                            s10['C' + str(9 + chek)].value = result_list1[chek]
                            s10['D' + str(9 + chek)].value = result_list2[chek]
                            s10['E' + str(9 + chek)].value = result_list4[chek]
                            if s10['C' + str(9 + chek)].value == '오류' or s10['D' + str(9 + chek)].value == '오류' or s10[
                                'E' + str(9 + chek)].value == '오류':
                                s10['F' + str(9 + chek)].value = '오류'
                            else:
                                s10['F' + str(9 + chek)].value = '정상'

                            exel_font_set(s10['A' + str(9 + chek)])
                            exel_font_set(s10['B' + str(9 + chek)])
                            exel_font_set(s10['C' + str(9 + chek)])
                            exel_font_set(s10['D' + str(9 + chek)])
                            exel_font_set(s10['E' + str(9 + chek)])
                            exel_font_set(s10['F' + str(9 + chek)])
                            result_list8.append(s10['F' + str(9 + chek)].value)

                        ##     # s['A'+str(8 + j)]
                        ##수준평가
                        s1 = wb_result['00.수준평가']
                        s1['C5'].value = html_meta_df.iloc[0, 3]
                        s1['C6'].value = api_name
                        s1['C7'].value = html_meta_df.iloc[2, 1]
                        s1['C8'].value = html_meta_df.iloc[2, 3]
                        s1['C9'].value = len(sheet_name_list)
                        s1['C10'].value = html_meta_df.iloc[8, 1]
                        s1['C11'].value = str(ws_data['L2'].value)
                        s1['D14'].value = operation_len
                        s1['D15'].value = operation_len
                        s1['D16'].value = operation_len
                        s1['D17'].value = operation_len
                        s1['D18'].value = operation_len

                        s1['E14'].value = operation_len - result_list1.count('오류')
                        s1['F15'].value = operation_len - result_list2.count('오류')
                        s1['E16'].value = operation_len - result_list3.count('오류')
                        s1['F17'].value = operation_len - result_list4.count('오류')
                        s1['E18'].value = operation_len
                        s1['E18'].value = operation_len - result_list5.count('오류')



                        s2 = wb_result['02.추가 진단']
                        s2['C5'].value = html_meta_df.iloc[0, 3]
                        s2['C6'].value = api_name
                        s2['C7'].value = html_meta_df.iloc[2, 1]
                        s2['C8'].value = html_meta_df.iloc[2, 3]
                        s2['C9'].value = len(sheet_name_list)
                        s2['C10'].value = html_meta_df.iloc[8, 1]
                        s2['C11'].value = str(ws_data['L2'].value)



                        s2['D15'].value = operation_len
                        s2['F15'].value = operation_len - result_list8.count('오류')

                        shutil.move(path, './complete/' + fname)
                    wb_result.save(result)
                    try:
                        driver.close()
                    except:
                        pass
                    self.label2.setText("보고서 생성 완료")
                    self.label2.repaint()

            except Exception as e:
                traceback.print_exc()
                if str(e).startswith('Message: Could not locate element with visible text:'):
                    self.messagebox_open('등록된 오퍼레이션 "'+str(e).split(':')[-1].strip()+ '"은 포털 오퍼레이션에 존재하지 않습니다.')
                elif str(e).startswith("Invalid URL 'nan':"):
                    self.messagebox_open('입력 리소스 "'+fname+'" 포털URL에 빈칸인 부분이 있습니다. 확인해주세요. (포털 URL입력 필수)')
                elif "Permission" in str(e):
                    self.messagebox_open('리소스 혹은 보고서 엑셀 파일이 사용중입니다. 해당 파일을 종료한 뒤 진단시작을 해주세요.')
                elif str(e).startswith("'numpy.float64' object has no"):
                    self.messagebox_open('입력 리소스 "'+fname+'" 오퍼레이션에 빈칸인 부분이 있습니다. 확인해주세요.(오퍼레이션 입력 필수)')
                elif str(e).startswith('Message: no such element:'):
                    self.messagebox_open('포털과의 연결상태가 좋지 않습니다. 공공데이터 포털 페이지를 확인해주세요.')
                else:
                    self.messagebox_open(e)
                self.label2.setText("보고서 생성 에러")
                self.label2.repaint()

                pass

    def messagebox_open(self,e):
        msgBox = QMessageBox()
        msgBox.setIcon(QMessageBox.Information)
        msgBox.setText(str(e))
        msgBox.setWindowTitle("파일문제")
        msgBox.setStandardButtons(QMessageBox.Ok)
        returnValue = msgBox.exec()



if __name__ == '__main__':
    app = QApplication(sys.argv)
    ex = MyApp()
    sys.exit(app.exec_())




