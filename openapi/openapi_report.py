import shutil
import traceback
from openpyxl.styles import Border, Side
import re
from difflib import SequenceMatcher
import requests
from bs4 import BeautifulSoup
from urllib import parse
from urllib import request
from urllib.error import HTTPError
import ssl
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
import module.module as md
from module.module import style_range, exel_font_set
import openpyxl
import sys, os
from PyQt5.QtWidgets import *
from PyQt5.QtCore import *
import pandas as pd


class Thread1(QThread):
    def __init__(self, parent):
        super().__init__(parent)

    def run(self):
        try:
            ##csv데이터
            api_table = pd.read_csv('./data/API.csv', dtype={"APICD": str})
            operation_req_table = pd.read_csv('./data/OPERATION_REQ.csv', dtype={"APICD": str, "OPERCD": str})
            operation_res_table = pd.read_csv('./data/OPERATION_RES.csv', dtype={"APICD": str, "OPERCD": str})

            ##변수
            servicekey_tail = 'WBaXX3pce9C9AKfYTQc5%2FXVYPXYJWfHVzWNaird%2Fv0f8C0zKhPFhjY10Tuf2QuiA83hfkGLzHknlOz5FWPbaDQ%3D%3D'
            fname_list_before = os.listdir('./resource/')
            fname_list = []
            result_path = './report/'
            path = './resource/'

            ##파일이 정상적인 엑셀 파일인지 체크
            fname_list = md.filechk(self, fname_list, fname_list_before)


            for fname in fname_list:
                print(fname)
                self.parent().label.setText(fname)
                self.parent().label.repaint()

                ##변수
                data = path + fname
                wb_data = openpyxl.load_workbook(data, data_only=True)
                sheet_name_list = wb_data.sheetnames
                ws_data_firstsheet = wb_data[sheet_name_list[0]]
                apiid = re.sub(r'[^0-9]', '', str(ws_data_firstsheet['L2'].value))
                s6cnt = 0
                s7cnt = 0
                opername_list = []
                sheet_name_list = wb_data.sheetnames
                temp_req = []
                temp_res = []
                num1_cnt1 = 0
                portal_temp_res = []
                result_list1 = []
                result_list2 = []
                result_list3 = []
                result_list4 = []
                result_list5 = []
                result_list6 = []
                result_list7 = []
                result_list8 = []
                operation_len = len(sheet_name_list)
                result_file = api_name.replace('/','').replace("\r",'').replace(":",'').replace("*",'')\
                              .replace("?",'').replace('"','').replace("<",'').replace(">",'').replace("|",'').replace('\t','').replace("\\",'').replace("\n","")\
                              +"_"+str(int(apiid))+"_" + '오픈API_진단결과보고서.xlsx'
                templet = './data/보고서_템플릿.xlsx'


                try:
                    apilist_df = api_table[api_table['APICD'] == str(int(apiid))]
                    dataformmat = apilist_df.iloc[0]['DATAFORMMAT']
                    docxname = apilist_df.iloc[0]['DOCXNAME']
                    api_name = apilist_df.iloc[0]['APINM']
                    api_type = apilist_df.iloc[0]['APITYPE']
                    organ_name = apilist_df.iloc[0]['ORGANNM']
                    api_portal_url = 'https://www.data.go.kr/data/' + apiid + '/openapi.do'
                except Exception as e:
                    traceback.print_exc()
                    self.parent().label2.setText("보고서 생성 에러: 리소스의  포털URL이 제대로 입력되었는지 확인해주세요.")
                    self.parent().label2.repaint()
                    return None

                self.parent().label2.setText("보고서 생성 중")
                self.parent().label2.repaint()

                ##오퍼레이션 별로 시트를 나누기에 시트 수로 오퍼레이션 나누기
                for i in range(0, len(sheet_name_list)):

                    oper_req_df = operation_req_table[operation_req_table['APICD'] == str(int(apiid))]

                    print(i)
                    ##시트 정하기
                    sheet = sheet_name_list[i]
                    df_data = pd.read_excel(data, engine="openpyxl", sheet_name=sheet)
                    ws_data = wb_data[sheet]
                    md.mergechk(self, sheet, ws_data, result_path, result_file)

                    md.oper_code_name_setter(self,df_data,opername_list,oper_req_df)
                    if str(type(df_data.iloc[0, 0])) == "<class 'str'>":
                        opername = df_data.iloc[0, 0]
                        opername_list.append(opername)

                    else:
                        opercode = int(df_data.iloc[0, 0])
                        opername = ''

                    try:
                        if opername == '':
                            opername = oper_req_df[oper_req_df['OPERCD'] == str(int(opercode))].iloc[0]['OPERNM']
                            opername_list.append(opername)
                        else:
                            opercode = oper_req_df[oper_req_df['OPERNM'] == opername].iloc[0]['OPERCD']
                        is_opername = oper_req_df['OPERNM'] == opername
                        same_oper_df = oper_req_df[is_opername]
                        if len(same_oper_df['OPERNM'].tolist()) < 1:
                            print(opername)
                            exit()
                    except Exception as e:
                            traceback.print_exc()
                            self.parent().label2.setText("보고서 생성 에러: 리소스의 오퍼레이션명 혹은 포털URL이 제대로 입력되었는지 확인해주세요.")
                            self.parent().label2.repaint()
                            if os.path.isfile(result_path + result_file):
                                os.remove(result_path + result_file)
                            return None

                    search_apicd_res = operation_res_table['APICD'] == str(int(apiid))
                    search_opercd_res = operation_res_table['OPERCD'] == str(int(opercode))
                    oper_req_df = oper_req_df[oper_req_df['OPERCD'] == str(int(opercode))]
                    oper_res_df = operation_res_table[search_apicd_res & search_opercd_res]

                    try:

                        api_req_link = str(df_data.iloc[0, 12]).replace(" ", '').replace("\n", "").replace("\t",
                                                                                                           "").replace(
                            "\r", "")
                        api_req_link_head = api_req_link.split('?')[0]
                        api_req_link_param_all = api_req_link.split('?')[1]
                        api_req_link_param_list = api_req_link_param_all.split('&')
                        changed_param = ''
                        for api_req_param in api_req_link_param_list:
                            if api_req_param.lower().startswith('servicekey'):
                                servicekey_head = api_req_param.split('=')[0]
                                api_req_param = servicekey_head + '=' + servicekey_tail
                            if api_req_param.lower().startswith('authapikey'):
                                servicekey_head = api_req_param.split('=')[0]
                                api_req_param = servicekey_head + '=' + servicekey_tail
                            if changed_param == '':
                                changed_param = changed_param + api_req_param
                            else:
                                changed_param = changed_param + '&' + api_req_param

                        api_req_link = api_req_link_head + '?' + changed_param
                    except:
                        if str(df_data.iloc[0, 12]) == 'nan':
                            api_req_link = '요청메세지 없음'
                        else:
                            api_req_link = str(df_data.iloc[0, 12]).replace(" ", '').replace("\n", "").replace("\t",
                                                                                                               "").replace(
                                "\r", "")

                    try:
                        api_req_link = parse.urlparse(api_req_link)
                        context = ssl._create_unverified_context()
                        query = parse.parse_qs(api_req_link.query)
                        api_req_link = api_req_link_head + '?' + parse.urlencode(query, doseq=True)
                        headers = {'User-Agent':' Mozilla/5.0 (Windows NT 6.1; WOW64; rv:12.0) Gecko/20100101 Firefox/12.0'}
                        api_req = request.Request(api_req_link, headers=headers)

                        try:
                            req_html = request.urlopen(api_req, data=None, context=context)
                        except HTTPError as e:
                            req_html = e
                        req_html = req_html.read().decode("utf-8")
                        req_html = BeautifulSoup(req_html, 'html.parser')
                        req_html = str(req_html)


                    except Exception as e:
                        traceback.print_exc()
                        if str(e).startswith("'utf-8' codec can't decode"):
                            req_html = '이미지'
                        else:
                            req_html = "응답없음"

                    if dataformmat == 'JSON+XML':
                        if req_html == '이미지':
                            req_html2 = '이미지'
                            req_html3 = '이미지'
                        else:
                            jsonurl_main1 = str(api_req_link).split('?')[0]
                            jsonurl_main2 = str(api_req_link).split('?')[0]
                            try:
                                jsonurl_pram_list = str(api_req_link).split('?')[1].split('&')
                            except IndexError:
                                jsonurl_pram_list = []
                            returnparam_list = ['_returnType', '_type', 'act', 'apiType', 'contentType', 'dataFormat',
                                                'dataFormat', 'dataType', 'dateType','jsonVD'
                                , 'format', 'output', 'resultType', 'retunType', 'returnType', 'service_Type', 'type',
                                                'viewType', 'ViewType']
                            selected_returnparam = ''

                            for returnparam in returnparam_list:
                                if returnparam in oper_req_df['COLNMEN'].tolist():
                                    selected_returnparam = returnparam
                                    break

                            for jsonurl_pram_i in range(0, len(jsonurl_pram_list)):
                                if selected_returnparam == '':
                                    if jsonurl_pram_i == 0:
                                        jsonurl_main1 = jsonurl_main1 + '?' + jsonurl_pram_list[jsonurl_pram_i]
                                        jsonurl_main2 = jsonurl_main2 + '?' + jsonurl_pram_list[jsonurl_pram_i]
                                    elif jsonurl_pram_i == len(jsonurl_pram_list) - 1:
                                        jsonurl_main1 = jsonurl_main1 + '&' + jsonurl_pram_list[
                                            jsonurl_pram_i] + '&resultType=xml'
                                        jsonurl_main2 = jsonurl_main2 + '&' + jsonurl_pram_list[
                                            jsonurl_pram_i] + '&resultType=json'
                                    else:
                                        jsonurl_main1 = jsonurl_main1 + '&' + jsonurl_pram_list[jsonurl_pram_i]
                                        jsonurl_main2 = jsonurl_main2 + '&' + jsonurl_pram_list[jsonurl_pram_i]

                                elif selected_returnparam in str(api_req_link):
                                    if jsonurl_pram_list[jsonurl_pram_i].startswith(selected_returnparam):
                                        if jsonurl_pram_i == 0:
                                            if selected_returnparam == 'jsonVD':
                                                jsonurl_main1 = jsonurl_main1
                                                jsonurl_main2 = jsonurl_main2 + '?' + selected_returnparam + '=Y'
                                            else:
                                                jsonurl_main1 = jsonurl_main1 + '?' + selected_returnparam + '=xml'
                                                jsonurl_main2 = jsonurl_main2 + '?' + selected_returnparam + '=json'
                                        else:
                                            if selected_returnparam == 'jsonVD':
                                                jsonurl_main1 = jsonurl_main1
                                                jsonurl_main2 = jsonurl_main2 + '&' + selected_returnparam + '=Y'
                                            else:
                                                jsonurl_main1 = jsonurl_main1 + '&' + selected_returnparam + '=xml'
                                                jsonurl_main2 = jsonurl_main2 + '&' + selected_returnparam + '=json'
                                    else:
                                        if jsonurl_pram_i == 0:
                                            jsonurl_main1 = jsonurl_main1 + '?' + jsonurl_pram_list[jsonurl_pram_i]
                                            jsonurl_main2 = jsonurl_main2 + '?' + jsonurl_pram_list[jsonurl_pram_i]
                                        else:
                                            jsonurl_main1 = jsonurl_main1 + '&' + jsonurl_pram_list[jsonurl_pram_i]
                                            jsonurl_main2 = jsonurl_main2 + '&' + jsonurl_pram_list[jsonurl_pram_i]
                                else:
                                    if jsonurl_pram_i == 0:
                                        jsonurl_main1 = jsonurl_main1 + '?' + jsonurl_pram_list[jsonurl_pram_i]
                                        jsonurl_main2 = jsonurl_main2 + '?' + jsonurl_pram_list[jsonurl_pram_i]

                                    elif jsonurl_pram_i == len(jsonurl_pram_list) - 1:
                                        if selected_returnparam == 'jsonVD':
                                            jsonurl_main1 = jsonurl_main1 + '&' + jsonurl_pram_list[
                                                jsonurl_pram_i]
                                            jsonurl_main2 = jsonurl_main2 + '&' + jsonurl_pram_list[
                                                jsonurl_pram_i] + '&' + selected_returnparam + '=Y'
                                        else:
                                            jsonurl_main1 = jsonurl_main1 + '&' + jsonurl_pram_list[
                                                jsonurl_pram_i] + '&' + selected_returnparam + '=xml'
                                            jsonurl_main2 = jsonurl_main2 + '&' + jsonurl_pram_list[
                                                jsonurl_pram_i] + '&' + selected_returnparam + '=json'

                                    else:
                                        jsonurl_main1 = jsonurl_main1 + '&' + jsonurl_pram_list[jsonurl_pram_i]
                                        jsonurl_main2 = jsonurl_main2 + '&' + jsonurl_pram_list[jsonurl_pram_i]
                            jsonurl_main1 = jsonurl_main1.replace(' ', '').replace('\n', '').replace('\t', '').replace(
                                '\r', '')
                            jsonurl_main2 = jsonurl_main2.replace(' ', '').replace('\n', '').replace('\t', '').replace(
                                '\r', '')
                            print(jsonurl_main1)

                            if req_html == "응답없음":
                                req_html2 = "응답없음"
                            else:
                                try:
                                    print(jsonurl_main1)
                                    req_html2 = requests.get(url=jsonurl_main1, timeout=10).text.strip()


                                except:
                                    req_html2 = "응답없음"

                            if req_html == "응답없음":
                                req_html3 = "응답없음"
                            else:
                                try:
                                    req_html3 = requests.get(url=jsonurl_main2, timeout=10).text.strip()


                                except:
                                    req_html3 = "응답없음"

                    ##모인 데이터로 엑셀 조립하기
                    if i == 0:
                        result = shutil.copy(templet, result_path + result_file)
                        wb_result = openpyxl.load_workbook(result)

                    s3 = wb_result['10.S-1']

                    req_name_index = []
                    res_name_index = []
                    portal_res_name_index = []
                    portal_eng_name_list = []
                    res_eng_name_list = []
                    ##요청항목 리스트 기준
                    for req_l in range(0, len(df_data['항목명(영문)'])):  # 요청항목 영문명 길이만큼 반복
                        if str(df_data['항목명(영문)'].iloc[req_l]).lower() == 'nan':  # 요청항목이 NaN값이면 중단
                            break
                        req_name_index.append(req_l)  # 요청항목 인덱스를 담음

                    temp_req.append(len(req_name_index))  # 인덱스 길이를 임시 리스트에 저장

                    ## 출력항목 리스트 기준
                    for res_l in range(0, len(df_data['응답항목명(영문)'])):  # 응답항목 영문명 길이만큼 반복
                        if str(df_data['응답항목명(영문)'].iloc[res_l]).lower() == 'nan':  # 응답항목이 NaN값이면 중단
                            break
                        res_name_index.append(res_l)  # 응답항목 인덱스를 담음
                        res_eng_name_list.append(str(df_data['응답항목명(영문)'].iloc[res_l]).lower)  # 응답항목 이름을 리스트에 담음
                    temp_res.append(len(res_name_index))  # 인덱스 길이를 임시 리스트에 저장

                    ##포털
                    for res_r in range(0, len(oper_res_df['OPERCD'])):  # 응답항목 영문명 길이만큼 반복
                        if str(oper_res_df['COLNMEN'].iloc[res_r]).lower() == 'nan':  # 응답항목이 NaN값이면 중단
                            break
                        portal_res_name_index.append(res_r)  # 응답항목 인덱스를 담음
                        portal_eng_name_list.append(oper_res_df['COLNMEN'].iloc[res_r])  # 응답항목 이름을 리스트에 담음
                    if len(portal_res_name_index) == 0:
                        portal_res_name_index.append(1)
                        portal_temp_res.append(1)
                    else:
                        portal_temp_res.append(len(portal_res_name_index))



                    ##오퍼레이션 추가 고려 요청항목 길이
                    req_length = 0
                    if i != 0:
                        for muimi in range(0, i):
                            req_length = req_length + temp_req[muimi]  ##임시 리스트에 저장한 값을 더해나감
                    ##오퍼레이션 추가 고려 응답항목 길이
                    res_length = 0
                    if i != 0:
                        for muimi in range(0, i):
                            res_length = res_length + temp_res[muimi]

                    ##xml 태그 이름 추출
                    portal_res_lenth = 0

                    if i != 0:
                        for muimi in range(0, i):
                            portal_res_lenth = portal_res_lenth + portal_temp_res[muimi]

                    real_tagnames = []
                    if req_html.startswith('<') or req_html.startswith('This XML'):
                        tagnams = re.findall('<.*?>', req_html)
                        for tagnam in tagnams:
                            if not str(tagnam).startswith('</') and not str(tagnam).startswith('<?') and not str(
                                    tagnam).startswith(
                                    '<!'):
                                real_tagnames.append(tagnam.lstrip('<').rstrip('>').lower())
                    elif '{' in req_html:
                        real_tagnames = re.findall('"([^"]*)"', req_html)

                    real_tagnames = set(real_tagnames)
                    real_tagnames = sorted(real_tagnames)
                    num1htmllist = oper_req_df['COLNMEN'].tolist()
                    num1reqlist = df_data['항목명(영문)'].values.tolist()
                    ## 1번항목 *iloc(y,x)
                    ## 1번은 요청항목 기준
                    resultlist1_temp = []
                    num1reqlist_temp = []
                    num1htmllist_temp = []
                    for j in num1reqlist:  ##리소스에 적혀있는 요청변수만큼 반복한다
                        if str(j) == 'nan':  ##만약 리소스에 적혀 있는 요청변수가 비어있다면 반복을 멈춘다
                            break
                        if len(num1htmllist) == 0:  ##만약 포털이 비어있다면 멈춘다
                            break
                        if num1htmllist[0] == '검색 결과가 없습니다.':
                            break
                        for jk in num1htmllist:

                            if str(j).lower().strip().replace(" ", "") == str(jk).lower().strip().replace(" ", ""):
                                num1reqlist_temp.append(j)
                                num1htmllist_temp.append(jk)

                                num1_cnt1 = num1_cnt1 + 1
                                s3['A' + str(8 + num1_cnt1 - 1)].value = str(num1_cnt1)
                                exel_font_set(s3['A' + str(8 + num1_cnt1 - 1)])

                                s3['B' + str(8 + num1_cnt1 - 1)].value = opername
                                exel_font_set(s3['B' + str(8 + num1_cnt1 - 1)])

                                s3['C' + str(8 + num1_cnt1 - 1)].value = j
                                exel_font_set(s3['C' + str(8 + num1_cnt1 - 1)])
                                try:
                                    s3['D' + str(8 + num1_cnt1 - 1)].value = \
                                        df_data[df_data['항목명(영문)'] == j]['항목명(국문)'].values[0]
                                    exel_font_set(s3['D' + str(8 + num1_cnt1 - 1)])
                                except IndexError:
                                    s3['D' + str(8 + num1_cnt1 - 1)].value = ''
                                    exel_font_set(s3['D' + str(8 + num1_cnt1 - 1)])
                                try:
                                    s3['E' + str(8 + num1_cnt1 - 1)].value = \
                                        df_data[df_data['항목명(영문)'] == j]['항목크기'].values[0]
                                    exel_font_set(s3['E' + str(8 + num1_cnt1 - 1)])
                                except IndexError:
                                    s3['E' + str(8 + num1_cnt1 - 1)].value = ''
                                    exel_font_set(s3['E' + str(8 + num1_cnt1 - 1)])
                                try:
                                    s3['F' + str(8 + num1_cnt1 - 1)].value = \
                                        df_data[df_data['항목명(영문)'] == j]['항목구분'].values[0]
                                    exel_font_set(s3['F' + str(8 + num1_cnt1 - 1)])
                                except IndexError:
                                    s3['F' + str(8 + num1_cnt1 - 1)].value = ''
                                    exel_font_set(s3['F' + str(8 + num1_cnt1 - 1)])
                                s3['G' + str(8 + num1_cnt1 - 1)].value = str(num1_cnt1)
                                exel_font_set(s3['G' + str(8 + num1_cnt1 - 1)])
                                s3['H' + str(8 + num1_cnt1 - 1)].value = opername
                                exel_font_set(s3['H' + str(8 + num1_cnt1 - 1)])
                                s3['I' + str(8 + num1_cnt1 - 1)].value = jk
                                exel_font_set(s3['I' + str(8 + num1_cnt1 - 1)])
                                s3['J' + str(8 + num1_cnt1 - 1)].value = \
                                    oper_req_df[oper_req_df['COLNMEN'] == jk]['COLNMKR'].values[0]
                                exel_font_set(s3['J' + str(8 + num1_cnt1 - 1)])
                                s3['K' + str(8 + num1_cnt1 - 1)].value = \
                                    oper_req_df[oper_req_df['COLNMEN'] == jk]['COLSIZE'].values[0]
                                exel_font_set(s3['K' + str(8 + num1_cnt1 - 1)])
                                s3['L' + str(8 + num1_cnt1 - 1)].value = \
                                    oper_req_df[oper_req_df['COLNMEN'] == jk]['COLTYPE'].values[0]
                                exel_font_set(s3['L' + str(8 + num1_cnt1 - 1)])
                                s3['M' + str(8 + num1_cnt1 - 1)].value = '정상'
                                resultlist1_temp.append(s3['M' + str(8 + num1_cnt1 - 1)].value)
                                exel_font_set(s3['M' + str(8 + num1_cnt1 - 1)])

                    num1reqlist = [x for x in num1reqlist if x not in num1reqlist_temp]
                    num1htmllist = [x for x in num1htmllist if x not in num1htmllist_temp]

                    if len(num1reqlist) > 0:
                        for j in num1reqlist:
                            if str(j) == 'nan':
                                continue
                            num1_cnt1 = num1_cnt1 + 1
                            s3['A' + str(8 + num1_cnt1 - 1)].value = str(num1_cnt1)
                            exel_font_set(s3['A' + str(8 + num1_cnt1 - 1)])
                            s3['G' + str(8 + num1_cnt1 - 1)].value = str(num1_cnt1)
                            exel_font_set(s3['G' + str(8 + num1_cnt1 - 1)])
                            s3['B' + str(8 + num1_cnt1 - 1)].value = opername
                            exel_font_set(s3['B' + str(8 + num1_cnt1 - 1)])
                            s3['H' + str(8 + num1_cnt1 - 1)].value = opername
                            exel_font_set(s3['H' + str(8 + num1_cnt1 - 1)])
                            s3['C' + str(8 + num1_cnt1 - 1)].value = j
                            exel_font_set(s3['C' + str(8 + num1_cnt1 - 1)])
                            s3['D' + str(8 + num1_cnt1 - 1)].value = \
                                df_data[df_data['항목명(영문)'] == j]['항목명(국문)'].values[0]
                            exel_font_set(s3['D' + str(8 + num1_cnt1 - 1)])
                            s3['E' + str(8 + num1_cnt1 - 1)].value = \
                                df_data[df_data['항목명(영문)'] == j]['항목크기'].values[0]
                            exel_font_set(s3['E' + str(8 + num1_cnt1 - 1)])
                            s3['F' + str(8 + num1_cnt1 - 1)].value = \
                                df_data[df_data['항목명(영문)'] == j]['항목구분'].values[0]
                            exel_font_set(s3['F' + str(8 + num1_cnt1 - 1)])
                            s3['I' + str(8 + num1_cnt1 - 1)].value = '동일항목없음'
                            exel_font_set(s3['I' + str(8 + num1_cnt1 - 1)])
                            s3['J' + str(8 + num1_cnt1 - 1)].value = '동일항목없음'
                            exel_font_set(s3['J' + str(8 + num1_cnt1 - 1)])
                            exel_font_set(s3['K' + str(8 + num1_cnt1 - 1)])
                            exel_font_set(s3['L' + str(8 + num1_cnt1 - 1)])
                            s3['M' + str(8 + num1_cnt1 - 1)].value = '오류'
                            resultlist1_temp.append(s3['M' + str(8 + num1_cnt1 - 1)].value)
                            exel_font_set(s3['M' + str(8 + num1_cnt1 - 1)])

                    if len(num1htmllist) > 0:
                        for j in num1htmllist:
                            if str(j) == 'nan':
                                continue
                            elif str(j) == '검색 결과가 없습니다.':
                                continue
                            num1_cnt1 = num1_cnt1 + 1
                            s3['G' + str(8 + num1_cnt1 - 1)].value = str(num1_cnt1 - 1 + 1)
                            exel_font_set(s3['G' + str(8 + num1_cnt1 - 1)])
                            s3['A' + str(8 + num1_cnt1 - 1)].value = str(num1_cnt1 - 1 + 1)
                            exel_font_set(s3['A' + str(8 + num1_cnt1 - 1)])
                            s3['H' + str(8 + num1_cnt1 - 1)].value = opername
                            exel_font_set(s3['H' + str(8 + num1_cnt1 - 1)])
                            s3['B' + str(8 + num1_cnt1 - 1)].value = opername
                            exel_font_set(s3['B' + str(8 + num1_cnt1 - 1)])
                            s3['I' + str(8 + num1_cnt1 - 1)].value = j
                            exel_font_set(s3['I' + str(8 + num1_cnt1 - 1)])
                            s3['J' + str(8 + num1_cnt1 - 1)].value = \
                                oper_req_df[oper_req_df['COLNMEN'] == j]['COLNMKR'].values[0]
                            exel_font_set(s3['J' + str(8 + num1_cnt1 - 1)])
                            s3['K' + str(8 + num1_cnt1 - 1)].value = \
                                oper_req_df[oper_req_df['COLNMEN'] == j]['COLSIZE'].values[0]
                            exel_font_set(s3['K' + str(8 + num1_cnt1 - 1)])
                            s3['L' + str(8 + num1_cnt1 - 1)].value = \
                                oper_req_df[oper_req_df['COLNMEN'] == j]['COLTYPE'].values[0]
                            exel_font_set(s3['L' + str(8 + num1_cnt1 - 1)])
                            s3['M' + str(8 + num1_cnt1 - 1)].value = '오류'
                            resultlist1_temp.append(s3['M' + str(8 + num1_cnt1 - 1)].value)
                            exel_font_set(s3['M' + str(8 + num1_cnt1 - 1)])
                            exel_font_set(s3['B' + str(8 + num1_cnt1 - 1)])
                            s3['C' + str(8 + num1_cnt1 - 1)].value = '동일항목없음'
                            s3['D' + str(8 + num1_cnt1 - 1)].value = '동일항목없음'
                            exel_font_set(s3['C' + str(8 + num1_cnt1 - 1)])
                            exel_font_set(s3['D' + str(8 + num1_cnt1 - 1)])
                            exel_font_set(s3['E' + str(8 + num1_cnt1 - 1)])
                            exel_font_set(s3['F' + str(8 + num1_cnt1 - 1)])

                    if '오류' in resultlist1_temp:
                        result_list1.append('오류')
                    else:
                        result_list1.append('정상')

                    s4 = wb_result['20.S-2']
                    ## 2번항목
                    s4['A' + str(8 + i)].value = str(i + 1)
                    exel_font_set(s4['A' + str(8 + i)])
                    s4['B' + str(8 + i)].value = opername
                    exel_font_set(s4['B' + str(8 + i)])
                    s4['C' + str(8 + i)].value = api_req_link
                    exel_font_set(s4['C' + str(8 + i)])
                    s4['D' + str(8 + i)].value = df_data.iloc[0, 13]
                    exel_font_set(s4['D' + str(8 + i)])
                    s4['E' + str(8 + i)].value = str(i + 1)
                    exel_font_set(s4['E' + str(8 + i)])
                    s4['F' + str(8 + i)].value = opername
                    exel_font_set(s4['F' + str(8 + i)])
                    s4['G' + str(8 + i)].value = api_req_link
                    exel_font_set(s4['G' + str(8 + i)])
                    s4['H' + str(8 + i)].value = req_html[:1000]
                    exel_font_set(s4['H' + str(8 + i)])
                    matchraio = int(
                        SequenceMatcher(None, str(req_html.replace(
                            'This XML file does not appear to have any style information associated with it. The document tree is shown below.',
                            '')).strip().replace('\n', '').replace('\t', '').replace(' ', '')[:200],
                                        str(df_data.iloc[0, 13]).strip().replace('\n', '').replace('\t',
                                                                                                   '').replace(' ',
                                                                                                               '')[
                                        :200]).ratio() * 100)
                    if 'normal service' in req_html.lower():
                        s4['I' + str(8 + i)].value = "정상"
                    elif '<result_msg>ok</result_msg>' in req_html.replace(' ', '').replace('\n', '').lower():
                        s4['I' + str(8 + i)].value = "정상"
                    elif 'normal_service' in req_html.lower():
                        s4['I' + str(8 + i)].value = "정상"
                    elif 'non_error' in req_html.lower():
                        s4['I' + str(8 + i)].value = "정상"
                    elif req_html == '응답없음':
                        s4['I' + str(8 + i)].value = "오류"
                    elif 'success' in req_html.lower():
                        s4['I' + str(8 + i)].value = "정상"
                    elif '정상' in req_html:
                        s4['I' + str(8 + i)].value = "정상"
                    elif '성공' in req_html:
                        s4['I' + str(8 + i)].value = "정상"
                    elif '이미지' in req_html:
                        s4['I' + str(8 + i)].value = "정상"
                    elif 'success_info' in req_html.lower():
                        s4['I' + str(8 + i)].value = "정상"
                    elif matchraio > 70:
                        s4['I' + str(8 + i)].value = "정상"
                    elif '<successYN>N</successYN>' in req_html.lower():
                        s4['I' + str(8 + i)].value = "오류"
                    elif 'normal' in req_html.lower():
                        s4['I' + str(8 + i)].value = '정상'
                    elif 'nomal' in req_html.lower():
                        s4['I' + str(8 + i)].value = '정상'
                    elif '<successYN>Y</successYN>' in req_html.lower():
                        s4['I' + str(8 + i)].value = '정상'
                    elif '<resultcode>00' in req_html.lower():
                        s4['I' + str(8 + i)].value = '정상'
                    elif 'wfs' in req_html.lower():
                        s4['I' + str(8 + i)].value = '정상'
                    elif 'soapenv' in req_html.lower():
                        s4['I' + str(8 + i)].value = "오류"
                    elif 'errorcode' in req_html.lower():
                        s4['I' + str(8 + i)].value = "오류"
                    elif 'service error' in req_html.lower():
                        s4['I' + str(8 + i)].value = "오류"
                    elif 'SERVICE KEY IS NOT' in req_html.upper():
                        s4['I' + str(8 + i)].value = "오류"
                    elif 'APPLICATION ERROR' in req_html.upper():
                        s4['I' + str(8 + i)].value = "오류"
                    elif 'DB ERROR' in req_html.upper():
                        s4['I' + str(8 + i)].value = "오류"
                    elif 'NODATA ERROR' in req_html.upper():
                        s4['I' + str(8 + i)].value = "오류"
                    elif 'HTTP ERROR' in req_html.upper():
                        s4['I' + str(8 + i)].value = "오류"
                    elif 'SERVICETIMEOUT ERROR' in req_html.upper():
                        s4['I' + str(8 + i)].value = "오류"
                    elif 'INVALID_REQUEST PARAMETER_ERROR' in req_html.upper():
                        s4['I' + str(8 + i)].value = "오류"
                    elif 'NO MANDATORY REQUEST PARAMETERS ERROR' in req_html.upper():
                        s4['I' + str(8 + i)].value = "오류"
                    elif 'NO OPENAPI SERVICE ERROR' in req_html.upper():
                        s4['I' + str(8 + i)].value = "오류"
                    elif 'SERVICE ACCESS DENIED ERROR' in req_html.upper():
                        s4['I' + str(8 + i)].value = "오류"
                    elif 'LIMITED NUMBER OF SERVICE REQUESTS EXCEEDS ERROR' in req_html.upper():
                        s4['I' + str(8 + i)].value = "오류"
                    elif 'SERVICE KEY IS NOT REGISTERED ERROR' in req_html.upper():
                        s4['I' + str(8 + i)].value = "오류"
                    elif 'DEADLINE HAS EXPIRED ERROR' in req_html.upper():
                        s4['I' + str(8 + i)].value = "오류"
                    elif 'UNREGISTERED IP ERROR' in req_html.upper():
                        s4['I' + str(8 + i)].value = "오류"
                    elif 'INVALID REQUEST PARAMETER ERROR' in req_html.upper():
                        s4['I' + str(8 + i)].value = "오류"
                    else:
                        s4['I' + str(8 + i)].value = "오류"

                    exel_font_set(s4['I' + str(8 + i)])
                    result_list2.append(s4['I' + str(8 + i)].value)

                    ## 3번 항목
                    s5 = wb_result['30.S-3']
                    result3_temp = []
                    for j in range(0, num1_cnt1):
                        s5['A' + str(8 + j)].value = s3['A' + str(8 + j)].value
                        exel_font_set(s5['A' + str(8 + j)])
                        s5['B' + str(8 + j)].value = s3['B' + str(8 + j)].value
                        exel_font_set(s5['B' + str(8 + j)])
                        s5['C' + str(8 + j)].value = s3['C' + str(8 + j)].value
                        exel_font_set(s5['C' + str(8 + j)])
                        s5['D' + str(8 + j)].value = s3['D' + str(8 + j)].value
                        exel_font_set(s5['D' + str(8 + j)])
                        s5['E' + str(8 + j)].value = s3['E' + str(8 + j)].value
                        exel_font_set(s5['E' + str(8 + j)])
                        s5['F' + str(8 + j)].value = s3['F' + str(8 + j)].value
                        exel_font_set(s5['F' + str(8 + j)])
                        s5['G' + str(8 + j)].value = s3['G' + str(8 + j)].value
                        exel_font_set(s5['G' + str(8 + j)])
                        s5['H' + str(8 + j)].value = s3['H' + str(8 + j)].value
                        exel_font_set(s5['H' + str(8 + j)])
                        s5['I' + str(8 + j)].value = s3['I' + str(8 + j)].value
                        exel_font_set(s5['I' + str(8 + j)])
                        s5['J' + str(8 + j)].value = s3['J' + str(8 + j)].value
                        exel_font_set(s5['J' + str(8 + j)])
                        s5['K' + str(8 + j)].value = s3['K' + str(8 + j)].value
                        exel_font_set(s5['K' + str(8 + j)])
                        s5['L' + str(8 + j)].value = s3['L' + str(8 + j)].value
                        exel_font_set(s5['L' + str(8 + j)])
                        if s5['M' + str(8 + j)].value == None:
                            if str(s5['I' + str(8 + j)].value) == '동일항목없음' or str(
                                    s5['C' + str(8 + j)].value) == '동일항목없음':
                                s5['M' + str(8 + j)].value = '오류'
                                result3_temp.append("오류")
                            elif str(s5['F' + str(8 + j)].value).startswith('0') and str(
                                    s5['L' + str(8 + j)].value).startswith('옵'):
                                s5['M' + str(8 + j)].value = '정상'
                            elif str(s5['F' + str(8 + j)].value).startswith('0') and str(
                                    s5['L' + str(8 + j)].value) == '':
                                s5['M' + str(8 + j)].value = '정상'
                            elif str(s5['F' + str(8 + j)].value) == '선택' and str(
                                    s5['L' + str(8 + j)].value).startswith('옵'):
                                s5['M' + str(8 + j)].value = '정상'
                            elif '0' in str(s5['F' + str(8 + j)].value) and str(
                                    s5['L' + str(8 + j)].value) == '':
                                s5['M' + str(8 + j)].value = '정상'
                            elif str(s5['F' + str(8 + j)].value) == '선택' and str(
                                    s5['L' + str(8 + j)].value) == '':
                                s5['M' + str(8 + j)].value = '정상'
                            elif '0' in str(s5['F' + str(8 + j)].value) and str(
                                    s5['L' + str(8 + j)].value) == 'nan':
                                s5['M' + str(8 + j)].value = '정상'
                            elif str(s5['F' + str(8 + j)].value) == 'nan' and str(
                                    s5['L' + str(8 + j)].value).startswith('옵'):
                                s5['M' + str(8 + j)].value = '정상'
                            elif str(s5['F' + str(8 + j)].value).startswith('옵') and str(
                                    s5['L' + str(8 + j)].value).startswith('옵'):
                                s5['M' + str(8 + j)].value = '정상'
                            elif str(s5['F' + str(8 + j)].value).startswith('필') and str(
                                    s5['L' + str(8 + j)].value).startswith('필'):
                                s5['M' + str(8 + j)].value = '정상'
                            elif str(s5['F' + str(8 + j)].value) == '선택' and str(
                                    s5['L' + str(8 + j)].value) == 'nan':
                                s5['M' + str(8 + j)].value = '정상'
                            elif str(s5['F' + str(8 + j)].value) == str(
                                    s5['L' + str(8 + j)].value):
                                s5['M' + str(8 + j)].value = '정상'
                            elif str(s5['F' + str(8 + j)].value).startswith('1') and str(
                                    s5['L' + str(8 + j)].value).startswith('1'):
                                s5['M' + str(8 + j)].value = '정상'
                            elif str(s5['F' + str(8 + j)].value).startswith('0') and str(
                                    s5['L' + str(8 + j)].value).startswith('0'):
                                s5['M' + str(8 + j)].value = '정상'
                            elif str(s5['F' + str(8 + j)].value).startswith('0') and str(
                                    s5['L' + str(8 + j)].value).startswith('필'):
                                s5['M' + str(8 + j)].value = '오류'
                                result3_temp.append("오류")
                            elif str(s5['F' + str(8 + j)].value).startswith('0') and str(
                                    s5['L' + str(8 + j)].value) == 'nan':
                                s5['M' + str(8 + j)].value = '정상'
                            elif str(s5['F' + str(8 + j)].value).startswith('1') and str(
                                    s5['L' + str(8 + j)].value).startswith('필'):
                                s5['M' + str(8 + j)].value = '정상'
                            else:
                                s5['M' + str(8 + j)].value = '오류'
                                result3_temp.append("오류")
                            exel_font_set(s5['M' + str(8 + j)])

                    if '오류' in result3_temp:
                        result_list3.append('오류')

                    ## 4번 항목
                    ## 응답항목 기준
                    s6 = wb_result['40.S-4']
                    if len(res_name_index) == 0 or str(df_data.iloc[0, 2]) == 'nan':
                        s6['A' + str(8 + s6cnt)].value = str(1 + s6cnt)
                        exel_font_set(s6['A' + str(8 + s6cnt)])
                        s6['B' + str(8 + s6cnt)].value = opername
                        exel_font_set(s6['B' + str(8 + s6cnt)])
                        s6['C' + str(8 + s6cnt)].value = '항목없음'
                        exel_font_set(s6['C' + str(8 + s6cnt)])
                        s6['D' + str(8 + s6cnt)].value = str(8 + s6cnt - 7)
                        exel_font_set(s6['D' + str(8 + s6cnt)])
                        s6['E' + str(8 + s6cnt)].value = opername
                        exel_font_set(s6['E' + str(8 + s6cnt)])
                        s6['F' + str(8 + s6cnt)].value = req_html[:1000]
                        if req_html == "응답없음":
                            s6['G' + str(8 + s6cnt)].value = '오류'
                        elif '<result_msg>ok</result_msg>' in req_html.replace(' ', '').replace('\n', '').lower():
                            s6['G' + str(8 + s6cnt)].value = '정상'
                        elif '이미지' in req_html:
                            s6['G' + str(8 + s6cnt)].value = '정상'
                        elif 'normal' in req_html.lower():
                            s6['G' + str(8 + s6cnt)].value = '정상'
                        elif 'non_error' in req_html.lower():
                            s6['G' + str(8 + s6cnt)].value = '정상'
                        elif 'nomal' in req_html.lower():
                            s6['G' + str(8 + s6cnt)].value = '정상'
                        elif '<successyn>y</successyn>' in req_html.lower():
                            s6['G' + str(8 + s6cnt)].value = '정상'
                        elif '정상' in req_html.lower():
                            s6['G' + str(8 + s6cnt)].value = '정상'
                        elif '성공' in req_html.lower():
                            s6['G' + str(8 + s6cnt)].value = '정상'
                        elif 'success' in req_html.lower():
                            s6['G' + str(8 + s6cnt)].value = '정상'
                        elif '>ok<' in req_html.lower():
                            s6['G' + str(8 + s6cnt)].value = '정상'
                        elif 'wfs' in req_html.lower():
                            s6['G' + str(8 + s6cnt)].value = '정상'
                        elif '<resultcode>00' in req_html.lower():
                            s6['G' + str(8 + s6cnt)].value = '정상'
                        elif 'soapenv' in req_html.lower():
                            s6['G' + str(8 + s6cnt)].value = "오류"
                        elif '<successyn>n</successyn>' in req_html.lower():
                            s6['G' + str(8 + s6cnt)].value = "오류"
                        elif 'errorcode' in req_html.lower():
                            s6['G' + str(8 + s6cnt)].value = "오류"
                        elif 'service error' in req_html.lower():
                            s6['G' + str(8 + s6cnt)].value = "오류"
                        elif 'SERVICE KEY IS NOT' in req_html.upper():
                            s6['G' + str(8 + s6cnt)].value = "오류"
                        elif 'APPLICATION ERROR' in req_html.upper():
                            s6['G' + str(8 + s6cnt)].value = "오류"
                        elif 'DB ERROR' in req_html.upper():
                            s6['G' + str(8 + s6cnt)].value = "오류"
                        elif 'NODATA ERROR' in req_html.upper():
                            s6['G' + str(8 + s6cnt)].value = "오류"
                        elif 'HTTP ERROR' in req_html.upper():
                            s6['G' + str(8 + s6cnt)].value = "오류"
                        elif 'SERVICETIMEOUT ERROR' in req_html.upper():
                            s6['G' + str(8 + s6cnt)].value = "오류"
                        elif 'INVALID_REQUEST PARAMETER_ERROR' in req_html.upper():
                            s6['G' + str(8 + s6cnt)].value = "오류"
                        elif 'NO MANDATORY REQUEST PARAMETERS ERROR' in req_html.upper():
                            s6['G' + str(8 + s6cnt)].value = "오류"
                        elif 'NO OPENAPI SERVICE ERROR' in req_html.upper():
                            s6['G' + str(8 + s6cnt)].value = "오류"
                        elif 'SERVICE ACCESS DENIED ERROR' in req_html.upper():
                            s6['G' + str(8 + s6cnt)].value = "오류"
                        elif 'LIMITED NUMBER OF SERVICE REQUESTS EXCEEDS ERROR' in req_html.upper():
                            s6['G' + str(8 + s6cnt)].value = "오류"
                        elif 'SERVICE KEY IS NOT REGISTERED ERROR' in req_html.upper():
                            s6['G' + str(8 + s6cnt)].value = "오류"
                        elif 'DEADLINE HAS EXPIRED ERROR' in req_html.upper():
                            s6['G' + str(8 + s6cnt)].value = "오류"
                        elif 'UNREGISTERED IP ERROR' in req_html.upper():
                            s6['G' + str(8 + s6cnt)].value = "오류"
                        elif 'INVALID REQUEST PARAMETER ERROR' in req_html.upper():
                            s6['G' + str(8 + s6cnt)].value = "오류"
                        elif len(set(res_eng_name_list)) == 0:
                            pass
                        elif len(set(res_eng_name_list) - set(real_tagnames)) / len(set(res_eng_name_list)) < 0.7:
                            s6['G' + str(8 + s6cnt)].value = '정상'
                        else:
                            s6['G' + str(8 + s6cnt)].value = '오류'
                        exel_font_set(s6['F' + str(8 + s6cnt)])
                        exel_font_set(s6['G' + str(8 + s6cnt)])
                        result_list4.append(s6['G' + str(8 + s6cnt)].value)
                        s6cnt = s6cnt + 1
                    else:
                        s6.merge_cells("D" + str(8 + s6cnt) + ":" + 'D' + str(
                            8 + s6cnt + len(res_name_index) - 1))
                        s6.merge_cells("E" + str(8 + s6cnt) + ":" + 'E' + str(
                            8 + len(res_name_index) + s6cnt - 1))
                        s6.merge_cells("F" + str(8 + s6cnt) + ":" + 'F' + str(
                            8 + len(res_name_index) + s6cnt - 1))
                        s6.merge_cells("G" + str(8 + s6cnt) + ":" + 'G' + str(
                            8 + len(res_name_index) + s6cnt - 1))
                        style_range(s6, "D" + str(8 + s6cnt) + ":" + 'D' + str(
                            8 + s6cnt + len(res_name_index) - 1),
                                                border=Border(left=Side(style='thin'), right=Side(style='thin'),
                                                              top=Side(style='thin'),
                                                              bottom=Side(style='thin')))
                        style_range(s6, "E" + str(8 + s6cnt) + ":" + 'E' + str(
                            8 + s6cnt + len(res_name_index) - 1),
                                                border=Border(left=Side(style='thin'), right=Side(style='thin'),
                                                              top=Side(style='thin'),
                                                              bottom=Side(style='thin')))
                        style_range(s6, "F" + str(8 + s6cnt) + ":" + 'F' + str(
                            8 + s6cnt + len(res_name_index) - 1),
                                                border=Border(left=Side(style='thin'), right=Side(style='thin'),
                                                              top=Side(style='thin'),
                                                              bottom=Side(style='thin')))
                        style_range(s6, "G" + str(8 + s6cnt) + ":" + 'G' + str(
                            8 + s6cnt + len(res_name_index) - 1),
                                                border=Border(left=Side(style='thin'), right=Side(style='thin'),
                                                              top=Side(style='thin'),
                                                              bottom=Side(style='thin')))
                        s6['D' + str(8 + s6cnt)].value = str(i + 1)
                        exel_font_set(s6['D' + str(8 + s6cnt)])
                        s6['E' + str(8 + s6cnt)].value = opername
                        exel_font_set(s6['E' + str(8 + s6cnt)])
                        s6['F' + str(8 + s6cnt)].value = req_html[:1000]
                        exel_font_set(s6['F' + str(8 + s6cnt)])
                        res_eng_name_list = []
                        s6cnt2 = s6cnt
                        for j in range(0, len(res_name_index)):
                            s6['A' + str(8 + s6cnt)].value = str(1 + s6cnt)
                            exel_font_set(s6['A' + str(8 + s6cnt)])
                            s6['B' + str(8 + s6cnt)].value = opername
                            exel_font_set(s6['B' + str(8 + s6cnt)])
                            s6['C' + str(8 + s6cnt)].value = df_data.iloc[j, 9]
                            exel_font_set(s6['C' + str(8 + s6cnt)])
                            res_eng_name_list.append(str(df_data.iloc[j, 9]))
                            s6cnt = s6cnt + 1

                        if req_html == "응답없음":
                            s6['G' + str(8 + s6cnt2)].value = '오류'
                        elif '<result_msg>ok</result_msg>' in req_html.replace(' ', '').replace('\n', '').lower():
                            s6['G' + str(8 + s6cnt2)].value = '정상'
                        elif '이미지' in req_html:
                            s6['G' + str(8 + s6cnt2)].value = '정상'
                        elif 'non_error' in req_html.lower():
                            s6['G' + str(8 + s6cnt2)].value = '정상'
                        elif 'normal' in req_html.lower():
                            s6['G' + str(8 + s6cnt2)].value = '정상'
                        elif '정상' in req_html.lower():
                            s6['G' + str(8 + s6cnt2)].value = '정상'
                        elif '성공' in req_html.lower():
                            s6['G' + str(8 + s6cnt2)].value = '정상'
                        elif 'success' in req_html.lower():
                            s6['G' + str(8 + s6cnt2)].value = '정상'
                        elif '>ok<' in req_html.lower():
                            s6['G' + str(8 + s6cnt2)].value = '정상'
                        elif 'nomal' in req_html.lower():
                            s6['G' + str(8 + s6cnt2)].value = '정상'
                        elif '<successYN>Y</successYN>' in req_html.lower():
                            s6['G' + str(8 + s6cnt2)].value = '정상'
                        elif '<resultcode>00' in req_html.lower():
                            s6['G' + str(8 + s6cnt2)].value = '정상'
                        elif 'wfs' in req_html.lower():
                            s6['G' + str(8 + s6cnt2)].value = '정상'
                        elif 'soapenv' in req_html.lower():
                            s6['G' + str(8 + s6cnt2)].value = "오류"
                        elif '<successYN>N</successYN>' in req_html.lower():
                            s6['G' + str(8 + s6cnt2)].value = "오류"
                        elif 'errorcode' in req_html.lower():
                            s6['G' + str(8 + s6cnt2)].value = "오류"
                        elif 'service error' in req_html.lower():
                            s6['G' + str(8 + s6cnt2)].value = "오류"
                        elif 'SERVICE KEY IS NOT' in req_html.upper():
                            s6['G' + str(8 + s6cnt2)].value = "오류"
                        elif 'APPLICATION ERROR' in req_html.upper():
                            s6['G' + str(8 + s6cnt2)].value = "오류"
                        elif 'DB ERROR' in req_html.upper():
                            s6['G' + str(8 + s6cnt2)].value = "오류"
                        elif 'NODATA ERROR' in req_html.upper():
                            s6['G' + str(8 + s6cnt2)].value = "오류"
                        elif 'HTTP ERROR' in req_html.upper():
                            s6['G' + str(8 + s6cnt2)].value = "오류"
                        elif 'SERVICETIMEOUT ERROR' in req_html.upper():
                            s6['G' + str(8 + s6cnt2)].value = "오류"
                        elif 'INVALID_REQUEST PARAMETER_ERROR' in req_html.upper():
                            s6['G' + str(8 + s6cnt2)].value = "오류"
                        elif 'NO MANDATORY REQUEST PARAMETERS ERROR' in req_html.upper():
                            s6['G' + str(8 + s6cnt2)].value = "오류"
                        elif 'NO OPENAPI SERVICE ERROR' in req_html.upper():
                            s6['G' + str(8 + s6cnt2)].value = "오류"
                        elif 'SERVICE ACCESS DENIED ERROR' in req_html.upper():
                            s6['G' + str(8 + s6cnt2)].value = "오류"
                        elif 'LIMITED NUMBER OF SERVICE REQUESTS EXCEEDS ERROR' in req_html.upper():
                            s6['G' + str(8 + s6cnt2)].value = "오류"
                        elif 'SERVICE KEY IS NOT REGISTERED ERROR' in req_html.upper():
                            s6['G' + str(8 + s6cnt2)].value = "오류"
                        elif 'DEADLINE HAS EXPIRED ERROR' in req_html.upper():
                            s6['G' + str(8 + s6cnt2)].value = "오류"
                        elif 'UNREGISTERED IP ERROR' in req_html.upper():
                            s6['G' + str(8 + s6cnt2)].value = "오류"
                        elif 'INVALID REQUEST PARAMETER ERROR' in req_html.upper():
                            s6['G' + str(8 + s6cnt2)].value = "오류"
                        elif len(set(res_eng_name_list)) == 0:
                            pass
                        elif len(set(res_eng_name_list) - set(real_tagnames)) / len(set(res_eng_name_list)) < 0.7:
                            s6['G' + str(8 + s6cnt2)].value = '정상'
                        else:
                            s6['G' + str(8 + s6cnt2)].value = '오류'
                        exel_font_set(s6['G' + str(8 + s6cnt2)])
                        result_list4.append(s6['G' + str(8 + s6cnt2)].value)

                    ## 5번 항목
                    ## 응답항목 기준
                    s7 = wb_result['50.S-5']
                    if len(oper_res_df['APICD'].tolist()) == 0 or oper_res_df['COLNMEN'].tolist()[0] == 'nan':
                        s7['A' + str(8 + s7cnt)].value = str(8 + s7cnt - 7)
                        exel_font_set(s7['A' + str(8 + s7cnt)])
                        s7['B' + str(8 + s7cnt)].value = opername
                        exel_font_set(s7['B' + str(8 + s7cnt)])
                        s7['C' + str(8 + s7cnt)].value = '항목없음'
                        exel_font_set(s7['C' + str(8 + s7cnt)])
                        s7['D' + str(8 + s7cnt)].value = str(8 + s7cnt - 7)
                        exel_font_set(s7['D' + str(8 + s7cnt)])
                        s7['E' + str(8 + s7cnt)].value = opername
                        exel_font_set(s7['E' + str(8 + s7cnt)])
                        if req_html == '이미지':
                            s7['F' + str(8 + s7cnt)].value = '이미지'
                        else:
                            s7['F' + str(8 + s7cnt)].value = ''

                        exel_font_set(s7['F' + str(8 + s7cnt)])
                        portal_engname_list = []

                        for portal_engname in oper_res_df['COLNMEN'].tolist():
                            portal_engname_list.append(portal_engname)
                        if req_html == "응답없음":
                            s7['G' + str(8 + s7cnt)].value = '오류'
                        elif '이미지' in req_html:
                            s7['G' + str(8 + s7cnt)].value = '정상'
                        elif 'normal' in req_html.lower():
                            s7['G' + str(8 + s7cnt)].value = '정상'
                        elif 'non_error' in req_html.lower():
                            s7['G' + str(8 + s7cnt)].value = '정상'
                        elif '정상' in req_html.lower():
                            s7['G' + str(8 + s7cnt)].value = '정상'
                        elif '성공' in req_html.lower():
                            s7['G' + str(8 + s7cnt)].value = '정상'
                        elif 'success' in req_html.lower():
                            s7['G' + str(8 + s7cnt)].value = '정상'
                        elif '>ok<' in req_html.lower():
                            s7['G' + str(8 + s7cnt)].value = '정상'
                        elif 'nomal' in req_html.lower():
                            s7['G' + str(8 + s7cnt)].value = '정상'
                        elif '<successyn>y</successyn>' in req_html.lower():
                            s7['G' + str(8 + s7cnt)].value = '정상'
                        elif '<resultcode>00' in req_html.lower():
                            s7['G' + str(8 + s7cnt)].value = '정상'
                        elif 'wfs' in req_html.lower():
                            s7['G' + str(8 + s7cnt)].value = '정상'
                        elif 'soapenv' in req_html.lower():
                            s7['G' + str(8 + s7cnt)].value = "오류"
                        elif 'errorcode' in req_html.lower():
                            s7['G' + str(8 + s7cnt)].value = "오류"
                        elif 'service error' in req_html.lower():
                            s7['G' + str(8 + s7cnt)].value = "오류"
                        elif 'SERVICE KEY IS NOT' in req_html.upper():
                            s7['G' + str(8 + s7cnt)].value = "오류"
                        elif '<successYN>N</successYN>' in req_html.lower():
                            s7['G' + str(8 + s7cnt)].value = "오류"
                        elif 'APPLICATION ERROR' in req_html.upper():
                            s7['G' + str(8 + s7cnt)].value = "오류"
                        elif 'DB ERROR' in req_html.upper():
                            s7['G' + str(8 + s7cnt)].value = "오류"
                        elif 'NODATA ERROR' in req_html.upper():
                            s7['G' + str(8 + s7cnt)].value = "오류"
                        elif 'HTTP ERROR' in req_html.upper():
                            s7['G' + str(8 + s7cnt)].value = "오류"
                        elif 'SERVICETIMEOUT ERROR' in req_html.upper():
                            s7['G' + str(8 + s7cnt)].value = "오류"
                        elif 'INVALID_REQUEST PARAMETER_ERROR' in req_html.upper():
                            s7['G' + str(8 + s7cnt)].value = "오류"
                        elif 'NO MANDATORY REQUEST PARAMETERS ERROR' in req_html.upper():
                            s7['G' + str(8 + s7cnt)].value = "오류"
                        elif 'NO OPENAPI SERVICE ERROR' in req_html.upper():
                            s7['G' + str(8 + s7cnt)].value = "오류"
                        elif 'SERVICE ACCESS DENIED ERROR' in req_html.upper():
                            s7['G' + str(8 + s7cnt)].value = "오류"
                        elif 'LIMITED NUMBER OF SERVICE REQUESTS EXCEEDS ERROR' in req_html.upper():
                            s7['G' + str(8 + s7cnt)].value = "오류"
                        elif 'SERVICE KEY IS NOT REGISTERED ERROR' in req_html.upper():
                            s7['G' + str(8 + s7cnt)].value = "오류"
                        elif 'DEADLINE HAS EXPIRED ERROR' in req_html.upper():
                            s7['G' + str(8 + s7cnt)].value = "오류"
                        elif 'UNREGISTERED IP ERROR' in req_html.upper():
                            s7['G' + str(8 + s7cnt)].value = "오류"
                        elif 'INVALID REQUEST PARAMETER ERROR' in req_html.upper():
                            s7['G' + str(8 + s7cnt)].value = "오류"
                        elif len(set(portal_engname_list)) == 0:
                            pass
                        elif len(set(portal_engname_list) - set(real_tagnames)) / len(
                                set(portal_engname_list)) < 0.7:
                            s7['G' + str(8 + s7cnt)].value = '정상'
                        else:
                            s7['G' + str(8 + s7cnt)].value = '오류'
                        exel_font_set(s7['G' + str(8 + s7cnt)])
                        result_list5.append(s7['G' + str(8 + s7cnt)].value)
                        s7cnt = s7cnt + 1
                    else:
                        s7.merge_cells("D" + str(8 + s7cnt) + ":" + 'D' + str(
                            8 + len(portal_res_name_index) + s7cnt - 1))
                        s7.merge_cells("E" + str(8 + s7cnt) + ":" + 'E' + str(
                            8 + len(portal_res_name_index) + s7cnt - 1))
                        s7.merge_cells("F" + str(8 + s7cnt) + ":" + 'F' + str(
                            8 + len(portal_res_name_index) + s7cnt - 1))
                        s7.merge_cells("G" + str(8 + s7cnt) + ":" + 'G' + str(
                            8 + len(portal_res_name_index) + s7cnt - 1))
                        style_range(s7, "D" + str(8 + s7cnt) + ":" + 'D' + str(
                            8 + len(portal_res_name_index) + s7cnt - 1),
                                                border=Border(left=Side(style='thin'), right=Side(style='thin'),
                                                              top=Side(style='thin'),
                                                              bottom=Side(style='thin')))
                        style_range(s7, "E" + str(8 + s7cnt) + ":" + 'E' + str(
                            8 + len(portal_res_name_index) + s7cnt - 1),
                                                border=Border(left=Side(style='thin'), right=Side(style='thin'),
                                                              top=Side(style='thin'),
                                                              bottom=Side(style='thin')))
                        style_range(s7, "F" + str(8 + s7cnt) + ":" + 'F' + str(
                            8 + len(portal_res_name_index) + s7cnt - 1),
                                                border=Border(left=Side(style='thin'), right=Side(style='thin'),
                                                              top=Side(style='thin'),
                                                              bottom=Side(style='thin')))
                        style_range(s7, "G" + str(8 + s7cnt) + ":" + 'G' + str(
                            8 + len(portal_res_name_index) + s7cnt - 1),
                                                border=Border(left=Side(style='thin'), right=Side(style='thin'),
                                                              top=Side(style='thin'),
                                                              bottom=Side(style='thin')))
                        s7cnt2 = s7cnt
                        for j in range(0, len(oper_res_df['APICD'].tolist())):
                            s7['A' + str(8 + s7cnt)].value = str(
                                + s7cnt + 1)
                            exel_font_set(s7['A' + str(8 + s7cnt)])
                            s7['B' + str(8 + s7cnt)].value = opername
                            exel_font_set(s7['B' + str(8 + s7cnt)])
                            s7['C' + str(8 + s7cnt)].value = oper_res_df['COLNMEN'].tolist()[j]
                            exel_font_set(s7['C' + str(8 + s7cnt)])
                            s7cnt = s7cnt + 1

                        s7['D' + str(8 + s7cnt2)].value = str(i + 1)
                        exel_font_set(s7['D' + str(8 + s7cnt2)])
                        s7['E' + str(8 + s7cnt2)].value = opername
                        exel_font_set(s7['E' + str(8 + s7cnt2)])
                        print(req_html)
                        s7['F' + str(8 + s7cnt2)].value = ILLEGAL_CHARACTERS_RE.sub(r'',
                                                                                    req_html[:1000].replace('\\n', ''))
                        exel_font_set(s7['F' + str(8 + s7cnt2)])
                        ##포털 항목명 영문 리스트 작성
                        portal_engname_list = []

                        for portal_engname in oper_res_df['COLNMEN']:
                            portal_engname_list.append(portal_engname)
                        if req_html == "응답없음":
                            s7['G' + str(8 + s7cnt2)].value = '오류'
                        elif '이미지' in req_html:
                            s7['G' + str(8 + s7cnt2)].value = '정상'
                        elif 'normal' in req_html.lower():
                            s7['G' + str(8 + s7cnt2)].value = '정상'
                        elif 'non_error' in req_html.lower():
                            s7['G' + str(8 + s7cnt2)].value = '정상'
                        elif '정상' in req_html.lower():
                            s7['G' + str(8 + s7cnt2)].value = '정상'
                        elif '성공' in req_html.lower():
                            s7['G' + str(8 + s7cnt2)].value = '정상'
                        elif 'success' in req_html.lower():
                            s7['G' + str(8 + s7cnt2)].value = '정상'
                        elif '>ok<' in req_html.lower():
                            s7['G' + str(8 + s7cnt2)].value = '정상'
                        elif 'nomal' in req_html.lower():
                            s7['G' + str(8 + s7cnt2)].value = '정상'
                        elif '<successYN>Y</successYN>' in req_html.lower():
                            s7['G' + str(8 + s7cnt2)].value = '정상'
                        elif '<resultcode>00' in req_html.lower():
                            s7['G' + str(8 + s7cnt2)].value = '정상'
                        elif 'wfs' in req_html.lower():
                            s7['G' + str(8 + s7cnt2)].value = '정상'
                        elif 'soapenv' in req_html.lower():
                            s7['G' + str(8 + s7cnt2)].value = "오류"
                        elif 'errorcode' in req_html.lower():
                            s7['G' + str(8 + s7cnt2)].value = "오류"
                        elif 'service error' in req_html.lower():
                            s7['G' + str(8 + s7cnt2)].value = "오류"
                        elif '<successYN>N</successYN>' in req_html.lower():
                            s7['G' + str(8 + s7cnt2)].value = "오류"
                        elif 'SERVICE KEY IS NOT' in req_html.upper():
                            s7['G' + str(8 + s7cnt2)].value = "오류"
                        elif 'APPLICATION ERROR' in req_html.upper():
                            s7['G' + str(8 + s7cnt2)].value = "오류"
                        elif 'DB ERROR' in req_html.upper():
                            s7['G' + str(8 + s7cnt2)].value = "오류"
                        elif 'NODATA ERROR' in req_html.upper():
                            s7['G' + str(8 + s7cnt2)].value = "오류"
                        elif 'HTTP ERROR' in req_html.upper():
                            s7['G' + str(8 + s7cnt2)].value = "오류"
                        elif 'SERVICETIMEOUT ERROR' in req_html.upper():
                            s7['G' + str(8 + s7cnt2)].value = "오류"
                        elif 'INVALID_REQUEST PARAMETER_ERROR' in req_html.upper():
                            s7['G' + str(8 + s7cnt2)].value = "오류"
                        elif 'NO MANDATORY REQUEST PARAMETERS ERROR' in req_html.upper():
                            s7['G' + str(8 + s7cnt2)].value = "오류"
                        elif 'NO OPENAPI SERVICE ERROR' in req_html.upper():
                            s7['G' + str(8 + s7cnt2)].value = "오류"
                        elif 'SERVICE ACCESS DENIED ERROR' in req_html.upper():
                            s7['G' + str(8 + s7cnt2)].value = "오류"
                        elif 'LIMITED NUMBER OF SERVICE REQUESTS EXCEEDS ERROR' in req_html.upper():
                            s7['G' + str(8 + s7cnt2)].value = "오류"
                        elif 'SERVICE KEY IS NOT REGISTERED ERROR' in req_html.upper():
                            s7['G' + str(8 + s7cnt2)].value = "오류"
                        elif 'DEADLINE HAS EXPIRED ERROR' in req_html.upper():
                            s7['G' + str(8 + s7cnt2)].value = "오류"
                        elif 'UNREGISTERED IP ERROR' in req_html.upper():
                            s7['G' + str(8 + s7cnt2)].value = "오류"
                        elif 'INVALID REQUEST PARAMETER ERROR' in req_html.upper():
                            s7['G' + str(8 + s7cnt2)].value = "오류"
                        elif len(set(portal_engname_list)) == 0:
                            pass
                        elif len(set(portal_engname_list) - set(real_tagnames)) / len(
                                set(portal_engname_list)) < 0.7:
                            s7['G' + str(8 + s7cnt2)].value = '정상'
                        else:
                            s7['G' + str(8 + s7cnt2)].value = '오류'
                        exel_font_set(s7['G' + str(8 + s7cnt2)])
                        result_list5.append(s7['G' + str(8 + s7cnt2)].value)

                    ## 7번 항목
                    s9 = wb_result['70.S-7']
                    print(dataformmat)
                    if dataformmat == 'XML' or dataformmat == 'JSON' or str(dataformmat) == 'nan':
                        s9['A' + str(9 + i)].value = str(i + 1)
                        exel_font_set(s9['A' + str(9 + i)])
                        s9['B' + str(9 + i)].value = opername
                        exel_font_set(s9['B' + str(9 + i)])
                        s9['C' + str(9 + i)].value = dataformmat
                        exel_font_set(s9['C' + str(9 + i)])
                        s9['D' + str(9 + i)].value = str(i + 1)
                        exel_font_set(s9['D' + str(9 + i)])
                        s9['E' + str(9 + i)].value = ILLEGAL_CHARACTERS_RE.sub(r'', req_html[:1000])
                        exel_font_set(s9['E' + str(9 + i)])
                        if '이미지' in req_html:
                            s9['F' + str(9 + i)].value = '정상'
                            exel_font_set(s9['F' + str(9 + i)])
                            result_list7.append('정상')
                        elif dataformmat == 'XML' and '</' in req_html[:1000]:
                            s9['F' + str(9 + i)].value = '정상'
                            exel_font_set(s9['F' + str(9 + i)])
                            result_list7.append('정상')
                        elif dataformmat == 'JSON' and '{' in req_html[:1000]:
                            s9['F' + str(9 + i)].value = '정상'
                            exel_font_set(s9['F' + str(9 + i)])
                            result_list7.append('정상')
                        else:
                            s9['F' + str(9 + i)].value = '오류'
                            exel_font_set(s9['F' + str(9 + i)])
                            result_list7.append('오류')
                    else:
                        s9['A' + str(9 + i * 2)].value = str(i * 2 + 1)
                        exel_font_set(s9['A' + str(9 + i * 2)])
                        s9['A' + str(9 + i * 2 + 1)].value = str(i * 2 + 1 + 1)
                        exel_font_set(s9['A' + str(9 + i * 2 + 1)])
                        s9['B' + str(9 + 2 * i)].value = opername
                        exel_font_set(s9['B' + str(9 + i)])
                        s9['B' + str(9 + 2 * i + 1)].value = opername
                        exel_font_set(s9['B' + str(9 + 2 * i + 1)])
                        s9['C' + str(9 + 2 * i)].value = 'XML'
                        exel_font_set(s9['C' + str(9 + 2 * i)])
                        s9['C' + str(9 + 2 * i + 1)].value = 'JSON'
                        exel_font_set(s9['C' + str(9 + 2 * i + 1)])
                        s9['D' + str(9 + 2 * i)].value = str(2 * i + 1)
                        exel_font_set(s9['D' + str(9 + 2 * i)])
                        s9['D' + str(9 + 2 * i + 1)].value = str(2 * i + 1 + 1)
                        exel_font_set(s9['D' + str(9 + 2 * i + 1)])
                        s9['E' + str(9 + 2 * i)].value = ILLEGAL_CHARACTERS_RE.sub(r'', req_html2[:1000])
                        exel_font_set(s9['E' + str(9 + 2 * i)])
                        s9['E' + str(9 + 2 * i + 1)].value = ILLEGAL_CHARACTERS_RE.sub(r'', req_html3[:1000])
                        exel_font_set(s9['E' + str(9 + 2 * i + 1)])
                        if '이미지' in req_html2:
                            s9['F' + str(9 + 2 * i)].value = '정상'
                            result_list7.append('정상')
                        elif '</' in req_html2[:1000]:
                            s9['F' + str(9 + 2 * i)].value = '정상'
                            result_list7.append('정상')
                        else:
                            s9['F' + str(9 + 2 * i)].value = '오류'
                            result_list7.append('오류')
                        if '이미지' in req_html3:
                            s9['F' + str(9 + 2 * i + 1)].value = '정상'
                            result_list7.append('정상')
                        elif '{' in req_html3[:1000]:
                            s9['F' + str(9 + 2 * i + 1)].value = '정상'
                            result_list7.append('정상')
                        else:
                            s9['F' + str(9 + 2 * i + 1)].value = '오류'
                            result_list7.append('오류')

                        exel_font_set(s9['F' + str(9 + 2 * i + 1)])
                        exel_font_set(s9['F' + str(9 + 2 * i)])

                if not opername == '문서오류':

                    s10cnt = 0
                    ##8번 항목
                    s10 = wb_result['80.S-8']
                    for chek in range(0, operation_len):
                        s10cnt = s10cnt + 1
                        s10['A' + str(9 + chek)].value = s10cnt
                        s10['B' + str(9 + chek)].value = opername_list[chek]
                        s10['C' + str(9 + chek)].value = result_list1[chek]
                        s10['D' + str(9 + chek)].value = result_list2[chek]
                        s10['E' + str(9 + chek)].value = result_list4[chek]
                        if s10['C' + str(9 + chek)].value == '오류' or s10['D' + str(9 + chek)].value == '오류' or s10[
                            'E' + str(9 + chek)].value == '오류':
                            s10['F' + str(9 + chek)].value = '오류'
                        else:
                            s10['F' + str(9 + chek)].value = '정상'

                        exel_font_set(s10['A' + str(9 + chek)])
                        exel_font_set(s10['B' + str(9 + chek)])
                        exel_font_set(s10['C' + str(9 + chek)])
                        exel_font_set(s10['D' + str(9 + chek)])
                        exel_font_set(s10['E' + str(9 + chek)])
                        exel_font_set(s10['F' + str(9 + chek)])
                        result_list8.append(s10['F' + str(9 + chek)].value)

                    ##     # s['A'+str(8 + j)]
                    ##수준평가
                    s1 = wb_result['00.수준평가']
                    s1['C5'].value = organ_name
                    s1['C6'].value = api_name
                    s1['C7'].value = api_type
                    s1['C8'].value = dataformmat
                    s1['C9'].value = operation_len
                    s1['C10'].value = docxname
                    s1['C11'].value = api_portal_url
                    s1['D14'].value = operation_len
                    s1['D15'].value = operation_len
                    s1['D16'].value = operation_len
                    s1['D17'].value = operation_len
                    s1['D18'].value = operation_len

                    s1['E14'].value = operation_len - result_list1.count('오류')
                    s1['F15'].value = operation_len - result_list2.count('오류')
                    s1['E16'].value = operation_len - result_list3.count('오류')
                    s1['F17'].value = operation_len - result_list4.count('오류')
                    s1['E18'].value = operation_len
                    s1['E18'].value = operation_len - result_list5.count('오류')

                    if len(result_list6) != 0:
                        s1['D19'].value = operation_len
                        s1['E19'].value = operation_len - result_list5.count('미제공')

                    s2 = wb_result['02.추가 진단']
                    s2['C5'].value = organ_name
                    s2['C6'].value = api_name
                    s2['C7'].value = api_type
                    s2['C8'].value = dataformmat
                    s2['C9'].value = operation_len
                    s2['C10'].value = docxname
                    s2['C11'].value = api_portal_url

                    if dataformmat == 'XML' or dataformmat == 'JSON' or str(dataformmat) == 'nan':
                        s2['E14'].value = operation_len
                        s2['F14'].value = operation_len - result_list7.count('오류')
                    else:
                        s2['E14'].value = operation_len
                        minus_cnt = 0
                        flag = True
                        for i in range(0, operation_len * 2):
                            if i % 2 == 0:
                                if result_list7[i] == '오류':
                                    minus_cnt = minus_cnt + 1
                                    flag = False
                                    continue
                            elif result_list7[i] == '오류' and flag:
                                minus_cnt = minus_cnt + 1
                            flag = True
                        s2['F14'].value = operation_len - minus_cnt

                    s2['D15'].value = operation_len
                    s2['F15'].value = operation_len - result_list8.count('오류')

                    shutil.move(data, './complete/' + fname)
                wb_result.save(result)
                self.parent().label2.setText("보고서 생성 완료")
                self.parent().label2.repaint()
                print(api_name)

        except Exception as e:
            traceback.print_exc()
            self.parent().label2.setText("에러메세지: " + str(e))
            self.parent().label2.repaint()
            if os.path.isfile(result_path + result_file):
                os.remove(result_path + result_file)
            return None



class MyApp(QWidget):

    def __init__(self):
        super().__init__()
        self.initUI()

    def initUI(self):

        self.grid = QGridLayout()
        self.setLayout(self.grid)

        self.pushButton = QPushButton("진단 시작", self)
        self.pushButton.clicked.connect(self.pushButtonClicked)
        self.label = QLabel()
        self.label2 = QLabel()

        self.grid.addWidget(self.pushButton, 0,0)
        self.grid.addWidget(self.label, 1, 0)
        self.grid.addWidget(self.label2, 3, 0)


        self.setWindowTitle('오픈API 보고서 작성v1.2')
        self.setGeometry(300, 100, 600, 300)

        self.th = Thread1(parent=self)

        self.setAcceptDrops(True)
        self.show()

    @pyqtSlot()
    def pushButtonClicked(self):
        self.th.start()
        self.th.working = True

    @pyqtSlot()
    def time_stop(self):
        self.th.working = False





if __name__ == '__main__':
    app = QApplication(sys.argv)
    ex = MyApp()
    sys.exit(app.exec_())




