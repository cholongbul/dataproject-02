import os
import shutil
import openpyxl

resourcepath = 'Z:\\104.DB적재소스\\리소스모음\\리소스\\'
resourcelist = os.listdir(resourcepath)
mergepath = 'C:\\Users\\admin\\Desktop\\리소스오류정리\\셀병합오류\\'
anothorcd = 'C:\\Users\\admin\\Desktop\\리소스오류정리\\리소스형식\\'
reserro = 'C:\\Users\\admin\\Desktop\\리소스오류정리\\응답에러\\'
error = 'C:\\Users\\admin\\Desktop\\리소스오류정리\\문서오류\\'
for resource in resourcelist:
    try:
        print(resource)
        wb_data = openpyxl.load_workbook(resourcepath + resource)
        sheetnames = wb_data.sheetnames
        link_list = []
        for sheetname in sheetnames:
            ws_data = wb_data[sheetname]
            if len(ws_data.merged_cells.ranges) > 0:
                shutil.copy(resourcepath + resource, mergepath + resource)
            if wb_data[sheetnames[0]]['L2'].value != ws_data['L2'].value:
                shutil.copy(resourcepath + resource, anothorcd + resource)
            if str(ws_data['A2'].value).replace(' ','').startswith('문서'):
                shutil.copy(resourcepath + resource, error + resource)
            if ws_data['N3'].value != None:
                shutil.copy(resourcepath + resource, reserro + resource)
    except:
        pass

            


