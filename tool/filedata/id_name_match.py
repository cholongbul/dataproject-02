import os
import traceback

import openpyxl
import pandas as pd
from copy import copy



def change(ws_data, ex_i, id_col, name_col, id_name_df):
    while True:
        print(ex_i)
        if ex_i + 2 == 1048577:
            break
        elif ws_data[name_col + str(ex_i + 2)].value == None:
            break
        else:
            ws_data[id_col + str(ex_i + 2)].value = str(ws_data[name_col + str(ex_i + 2)].value).replace('.csv', '').replace('-','').replace(' ','')
            is_id_name = id_name_df['id'] == str(ws_data[name_col + str(ex_i + 2)].value).replace('.csv', '').replace('-','').replace(' ','')
            ws_data[name_col + str(ex_i + 2)].value = id_name_df[is_id_name]['name'].tolist()[0]
            ex_i = ex_i + 1

def change2(ws_data, ex_i, id_name_df):
    print(ex_i)
    print(ws_data['C' + str(ex_i + 2)].value)
    ws_data['D' + str(ex_i + 1)].font = copy(ws_data['C' + str(ex_i + 1)].font)
    ws_data['D' + str(ex_i + 1)].border = copy(ws_data['C' + str(ex_i + 1)].border)
    ws_data['D' + str(ex_i + 1)].fill = copy(ws_data['C' + str(ex_i + 1)].fill)
    ws_data['D' + str(ex_i + 1)].protection = copy(ws_data['C' + str(ex_i + 1)].protection)
    ws_data['D' + str(ex_i + 1)].alignment = copy(ws_data['C' + str(ex_i + 1)].alignment)
    ws_data.column_dimensions['D'].width = copy(ws_data.column_dimensions['C'].width)

    ws_data['C' + str(ex_i + 1)].value = '목록명'
    ws_data['D' + str(ex_i + 1)].value = '개방데이터파일명'
    while True:
        if ws_data['C' + str(ex_i + 2)].value == None:
            break
        else:
            ws_data['D' + str(ex_i + 2)].font = copy(ws_data['C' + str(ex_i + 2)].font)
            ws_data['D' + str(ex_i + 2)].border = copy(ws_data['C' + str(ex_i + 2)].border)
            ws_data['D' + str(ex_i + 2)].fill = copy(ws_data['C' + str(ex_i + 2)].fill)
            ws_data['D' + str(ex_i + 2)].protection = copy(ws_data['C' + str(ex_i + 2)].protection)
            ws_data['D' + str(ex_i + 2)].alignment = copy(ws_data['C' + str(ex_i + 2)].alignment)

            ws_data['B' + str(ex_i + 2)].value = str(ws_data['C' + str(ex_i + 2)].value).replace('.csv', '').replace('-','').replace(' ','')
            is_id_name = id_name_df['id'] == str(ws_data['C' + str(ex_i + 2)].value).replace('.csv', '').replace('-','').replace(' ','')
            ws_data['D' + str(ex_i + 2)].value = id_name_df[is_id_name]['name'].tolist()[0]
            ws_data['C' + str(ex_i + 2)].value = id_name_df[is_id_name]['list'].tolist()[0]
            ex_i = ex_i + 1


path = 'C:\\Users\\admin\\Documents\\2.개방데이터\\22.개방데이터정리\\1000.3차제외파일 진단보고서\\'
file_list = os.listdir(path)
id_name_df = pd.read_csv('./id_name.csv')
def change3():
    try:
        for file in file_list:

            wb_data = openpyxl.load_workbook(path + file)
            sheetnameslist = wb_data.sheetnames
            for i in range(1,len(sheetnameslist)):
                ex_i = 0
                if i <3 :
                    ws_data = wb_data[sheetnameslist[i]]
                    while True:
                        if ws_data['B' + str(ex_i+2)].value == None:
                            break

                        is_id_name = id_name_df['id'] == str(ws_data['B' + str(ex_i+2)].value)
                        ws_data['C' +  str(ex_i+2)].value = id_name_df[is_id_name]['name'].tolist()[0]
                        ex_i = ex_i + 1
                else:
                    ws_data = wb_data[sheetnameslist[i]]
                    while True:
                        if ws_data['C' + str(ex_i+2)].value == None:
                            break
                        elif str(ws_data['C' +str(ex_i+2)].value).startswith('F2'):
                            ws_data['C' + str(ex_i+2)].value = str(ws_data['D' +str(ex_i+2)].value).replace('.csv', '')
                            is_id_name = id_name_df['id'] == str(ws_data['D' + str(ex_i+2)].value).replace('.csv',
                                                                                                             '').replace(
                                '-', '').replace(' ', '')
                            ws_data['D' +str(ex_i+2)].value = id_name_df[is_id_name]['name'].tolist()[0]
                        ex_i = ex_i + 1
            wb_data.save(path+file)
    except:
        traceback.print_exc()
        exit()

def change4():
    path = 'C:\\Users\\admin\\Documents\\2.개방데이터\\22.개방데이터정리\\1000.3차제외파일 진단보고서\\'
    file_list = os.listdir(path)
    id_name_df = pd.read_csv('./id_name.csv')
    try:
        for file in file_list:
            print(file)

            wb_data = openpyxl.load_workbook(path + file)
            sheetnameslist = wb_data.sheetnames
            for i in range(1, len(sheetnameslist)):
                ex_i = 0
                if i == 1 or i ==2:
                    ws_data = wb_data[sheetnameslist[i]]
                    while True:
                        r_id = ws_data['B' + str(ex_i + 2)].value
                        if ws_data['B' + str(ex_i + 2)].value == None:
                            break

                        is_id_name = id_name_df['id'] == str(ws_data['B' + str(ex_i + 2)].value)
                        ws_data['C' + str(ex_i + 2)].value = id_name_df[is_id_name]['name'].tolist()[0]
                        ex_i = ex_i + 1
                elif i >2:
                    ws_data = wb_data[sheetnameslist[i]]
                    while True:
                        r_id = ws_data['C' + str(ex_i + 2)].value
                        if ws_data['C' + str(ex_i + 2)].value == None:
                            break

                        is_id_name = id_name_df['id'] == str(ws_data['C' + str(ex_i + 2)].value)
                        ws_data['D' + str(ex_i + 2)].value = id_name_df[is_id_name]['name'].tolist()[0]
                        ex_i = ex_i + 1
            wb_data.save(path + file)
    except:
        traceback.print_exc()
        print(r_id)
        exit()

change4()