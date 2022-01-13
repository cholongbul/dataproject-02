import openpyxl
import os
import pandas as pd
import traceback

path = 'C:\\Users\\admin\\Documents\\2.개방데이터\\13.파일진단보고서\\'
file_list = os.listdir(path)

try:
    for file in file_list:
        print(file)

        wb_data = openpyxl.load_workbook(path + file,read_only=True)
        sheetnameslist = wb_data.sheetnames
        ws_data = wb_data[sheetnameslist[1]]
        ex_i = 0
        while True:
            if ex_i + 2 == 1048577:
                break
            elif ws_data['B' + str(ex_i + 2)].value == None:
                break
            else:
                id = ws_data['B' + str(ex_i + 2)].value
                ex_i = ex_i + 1
                w_file = open('idlist3.txt', 'a', encoding='utf8')
                w_file.write(str(id) + '\n')
                w_file.close()

except:
    traceback.print_exc()
    exit()

