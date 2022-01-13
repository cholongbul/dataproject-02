import os
import shutil

import openpyxl
from sqlalchemy import create_engine
from multiprocessing import Process, Queue
import pandas as pd
import cx_Oracle
import xlrd

##인스턴스 경로 지정
cx_Oracle.init_oracle_client(lib_dir=r"C:\\instantclient_11_2\\")

##오라클 연동
host='121.67.187.211'
port=1521
sid='orclcdb'
user='C##open1'
password='oracle-data4321'
dsn = cx_Oracle.makedsn(host, port, sid=sid)
conn = cx_Oracle.connect(user,password,dsn)
cursor = conn.cursor()
##원본파일
path = 'C:\\Users\\admin\\Downloads\\SQL_DIAG_20220105\\기관별\\행정안전부\\'
##이상파일
sqlerrorpath = 'C:\\Users\\admin\\Downloads\\SQL_DIAG_20220105\\sql이상파일\\'
files = os.listdir(path)
for file in files:
    result_wb = openpyxl.load_workbook('./SQL보고서_템플릿.xlsx')
    ws1 = result_wb['01.컬럼목록']
    ws2 = result_wb['C20000']
    tablename = file.split('.')[0]
    cursor.execute("""select 
        count(*)
        from C##OPEN1."""+tablename+"""
        """)
    x = cursor.fetchall()
    df_oracle = pd.DataFrame(x)
    datacnt = df_oracle[0][0]
    if file.endswith('xlsx') or file.endswith('xlsm'):
        wb_data = openpyxl.load_workbook(path + file)
        sheets = wb_data.sheetnames
        errorcnt_dict = {}
        for sheet in sheets:
            if sheet.startswith('C'):
                try:
                    ws = wb_data[sheet]
                    type = ws['B5'].value
                    sql = str(ws['B6'].value)
                    sql = sql.replace('C##OPENDATA','C##OPEN1').replace('\n', ' ').replace(';','').replace('_x000D_',' ')
                    if sql.lower().count('select') > 1:
                        sel_i = sql.lower().find('select',6)
                        sql = sql[:sel_i]
                    where_i = sel_i = sql.lower().find('where')
                    sql = 'select '+'count('+sheet+') from ' + 'C##OPEN1.'+tablename+' '+ sql[where_i:]
                    cursor.execute(sql)
                    x = cursor.fetchall()
                    df_oracle = pd.DataFrame(x)
                    errorcnt = df_oracle[0][0]
                    errorcnt_dict[sheet] = errorcnt
                    if int(df_oracle[0][0]) > int(datacnt)/2:
                        shutil.move(path + file, sqlerrorpath + file)
                        print(sql)
                        break
                except cx_Oracle.DatabaseError:
                    shutil.move(path + file, sqlerrorpath + file)
                    print(sql)
                    break

            ws = wb_data[sheet]
            ws_cnt = 0
            while True:
            if ws['A'+str(2+ws_cnt)].value == None:
                break
            ws1['A'+str(2+ws_cnt)].value = 2+ws_cnt
            ws1['B' + str(2 + ws_cnt)].value = tablename
            ws1['C' + str(2 + ws_cnt)].value = ws['B' + str(2 + ws_cnt)].value
            ws1['D' + str(2 + ws_cnt)].value = ws['C' + str(2 + ws_cnt)].value
            ws1['E' + str(2 + ws_cnt)].value = ws['D' + str(2 + ws_cnt)].value
            ws1['F' + str(2 + ws_cnt)].value = ws['E' + str(2 + ws_cnt)].value
            ws1['G' + str(2 + ws_cnt)].value = datacnt
            ws1['H' + str(2 + ws_cnt)].value = ws['A' + str(2 + ws_cnt)].value


        wb_data.close()

    if file.endswith('xls'):

        wb_data = xlrd.open_workbook(path + file)
        sheets = wb_data.sheet_names()
        for sheet in sheets:
            if sheet.startswith('C'):
                try:
                    ws = wb_data[sheet]
                    type = ws.cell_value(4,1)
                    sql = ws.cell_value(5,1)
                    sql = sql.replace('C##OPENDATA', 'C##OPEN1').replace('\n', ' ').replace(';', '').replace('_x000D_',' ')
                    if sql.lower().count('select') > 1:
                        sel_i = sql.lower().find('select', 6)
                        sql = sql[:sel_i]

                    where_i = sel_i = sql.lower().find('where')
                    sql = 'select '+'count('+sheet+') from ' + 'C##OPEN1.'+tablename+' '+ sql[where_i:]

                    cursor.execute(sql)
                    x = cursor.fetchall()
                    df_oracle = pd.DataFrame(x)
                    print(df_oracle[0][0])
                    if int(df_oracle[0][0]) > int(datacnt)/2:
                        shutil.move(path + file, sqlerrorpath + file)
                        print(sql)
                        break

                except cx_Oracle.DatabaseError:
                    shutil.move(path+file,sqlerrorpath+file)
                    print(sql)
                    break

#     filename = '#fromdb'
#     colname = '#fromdb'
#     itemname = '#fromdb'
#     if itemname == '문자열':
#         sqlYn = ''
#     else:
#         sqlYn = 'Y'
#
#
#     cursor.execute("""select
#     C5
#     from C##OPEN1.F100014060
#     where  "index" <> 0
#     and C5 not in ('Active', 'Failed startup', 'Grounded', 'Startup')
#     """)
# x = cursor.fetchall()
# df_oracle = pd.DataFrame(x)
# #sql입력값
# print(df_oracle)
# fileid = 'F100019491'
# sqltype_df = pd.read_csv('./sql형식.csv')
# is_same_id = sqltype_df

