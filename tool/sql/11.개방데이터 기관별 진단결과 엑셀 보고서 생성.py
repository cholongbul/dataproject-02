import os
import datetime
import cx_Oracle as ora
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Font, Color
from openpyxl.styles import Alignment
from openpyxl.styles import PatternFill
from openpyxl.styles import Border, Side
from openpyxl.worksheet.dimensions import ColumnDimension, DimensionHolder
from openpyxl.utils import get_column_letter

# Variables Define
#filePath = 'D:\\org_list\\'
filePath = "Y:\\150.품질분석_배치시스템구현_산출물\\01.기관별_품질분석_결과보고서\\"
fileName = ' '
HAIR_BORDER1 = Border(left=Side(border_style=None), right=Side(border_style=None), top=Side(border_style='hair'),
                      bottom=Side(border_style='hair'))
fc_red = Font(color="FF0000") # Red
now = datetime.datetime.now()

# DB Connection
ora.init_oracle_client(lib_dir=r"C:\\anaconda\\")
conn = ora.connect("C##BRION/1234@192.168.3.20:1521/ORCL")
# conn = ora.connect("C##BRION/1234@localhost:1521/ORCL")
curs = conn.cursor()
curs.rowfactory = lambda *args: dict(zip([d[0] for d in curs.description], args))


# Table Dictionary Referencing Define
def makeDictFactory(cursor):
    columnNames = [d[0] for d in cursor.description]

    def createRow(*args):
        return dict(zip(columnNames, args))

    return createRow

# Excel File Read
#wbs = load_workbook('D:\\기관별현황_SAMPLE.xlsx')

#print(wbs.get_sheet_names())  # old
#print(wbs.sheetnames)          # new

counter = 0
# 출력 대상 파일 목록 DB 조회
sql0 = "SELECT ORG_CD ,ORG_NM  FROM DIAG_T520_ORG_VAL_1231  ORDER BY ORG_CD "
#sql0 = "SELECT ORG_CD ,ORG_NM  FROM DIAG_T520_ORG_VAL_1231 WHERE ORG_CD IN ('1051000','1192000')  ORDER BY ORG_CD "
curs.execute(sql0)
curs.rowfactory = makeDictFactory(curs)
tbl0 = curs.fetchall()
print(tbl0)
print('Start')
for row in tbl0:
    counter = counter + 1

    print('Counter : ',counter)
    org_cd = row['ORG_CD']
    print('ORG_NM :' , row['ORG_CD'] , row['ORG_NM'])
    wbs = load_workbook('D:\\기관별현황_SAMPLE.xlsx')
    ws1 = wbs['기관별 값 진단 결과보고서']
    ws2 = wbs['대상파일목록']
    ws3 = wbs['기관별_진단규칙 및 오류목록']

    ws1['G3'] = now.strftime('%Y-%m-%d %H:%M')

    # 출력 대상 파일 목록 DB 조회
    sql1 = "SELECT  ORG_CD, ORG_NM, ORG_EYN, ORG_ERR_TERR, ORG_EGRD, ORG_TCNT,ORG_DNT,ORG_CNT,ORG_RTNT,ORG_RCNT,ORG_RSNT," \
          "ORG_ERR_RSNT,ORG_ERR_RENT ,ORG_ERR_TERR FROM DIAG_T520_ORG_VAL_1231 WHERE ORG_CD = :1  ORDER BY ORG_CD "
    curs.execute(sql1,[org_cd])
    curs.rowfactory = makeDictFactory(curs)
    tbl1 = curs.fetchall()

    # 초기값 세팅(0)
    for rng in ws1['D16:G23']:
        for cell in rng:
            cell.value = 0
    # 계산식셀 지정
    ws1.cell(23, 4).value = '=SUM(D16:D22)'  # 진단대상 컬럼 Total
    ws1.cell(23, 5).value = '=SUM(E16:E22)'  # 전체건수
    ws1.cell(23, 6).value = '=SUM(F16:F22)'  # 오류데이터 건수
    ws1.cell(23, 7).value = '=ROUND((F23/E23) * 100,2)'  # 오류율
    # 변수셀 세팅
    for row in tbl1:
        fileName = row['ORG_NM'] + '_' + row['ORG_CD'] + '_종합진단_보고서_'  + str(row['ORG_ERR_RENT']) + '.xlsx'
        ws1['C05'] = row['ORG_NM'] + '(' + row['ORG_CD'] + ')'
        ws1['C07'] = row['ORG_EGRD']        #품질평가등급
        ws1['B10'] = row['ORG_TCNT']        #전체테이블갯수
        ws1['C10'] = row['ORG_RTNT']        #룰적용테이블갯수
        ws1['D10'] = row['ORG_DNT']         #기관총테이블건수
        ws1['E10'] = row['ORG_RSNT']        #룰적용데이터건수(Cell)
        ws1['F10'] = row['ORG_CNT']         #기관전체컬럼수
        ws1['G10'] = row['ORG_RCNT']        #룰적용컬럼수
        ws1['B13'] = row['ORG_ERR_RENT']    #룰적용오류데이터건수
        ws1['E13'] = row['ORG_ERR_TERR']    #기관 오류율

    sql2 = "SELECT COL_PAT_LCD, COL_PAT_LNM AS 검증유형, SUM(PAT_RCNT) AS 진단대상컬럼, SUM(PAT_RSNT) AS 전체데이터, SUM(PAT_RENT) AS 오류데이터 FROM DIAG_T502_TAB_PAT_1231 WHERE ORG_CD = :1 GROUP BY COL_PAT_LNM,COL_PAT_LCD"
    curs.execute(sql2, [org_cd])
    curs.rowfactory = makeDictFactory(curs)
    tbl2 = curs.fetchall()
    for row in tbl2:
        if row['COL_PAT_LCD'] == '02':    # 숫자형
            ws1['D16'] = row['진단대상컬럼']
            ws1['E16'] = row['전체데이터']
            ws1['F16'] = row['오류데이터']
            ws1.cell(16, 7).value = '=ROUND((F16/E16) * 100,2)'  # 오류율
        if row['COL_PAT_LCD'] == '06':  # 비율형
            ws1['D18'] = row['진단대상컬럼']
            ws1['E18'] = row['전체데이터']
            ws1['F18'] = row['오류데이터']
            ws1.cell(18, 7).value = '=ROUND((F18/E18) * 100,2)'  # 오류율
        if row['COL_PAT_LCD'] == '03':    # 날짜형
            ws1['D20'] = row['진단대상컬럼']
            ws1['E20'] = row['전체데이터']
            ws1['F20'] = row['오류데이터']
            ws1.cell(20, 7).value = '=ROUND((F20/E20) * 100,2)'  # 오류율
        if row['COL_PAT_LCD'] == '04':    # 번호형
            ws1['D21'] = row['진단대상컬럼']
            ws1['E21'] = row['전체데이터']
            ws1['F21'] = row['오류데이터']
            ws1.cell(21, 7).value = '=ROUND((F21/E21) * 100,2)'  # 오류율
        if row['COL_PAT_LCD'] == '05':    # 여부
            ws1['D19'] = row['진단대상컬럼']
            ws1['E19'] = row['전체데이터']
            ws1['F19'] = row['오류데이터']
            ws1.cell(19, 7).value = '=ROUND((F19/E19) * 100,2)'  # 오류율

    sql3 = "SELECT  ROWNUM,A1.* FROM (SELECT DISTINCT A1.TAB_ENG,A1.TOOL_FILE_ID , B1.데이터명 FROM DIAG_T501_TAB_COL_1231  A1, DIAG_TOPN_MAS_0902 B1 WHERE A1.TAB_ENG = B1.FILE_ID AND A1.ORG_CD = :1 ORDER BY TAB_ENG) A1"
    curs.execute(sql3, [org_cd])
    tbl3 = curs.fetchall()
    for row3 in tbl3:
        ws2.append(row3)
    # 센터 조정
    last_row = ws2.max_row
    for i in range(1,last_row+1):
        ws2.cell(i,1).alignment = Alignment(horizontal='center', vertical='center')
        ws2.cell(i,2).alignment = Alignment(horizontal='center', vertical='center')
        ws2.cell(i,3).alignment = Alignment(horizontal='center', vertical='center')
    # Border
    for i in range(1, last_row + 1):
        ws2.cell(i, 4).border = HAIR_BORDER1
    sql4 = "SELECT TOOL_FILE_ID,TAB_KOR,COL_NUM,COL_KOR,COL_PAT_LNM,COL_PAT_NM,COL_PAT_TYPE,ERR_SAM1,ERR_SAM2,ERR_SAM3,ERR_SAM4,ERR_SAM5 FROM DIAG_T503_COL_ERR_1231 WHERE ORG_CD = :1  AND NOT REGEXP_LIKE(ERR_SAM1, '[一-龥]')  ORDER BY TOOL_FILE_ID,COL_NUM"
    curs.execute(sql4, [org_cd])
    tbl4 = curs.fetchall()
    for row4 in tbl4:
        ws3.append(row4)
        # 센터 조정
    last_row = ws3.max_row
    for i in range(1, last_row + 1):
        ws3.cell(i, 1).alignment = Alignment(horizontal='center', vertical='center')
        ws3.cell(i, 3).alignment = Alignment(horizontal='center', vertical='center')
        ws3.cell(i, 5).alignment = Alignment(horizontal='center', vertical='center')
        ws3.cell(i, 6).alignment = Alignment(horizontal='center', vertical='center')
        ws3.cell(i, 7).alignment = Alignment(horizontal='center', vertical='center')
        ws3.cell(i, 8).alignment = Alignment(horizontal='center', vertical='center')
        ws3.cell(i, 9).alignment = Alignment(horizontal='center', vertical='center')
        ws3.cell(i,10).alignment = Alignment(horizontal='center', vertical='center')
        ws3.cell(i,11).alignment = Alignment(horizontal='center', vertical='center')
        ws3.cell(i,12).alignment = Alignment(horizontal='center', vertical='center')
    # 오류셀 Red_fill
    for row in range(2,last_row + 1):
        for col in range(8, 13):
            if ws3.cell(row,col).value != '오류없음':
                ws3.cell(row, col).font = fc_red
    wbs.save(filePath + fileName)
curs.close()




