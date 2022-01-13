import os
import datetime
import cx_Oracle as ora
from openpyxl import Workbook
from openpyxl.styles import Font, Color
from openpyxl.styles import Alignment
from openpyxl.styles import PatternFill
from openpyxl.styles import Border, Side
from openpyxl.worksheet.dimensions import ColumnDimension, DimensionHolder
from openpyxl.utils import get_column_letter

# Variables Define
file_name = ''
now = datetime.datetime.now()
# DB Connection
ora.init_oracle_client(lib_dir=r"C:\\anaconda\\")
conn = ora.connect("C##BRION/1234@192.168.3.20:1521/ORCL")
# conn = ora.connect("C##BRION/1234@localhost:1521/ORCL")
curs = conn.cursor()
curs.rowfactory = lambda *args: dict(zip([d[0] for d in curs.description], args))


# Table Dictionary Referencing Define
def makeDictFactory(cursor):
    columnNames = [d[0] for d in cursor.description]

    def createRow(*args):
        return dict(zip(columnNames, args))

    return createRow


# Excel Variable Define
# font_sub = Font(name="맑은 고딕",size=11,bold=True,color="FFFFFF")
fc_red = Font(color="FF0000") # Red
g_color = 'C0C0C0'  # Gray  옅은 Gray = '808080'
b_color = '392f31'  # 짙은 Gray
g_fill = PatternFill(start_color=g_color, end_color=g_color, fill_type='solid')
b_fill = PatternFill(start_color=b_color, end_color=b_color, fill_type='solid')
align = Alignment(horizontal='center', vertical='center')
HAIR_BORDER1 = Border(left=Side(border_style=None), right=Side(border_style=None), top=Side(border_style='hair'),
                      bottom=Side(border_style='hair'))
THIN_BORDER2 = Border(left=Side(border_style=None), right=Side(border_style=None), top=Side(border_style='thin'),
                      bottom=Side(border_style='thin'))

# Excel Frame Define
wbk = Workbook()
ws1 = wbk.active
# Sheet Name Define
ws1.title = "개방데이터(파일) 값 진단 결과보고서"

#####################################################
# Main Logic
#####################################################

#################################
# Source Path Dir
#path_in_dir = 'D:/tab_1000.txt'
path_in_dir = 'D:/tab_1.txt'
##################################
# Target Table Read
f = open(path_in_dir, 'r')
lines = f.readlines()
# print('lines :', lines)  # ['FILE_000000002262121\n']
counter = 0
# 출력 대상 파일 목록 DB 조회
sql0 = "SELECT DISTINCT TAB_ENG  FROM DIAG_T502_TAB_PAT_1231 WHERE TAB_ENG NOT IN ('FILE_000000002336826')  ORDER BY TAB_ENG "

curs.execute(sql0)
curs.rowfactory = makeDictFactory(curs)
tbl0 = curs.fetchall()
print('start')
for row in tbl0:
    print(row)
    counter = counter + 1
    print('Counter :', counter)
    tab_eng = (row['TAB_ENG'])
    print(tab_eng)
    # Sheet1 : '개방데이터(파일) 값 진단 결과보고서' DB 조회
    sql1 = "SELECT ORG_CD,ORG_NM,FILE_NM,COL_PAT_LCD,COL_PAT_LNM,COL_CNT,PAT_RCNT,PAT_RSNT,PAT_RENT,PAT_RERR," \
           "DAT_NM,TAB_ENG,TOOL_FILE_ID ,TAB_RENT FROM DIAG_T502_TAB_PAT_1231 WHERE TAB_ENG = :1 " \
           "ORDER BY TAB_ENG,COL_PAT_LCD "
    curs.execute(sql1, [tab_eng])
    curs.rowfactory = makeDictFactory(curs)
    tbl1 = curs.fetchall()
    print('tbl1= : ',tbl1)
    wbk = Workbook()
    ws1 = wbk.active
    ws1.title = "개방데이터(파일) 값 진단 결과보고서"
    # Border Define
    for rng in ws1['B4:G8']:
        for cell in rng:
            cell.border = HAIR_BORDER1
    for rng in ws1['B11:G18']:
        for cell in rng:
            cell.border = HAIR_BORDER1
    for rng in ws1['B18:G18']:
        for cell in rng:
            cell.border = THIN_BORDER2
    # 센터 조정
    ws1['B02'].alignment = Alignment(horizontal='center', vertical='center')
    ws1['F03'].alignment = Alignment(horizontal='right', vertical='center')
    ws1['F04'].alignment = Alignment(horizontal='right', vertical='center')
    ws1['B04'].alignment = Alignment(horizontal='center', vertical='center')
    ws1['B05'].alignment = Alignment(horizontal='center', vertical='center')
    ws1['B07'].alignment = Alignment(horizontal='center', vertical='center')
    ws1['B08'].alignment = Alignment(horizontal='center', vertical='center')
    ws1['B11'].alignment = Alignment(horizontal='center', vertical='center')
    ws1['B18'].alignment = Alignment(horizontal='center', vertical='center')
    ws1['C04'].alignment = Alignment(horizontal='center', vertical='center')
    ws1['C05'].alignment = Alignment(horizontal='center', vertical='center')
    ws1['C07'].alignment = Alignment(horizontal='center', vertical='center')
    ws1['C08'].alignment = Alignment(horizontal='right', vertical='center')
    ws1['C10'].alignment = Alignment(horizontal='center', vertical='center')

    ws1['C11'].alignment = Alignment(horizontal='center', vertical='center')
    ws1['C12'].alignment = Alignment(horizontal='center', vertical='center')
    ws1['C13'].alignment = Alignment(horizontal='center', vertical='center')
    ws1['C14'].alignment = Alignment(horizontal='center', vertical='center')
    ws1['C15'].alignment = Alignment(horizontal='center', vertical='center')
    ws1['C16'].alignment = Alignment(horizontal='center', vertical='center')
    ws1['C17'].alignment = Alignment(horizontal='center', vertical='center')

    ws1['B10'].alignment = Alignment(horizontal='center', vertical='center')
    ws1['D10'].alignment = Alignment(horizontal='center', vertical='center')
    ws1['E10'].alignment = Alignment(horizontal='center', vertical='center')
    ws1['F10'].alignment = Alignment(horizontal='center', vertical='center')
    ws1['G10'].alignment = Alignment(horizontal='center', vertical='center')

    ws1['D11'].alignment = Alignment(horizontal='right', vertical='center')
    ws1['D12'].alignment = Alignment(horizontal='right', vertical='center')
    ws1['D13'].alignment = Alignment(horizontal='right', vertical='center')
    ws1['D14'].alignment = Alignment(horizontal='right', vertical='center')
    ws1['D15'].alignment = Alignment(horizontal='right', vertical='center')
    ws1['D16'].alignment = Alignment(horizontal='right', vertical='center')
    ws1['D17'].alignment = Alignment(horizontal='right', vertical='center')
    # for num in range(11, 18):
    #    ws1.cell(num,5).fill = b_fill
    ws1['E11'].alignment = Alignment(horizontal='right', vertical='center')
    ws1['E12'].alignment = Alignment(horizontal='right', vertical='center')
    ws1['E13'].alignment = Alignment(horizontal='right', vertical='center')
    ws1['E14'].alignment = Alignment(horizontal='right', vertical='center')
    ws1['E15'].alignment = Alignment(horizontal='right', vertical='center')
    ws1['E16'].alignment = Alignment(horizontal='right', vertical='center')
    ws1['E17'].alignment = Alignment(horizontal='right', vertical='center')

    ws1['F11'].alignment = Alignment(horizontal='right', vertical='center')
    ws1['F12'].alignment = Alignment(horizontal='right', vertical='center')
    ws1['F13'].alignment = Alignment(horizontal='right', vertical='center')
    ws1['F14'].alignment = Alignment(horizontal='right', vertical='center')
    ws1['F15'].alignment = Alignment(horizontal='right', vertical='center')
    ws1['F16'].alignment = Alignment(horizontal='right', vertical='center')
    ws1['F17'].alignment = Alignment(horizontal='right', vertical='center')

    ws1['G03'].alignment = Alignment(horizontal='right', vertical='center')
    ws1['G11'].alignment = Alignment(horizontal='right', vertical='center')
    ws1['G12'].alignment = Alignment(horizontal='right', vertical='center')
    ws1['G13'].alignment = Alignment(horizontal='right', vertical='center')
    ws1['G14'].alignment = Alignment(horizontal='right', vertical='center')
    ws1['G15'].alignment = Alignment(horizontal='right', vertical='center')
    ws1['G16'].alignment = Alignment(horizontal='right', vertical='center')
    ws1['G17'].alignment = Alignment(horizontal='right', vertical='center')

    ws1.cell(4, 2).fill = b_fill
    ws1.cell(row=2, column=2).font = Font(size=12, bold=True)

    ws1.cell(5, 2).fill = g_fill
    ws1.cell(7, 2).fill = g_fill
    ws1.cell(8, 2).fill = g_fill
    ws1.cell(11, 2).fill = g_fill
    for rng in ws1['B18:G18']:
        for cell in rng:
            cell.fill = g_fill

    # 셀병합
    ws1.merge_cells('B02:G02')  # '개방데이터(파일) 값 진단 종합 현황'
    ws1.merge_cells('B04:G04')  # '진단 데이터베이스 기본 정보'
    ws1.merge_cells('B05:B06')  # '기관명'
    ws1.merge_cells('C05:G06')  # 기관명 변수값
    ws1.merge_cells('C07:G07')  # 파일명 변수값
    ws1.merge_cells('C08:G08')  # 전체컬럼수
    ws1.merge_cells('B11:B17')  # '유효성 진단'
    ws1.merge_cells('B18:C18')  # '전체'
    # 셀 폭조정
    ws1.column_dimensions['A'].width = 2
    ws1.column_dimensions['B'].width = 20
    ws1.column_dimensions['C'].width = 15
    ws1.column_dimensions['D'].width = 15
    ws1.column_dimensions['E'].width = 15
    ws1.column_dimensions['F'].width = 15
    ws1.column_dimensions['G'].width = 20
    ################################################
    # Cell Frame Define
    ################################################
    ws1['B2'] = '개방데이터(파일) 값 진단 종합 현황'
    ws1['F3'] = '출력일'
    ws1['G3'] = now.strftime('%Y-%m-%d %H:%M')

    ws1.cell(row=4, column=2).font = Font(size=11, bold=True, color="FFFFFF")
    ws1['B04'] = '진단 데이터베이스 기본 정보'
    ws1['B05'] = '기관명'
    ws1['B07'] = '진단대상 파일명'
    ws1['B08'] = '진단대상 전체컬럼수'
    ws1['B10'] = '분석영역'
    ws1.cell(10, 2).fill = b_fill
    ws1.cell(row=10, column=2).font = Font(size=11, bold=True, color="FFFFFF")
    ws1.cell(10, 3).fill = b_fill
    ws1.cell(row=10, column=3).font = Font(size=11, bold=True, color="FFFFFF")
    ws1.cell(10, 4).fill = b_fill
    ws1.cell(row=10, column=4).font = Font(size=11, bold=True, color="FFFFFF")
    ws1.cell(10, 5).fill = b_fill
    ws1.cell(row=10, column=5).font = Font(size=11, bold=True, color="FFFFFF")
    ws1.cell(10, 6).fill = b_fill
    ws1.cell(row=10, column=6).font = Font(size=11, bold=True, color="FFFFFF")
    ws1.cell(10, 7).fill = b_fill
    ws1.cell(row=10, column=7).font = Font(size=11, bold=True, color="FFFFFF")
    ws1['C10'] = '검증유형'
    ws1['D10'] = '진단대상 컬럼'
    ws1['E10'] = '전체데이터'
    ws1['F10'] = '오류데이터'
    ws1['G10'] = '오류율(%)'
    ws1['B11'] = '유효성 진단'
    ws1['C11'] = '금액'
    ws1['C12'] = '수량'
    ws1['C13'] = '율'
    ws1['C14'] = '여부'
    ws1['C15'] = '날짜'
    ws1['C16'] = '번호'
    ws1['C17'] = '코드'
    ws1['B18'] = '전체'
    # 변수셀 초기값 세팅(0)
    for rng in ws1['D11:G18']:
        for cell in rng:
            cell.value = 0
    # 계산식셀 지정
    ws1.cell(18, 4).value = '=SUM(D11:D17)'  # 진단대상 컬럼 Total
    ws1.cell(18, 5).value = '=SUM(E11:E17)'  # 전체건수
    ws1.cell(18, 6).value = '=SUM(F11:F17)'  # 오류데이터 건수
    err_sum = ws1['G11'].value + ws1['G11'].value + ws1['G12'].value + ws1['G13'].value + ws1['G14'].value + ws1['G15'].value + ws1['G16'].value + ws1['G17'].value
    if err_sum == 0:
        ws1['G18'].value = 0
    else:
        ws1.cell(18, 7).value = '=ROUND((F18/E18) * 100,2)'  # 오류율

    # 변수셀 세팅
    for row in tbl1:
        print('sql1_row: ', row)
        file_name = row['FILE_NM'] + '_' + str(row['TAB_RENT'])  # 엑셀파일명
        print('file_name:', file_name)  # 파일명 출력
        ws1['C05'] = row['ORG_NM']  # 기관명
        ws1['C07'] = row['DAT_NM']  # 데이터명
        ws1['C08'] = row['COL_CNT']  # 컬럼수
        ws1['B21'] = row['TAB_ENG']  # 검증용
        ws1['C21'] = row['TOOL_FILE_ID']  # 검증용
        if row['COL_PAT_LCD'] == '02':  # 숫자형
            ws1['D11'] = row['PAT_RCNT']
            ws1['E11'] = row['PAT_RSNT']
            ws1['F11'] = row['PAT_RENT']
            ws1['G11'] = row['PAT_RERR']
        if row['COL_PAT_LCD'] == '06':  # 비율형
            ws1['D13'] = row['PAT_RCNT']
            ws1['E13'] = row['PAT_RSNT']
            ws1['F13'] = row['PAT_RENT']
            ws1['G13'] = row['PAT_RERR']
        if row['COL_PAT_LCD'] == '03':  # 날짜형
            ws1['D15'] = row['PAT_RCNT']
            ws1['E15'] = row['PAT_RSNT']
            ws1['F15'] = row['PAT_RENT']
            ws1['G15'] = row['PAT_RERR']
        if row['COL_PAT_LCD'] == '04':  # 번호형
            ws1['D16'] = row['PAT_RCNT']
            ws1['E16'] = row['PAT_RSNT']
            ws1['F16'] = row['PAT_RENT']
            ws1['G16'] = row['PAT_RERR']
        if row['COL_PAT_LCD'] == '05':  # 여부형
            ws1['D14'] = row['PAT_RCNT']
            ws1['E14'] = row['PAT_RSNT']
            ws1['F14'] = row['PAT_RENT']
            ws1['G14'] = row['PAT_RERR']
        # 숫자 포맷 지정(천단위 콤마)
        ws1['D18'].number_format = '#,##0'
        ws1['E18'].number_format = '#,##0'
        for rng in ws1['D11:F17']:
            for cell in rng:
                cell.number_format = '#,##0'
    # Sheet2 : 진단규칙 및 오류목록 DB 조회
    sql2 = "SELECT TOOL_FILE_ID,COL_NUM,COL_KOR,COL_PAT_LNM,COL_PAT_TYPE,COL_PAT_NM,ERR_SAM1,ERR_SAM2,ERR_SAM3," \
           "ERR_SAM4,ERR_SAM5  FROM DIAG_T503_COL_ERR_1231 WHERE TAB_ENG = :1 ORDER BY TOOL_FILE_ID,COL_NUM "
    curs.execute(sql2, [tab_eng])
    #    curs.rowfactory = makeDictFactory(curs)
    tbl2 = curs.fetchall()
    ws2 = wbk.create_sheet()
    # Sheet2 Frame Define
    ws2.title = "진단규칙 및 오류목록"
    ws2['A1'] = '개방데이터ID'
    ws2['B1'] = '컬럼순번'
    ws2['C1'] = '컬럼명'
    ws2['D1'] = '검증유형'
    ws2['E1'] = '검증유형 상세'
    ws2['F1'] = '검증유형 설명'
    ws2['G1'] = '오류샘플 1'
    ws2['H1'] = '오류샘플 2'
    ws2['I1'] = '오류샘플 3'
    ws2['J1'] = '오류샘플 4'
    ws2['K1'] = '오류샘플 5'
    # 배경색 지정
    for rng in ws2['A1:K1']:
        for cell in rng:
            cell.fill = g_fill
    # 셀 폭조정
    for col in range(1, 12):  # col_names = [row[0] for row in cur.description]
        ws2.column_dimensions[get_column_letter(col)].width = 15
    #    print('tbl2:', tbl2)
    for row in tbl2:
        #     print('sql2_row :', row)
        ws2.append(row)
    for row in range(1, 300):
        for col in range(1, 12):
            ws2.cell(row, col).alignment = align
    for rng in ws2['A01:K200']:
        for cell in rng:
            cell.border = HAIR_BORDER1
    for rng in ws2['C02:C200']:
        for cell in rng:
            cell.alignment = Alignment(horizontal='left', vertical='center')
    ws2.column_dimensions['C'].width = 30
    for row in range(2, 301):
        for col in range(7, 12):
            if ws2.cell(row, col).value != '오류없음':
                ws2.cell(row, col).font = fc_red
    file_path = 'Y:\\400.육안진단\\41.육안진단_From_DB_오류패턴_보수적_적용(정규패턴적용)\\' + file_name + '.xlsx'
    wbk.save(file_path)
curs.close()
f.close()
# conn.close
exit()



