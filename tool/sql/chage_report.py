import os
import traceback
import xlrd
import openpyxl
import win32com.client
import re
organ = '여성가족부'
path = 'C:\\Users\\admin\\Downloads\\SQL_DIAG_20220105\\기관별\\'+organ+'\\'
resultpath = 'C:\\Users\\admin\\Documents\\aa2\\시트정상화\\'
files = os.listdir(path)
result_wb = openpyxl.load_workbook('./결과물.xlsx')
result_ws = result_wb['Sheet1']
cnt = 1
for file in files:
    print(file)
    if file.endswith('xlsx') or file.endswith('xlsm'):


        wb_data = openpyxl.load_workbook(path + file)
        sheets = wb_data.sheetnames
        for sheet in sheets:
            if sheet.startswith('C'):
                ws = wb_data[sheet]
                table = ws['B1'].value
                type = ws['B5'].value
                sql = ws['B6'].value
                result_ws['A'+str(cnt)] = organ
                result_ws['B' + str(cnt)] = table
                result_ws['C' + str(cnt)] = sheet
                result_ws['D' + str(cnt)] = type
                result_ws['E' + str(cnt)] = sql
                cnt = cnt +1
        wb_data.close()

    if file.endswith('xls'):

        wb_data = xlrd.open_workbook(path + file)
        sheets = wb_data.sheet_names()
        for sheet in sheets:
            if sheet.startswith('C'):
                ws = wb_data[sheet]
                table = ws.cell_value(0,1)
                type = ws.cell_value(4,1)
                sql = ws.cell_value(5,1)
                result_ws['A' + str(cnt)] = organ
                result_ws['B' + str(cnt)] = table
                result_ws['C' + str(cnt)] = sheet
                result_ws['D' + str(cnt)] = type
                result_ws['E' + str(cnt)] = sql
                cnt = cnt + 1
result_wb.save('./'+organ+'.xlsx')
