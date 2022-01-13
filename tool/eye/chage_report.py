import os
import traceback

import openpyxl
import re
oddpath = 'C:\\Users\\admin\\Documents\\aa2\\시트이상\\'
resultpath = 'C:\\Users\\admin\\Documents\\aa2\\시트정상화\\'
files = os.listdir(oddpath)
for file in files:
    try:
        wb_data = openpyxl.load_workbook(oddpath + file)
        sheets = wb_data.sheetnames
        for sheet in sheets:
            if sheet.startswith('C'):
                if not re.match('^C[0-9]{1,}$',str(sheet)):
                    if ',' in sheet:
                        for col in sheet.split(','):
                            if '-' in col:
                                shetnum = int(col.split('-')[1].replace('C', '')) - int(
                                    col.split('-')[0].replace('C', '')) + 1
                                for i in range(0, shetnum):
                                    ws = wb_data[sheet]
                                    table = ws['B1'].value
                                    type = ws['B5'].value
                                    copy_ws = wb_data.copy_worksheet(ws)
                                    title = "C" + str(i + int(col.split('-')[0].replace('C', '')))
                                    copy_ws.title = title
                                    copy_ws['B3'].value = title

                                    try:
                                        if type == '숫자':
                                            copy_ws[
                                                'B6'].value = 'select \n' + title + '\nfrom C##OPENDATA.' + table + '\n' + 'where "index"<>0\n' + "and not regexp_like (" + title + ",'[0-9]');"
                                        elif type == '수량':
                                            copy_ws[
                                                'B6'].value = 'select \n' + title + '\nfrom C##OPENDATA.' + table + '\n' + 'where "index"<>0\n' + "and not regexp_like (" + title + ",'[0-9]');"
                                        elif type == '여부':
                                            copy_ws[
                                                'B6'].value = 'select \n' + title + '\nfrom C##OPENDATA.' + table + '\n' + 'where "index"<>0\n' + "and upper(" + title + ") not in ('Y','N');"
                                        elif type == '전화번호':
                                            copy_ws[
                                                'B6'].value = 'select \n' + title + '\nfrom C##OPENDATA.' + table + '\n' + 'where "index"<>0\n' + "and NOT REGEXP_LIKE(" + title + ", '[0-9|[0-9].[0-9]');"
                                        elif type == '날짜 YYYYMMDD':
                                            copy_ws[
                                                'B6'].value = 'select \n' + title + '\nfrom C##OPENDATA.' + table + '\n' + 'where "index"<>0\n' + "and not regexp_like (" + title + ", '^(19|20)\d{2}-(0[1-9]|1[012])-(0[1-9]|[12][0-9]|3[0-1])$');"
                                        elif type == 'YYYYMMDD':
                                            copy_ws[
                                                'B6'].value = 'select \n' + title + '\nfrom C##OPENDATA.' + table + '\n' + 'where "index"<>0\n' + "and not regexp_like (" + title + ", '^(19|20)\d{2}-(0[1-9]|1[012])-(0[1-9]|[12][0-9]|3[0-1])$');"
                                        elif type == 'HH:MM':
                                            copy_ws[
                                                'B6'].value = 'select \n' + title + '\nfrom C##OPENDATA.' + table + '\n' + 'where "index"<>0\n' + "and not REGEXP_LIKE(" + title + ", '^([1-9]|[01][0-9]|2[0-4])[:]([0-5][0-9])$');"
                                        elif type == 'Y, N 여부':
                                            copy_ws[
                                                'B6'].value = 'select \n' + title + '\nfrom C##OPENDATA.' + table + '\n' + 'where "index"<>0\n' + "and upper(" + title + ") not in ('Y', 'N');"
                                        else:
                                            print(file)
                                            print(col)
                                            print(type)
                                    except:
                                        traceback.print_exc()

                            else:
                                ws = wb_data[sheet]
                                table = ws['B1'].value
                                type = ws['B5'].value
                                copy_ws = wb_data.copy_worksheet(ws)
                                title = col
                                copy_ws.title = title
                                copy_ws['B3'].value = title

                                try:
                                    if type == '숫자':
                                        copy_ws[
                                            'B6'].value = 'select \n' + title + '\nfrom C##OPENDATA.' + table + '\n' + 'where "index"<>0\n' + "and not regexp_like (" + title + ",'[0-9]');"
                                    elif type == '수량':
                                        copy_ws[
                                            'B6'].value = 'select \n' + title + '\nfrom C##OPENDATA.' + table + '\n' + 'where "index"<>0\n' + "and not regexp_like (" + title + ",'[0-9]');"
                                    elif type == '여부':
                                        copy_ws[
                                                'B6'].value = 'select \n' + title + '\nfrom C##OPENDATA.' + table + '\n' + 'where "index"<>0\n' + "and upper(" + title + ") not in ('Y','N');"
                                    elif type == '전화번호':
                                        copy_ws[
                                                'B6'].value = 'select \n' + title + '\nfrom C##OPENDATA.' + table + '\n' + 'where "index"<>0\n' + "and NOT REGEXP_LIKE(" + title + ", '[0-9|[0-9].[0-9]');"
                                    elif type == '날짜 YYYYMMDD':
                                        copy_ws[
                                                'B6'].value = 'select \n' + title + '\nfrom C##OPENDATA.' + table + '\n' + 'where "index"<>0\n' + "and not regexp_like (" + title + ", '^(19|20)\d{2}-(0[1-9]|1[012])-(0[1-9]|[12][0-9]|3[0-1])$');"
                                    elif type == 'YYYYMMDD':
                                        copy_ws[
                                                'B6'].value = 'select \n' + title + '\nfrom C##OPENDATA.' + table + '\n' + 'where "index"<>0\n' + "and not regexp_like (" + title + ", '^(19|20)\d{2}-(0[1-9]|1[012])-(0[1-9]|[12][0-9]|3[0-1])$');"
                                    elif type == 'HH:MM':
                                        copy_ws[
                                                'B6'].value = 'select \n' + title + '\nfrom C##OPENDATA.' + table + '\n' + 'where "index"<>0\n' + "and not REGEXP_LIKE(" + title + ", '^([1-9]|[01][0-9]|2[0-4])[:]([0-5][0-9])$');"
                                    elif type == 'Y, N 여부':
                                        copy_ws[
                                                'B6'].value = 'select \n' + title + '\nfrom C##OPENDATA.' + table + '\n' + 'where "index"<>0\n' + "and upper(" + title + ") not in ('Y', 'N');"
                                    else:
                                        print(file)
                                        print(col)
                                        print(type)
                                except:
                                    traceback.print_exc()





                    elif '-' in sheet:
                        shetnum = int(sheet.split('-')[1].replace('C','')) - int(sheet.split('-')[0].replace('C','')) + 1
                        for i in range(0,shetnum):
                            ws = wb_data[sheet]
                            table = ws['B1'].value
                            type = ws['B5'].value
                            copy_ws = wb_data.copy_worksheet(ws)
                            title = "C"+str(i+int(sheet.split('-')[0].replace('C','')))
                            copy_ws.title = title
                            copy_ws['B3'].value = title

                            try:
                                if type == '숫자':
                                    copy_ws['B6'].value = 'select \n' + title + '\nfrom C##OPENDATA.'+table+'\n'+ 'where "index"<>0\n' +"and not regexp_like ("+title+",'[0-9]');"
                                elif type == '수량':
                                    copy_ws['B6'].value = 'select \n' + title + '\nfrom C##OPENDATA.'+table+'\n'+ 'where "index"<>0\n'  +"and not regexp_like ("+title+",'[0-9]');"
                                elif type == '여부':
                                    copy_ws[
                                                'B6'].value = 'select \n' + title + '\nfrom C##OPENDATA.'+table+'\n' + 'where "index"<>0\n' +"and upper("+title+") not in ('Y','N');"
                                elif type == '전화번호':
                                    copy_ws[
                                                'B6'].value = 'select \n' + title + '\nfrom C##OPENDATA.' + table + '\n' + 'where "index"<>0\n'   +"and NOT REGEXP_LIKE("+title+", '[0-9|[0-9].[0-9]');"
                                elif type == '날짜 YYYYMMDD':
                                    copy_ws[
                                                'B6'].value = 'select \n' + title + '\nfrom C##OPENDATA.' + table + '\n'+ 'where "index"<>0\n' +"and not regexp_like ("+title+", '^(19|20)\d{2}-(0[1-9]|1[012])-(0[1-9]|[12][0-9]|3[0-1])$');"
                                elif type == 'YYYYMMDD':
                                    copy_ws[
                                                'B6'].value = 'select \n' + title + '\nfrom C##OPENDATA.' + table + '\n' + 'where "index"<>0\n' +"and not regexp_like ("+title+", '^(19|20)\d{2}-(0[1-9]|1[012])-(0[1-9]|[12][0-9]|3[0-1])$');"
                                elif type == 'HH:MM':
                                    copy_ws[
                                                'B6'].value = 'select \n' + title + '\nfrom C##OPENDATA.' + table + '\n'   + 'where "index"<>0\n'  +"and not REGEXP_LIKE("+title+", '^([1-9]|[01][0-9]|2[0-4])[:]([0-5][0-9])$');"
                                elif type == 'Y, N 여부':
                                    copy_ws[
                                                'B6'].value = 'select \n' + title + '\nfrom C##OPENDATA.' + table + '\n' + 'where "index"<>0\n'  +"and upper("+title+") not in ('Y', 'N');"
                                else:
                                    print(file)
                                    print(sheet)
                                    print(type)
                            except:
                                traceback.print_exc()
                    try:
                        wb_data.remove(ws)
                    except:
                        traceback.print_exc()


        wb_data.save(resultpath+file)
    except:
        traceback.print_exc()
        print(sheet)


