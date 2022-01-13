import openpyxl
import os
def sheetname():
    path = 'C:\\Users\\admin\\Documents\\aa2\\D3014247\\'
    result_path = 'C:\\Users\\admin\\Documents\\8.육안\\2.이름수정\\'
    files = os.listdir(path)
    for file in files:

        print(file)
        wb_data = openpyxl.load_workbook(path + file)
        sheetnames = wb_data.sheetnames
        cnt = 0
        cnt2 = 0
        filename_dict = {}
        while True:
            if ws1['B' + str(5 + cnt)].value == None:
                break
            else:
                filename_dict[str(ws1['B' + str(5 + cnt)].value)] = str(ws1['C' + str(5 + cnt)].value)
                cnt = cnt + 1

        while True:
            if ws2['B' + str(2 + cnt2)].value == None:
                break
            else:
                ws2['B' + str(2 + cnt2)].value = filename_dict[str(ws2['B' + str(2 + cnt2)].value)]
                cnt2 = cnt2 + 1
        wb_data.save(result_path + file)
