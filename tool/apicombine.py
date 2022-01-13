import excellstyle
import os
from openpyxl.styles import Font, Border, Side, Alignment

import openpyxl
import shutil
import pandas as pd
from copy import copy

#
report_folder_list = os.listdir('./report_name/')
for report_folder in report_folder_list:
    report_list = os.listdir('./report_name/' + report_folder)
    result = shutil.copy('../document/오픈API_품질진단결과보고서(종합)_템플릿.xlsx', './report_combine/' +report_folder+ '_종합.xlsx')
    combine_wb = openpyxl.load_workbook(result)
    report_cnt = 0
    s1_cnt = 0
    s2_cnt = 0
    s3_cnt = 0
    s4_cnt = 0
    s5_cnt = 0
    s6_cnt = 0

    api_len = len(report_list)
    for report in report_list:
        wb_data = openpyxl.load_workbook('./report_name/'+report_folder+'/'+ report, data_only=True)

        ##보고서 변수 선언 정리
        estimate_data = wb_data['00.수준평가']
        organ_name = estimate_data['C5'].value
        api_name = estimate_data['C6'].value
        oper_len = estimate_data['C9'].value


        ##표지
        combine_ws_main = combine_wb['표지']
        combine_ws_main['B7'].value = "["+ organ_name+"]"


        ##서비스진단대상
        combine_ws_service = combine_wb['서비스진단대상']
        combine_ws_service['A2'].value = organ_name
        combine_ws_service['A' + str(6 + report_cnt)].value = api_name
        combine_ws_service['D2'] = api_len

        combine_ws_service['B' + str(6 + report_cnt)].value= estimate_data['D14'].value
        combine_ws_service['C' + str(6 + report_cnt)].value= estimate_data['E14'].value
        combine_ws_service['E' + str(6 + report_cnt)].value=estimate_data['D15'].value
        combine_ws_service['F' + str(6 + report_cnt)].value=estimate_data['F15'].value
        combine_ws_service['H' + str(6 + report_cnt)].value=estimate_data['D16'].value
        combine_ws_service['I' + str(6 + report_cnt)].value=estimate_data['E16'].value
        combine_ws_service['K' + str(6 + report_cnt)].value=estimate_data['D17'].value
        combine_ws_service['L' + str(6 + report_cnt)].value=estimate_data['F17'].value
        combine_ws_service['N' + str(6 + report_cnt)].value=estimate_data['D18'].value
        combine_ws_service['O' + str(6 + report_cnt)].value=estimate_data['E18'].value
        combine_ws_service['Q' + str(6 + report_cnt)].value=estimate_data['D19'].value
        combine_ws_service['R' + str(6 + report_cnt)].value=estimate_data['E19'].value

        ##s1항목
        combine_s1 = combine_wb['10.S-1']
        s1 = wb_data['10.S-1']
        s1_len = 0
        while True:
            if s1['A' + str(8+s1_len)].value == None:
                break
            s1_len = s1_len + 1
        for i in range(0,s1_len):
            combine_s1['A' + str(8 + i + s1_cnt)].value = api_name
            combine_s1['B' + str(8 + i + s1_cnt)].value = s1['A' + str(8 + i)].value
            combine_s1['C' + str(8 + i + s1_cnt)].value =s1['B' + str(8 + i)].value
            combine_s1['D' + str(8 + i + s1_cnt)].value =s1['C' + str(8 + i)].value
            combine_s1['E' + str(8 + i + s1_cnt)].value =s1['D' + str(8 + i)].value
            combine_s1['F' + str(8 + i + s1_cnt)].value =s1['E' + str(8 + i)].value
            combine_s1['G' + str(8 + i + s1_cnt)].value =s1['F' + str(8 + i)].value
            combine_s1['H' + str(8 + i + s1_cnt)].value =s1['G' + str(8 + i)].value
            combine_s1['I' + str(8 + i + s1_cnt)].value =s1['H' + str(8 + i)].value
            combine_s1['J' + str(8 + i + s1_cnt)].value =s1['I' + str(8 + i)].value
            combine_s1['K' + str(8 + i + s1_cnt)].value =s1['J' + str(8 + i)].value
            combine_s1['L' + str(8 + i + s1_cnt)].value =s1['K' + str(8 + i)].value
            combine_s1['M' + str(8 + i + s1_cnt)].value =s1['L' + str(8 + i)].value
            combine_s1['N' + str(8 + i + s1_cnt)].value =s1['M' + str(8 + i)].value
            excellstyle.exel_font_set(combine_s1['A' + str(8 + i + s1_cnt)])
            excellstyle.exel_font_set(combine_s1['B' + str(8 + i + s1_cnt)])
            excellstyle.exel_font_set(combine_s1['C' + str(8 + i + s1_cnt)])
            excellstyle.exel_font_set(combine_s1['D' + str(8 + i + s1_cnt)])
            excellstyle.exel_font_set(combine_s1['E' + str(8 + i + s1_cnt)])
            excellstyle.exel_font_set(combine_s1['F' + str(8 + i + s1_cnt)])
            excellstyle.exel_font_set(combine_s1['G' + str(8 + i + s1_cnt)])
            excellstyle.exel_font_set(combine_s1['H' + str(8 + i + s1_cnt)])
            excellstyle.exel_font_set(combine_s1['I' + str(8 + i + s1_cnt)])
            excellstyle.exel_font_set(combine_s1['J' + str(8 + i + s1_cnt)])
            excellstyle.exel_font_set(combine_s1['K' + str(8 + i + s1_cnt)])
            excellstyle.exel_font_set(combine_s1['L' + str(8 + i + s1_cnt)])
            excellstyle.exel_font_set(combine_s1['M' + str(8 + i + s1_cnt)])
            excellstyle.exel_font_set(combine_s1['N' + str(8 + i + s1_cnt)])


        s1_cnt = s1_cnt + s1_len

        ##s2항목
        combine_s2 = combine_wb['20.S-2']
        s2 = wb_data['20.S-2']
        for i in range(0,oper_len):
            combine_s2['A' + str(8 + i + s2_cnt)].value = api_name
            combine_s2['B' + str(8 + i + s2_cnt)].value = s2['A' + str(8 + i)].value
            combine_s2['C' + str(8 + i + s2_cnt)].value = s2['B' + str(8 + i)].value
            combine_s2['D' + str(8 + i + s2_cnt)].value = s2['C' + str(8 + i)].value
            combine_s2['E' + str(8 + i + s2_cnt)].value = s2['D' + str(8 + i)].value
            combine_s2['F' + str(8 + i + s2_cnt)].value = s2['E' + str(8 + i)].value
            combine_s2['G' + str(8 + i + s2_cnt)].value = s2['F' + str(8 + i)].value
            combine_s2['H' + str(8 + i + s2_cnt)].value = s2['G' + str(8 + i)].value
            combine_s2['I' + str(8 + i + s2_cnt)].value = s2['H' + str(8 + i)].value
            combine_s2['J' + str(8 + i + s2_cnt)].value = s2['I' + str(8 + i)].value
            excellstyle.exel_font_set(combine_s2['A' + str(8 + i + s2_cnt)])
            excellstyle.exel_font_set(combine_s2['B' + str(8 + i + s2_cnt)])
            excellstyle.exel_font_set(combine_s2['C' + str(8 + i + s2_cnt)])
            excellstyle.exel_font_set(combine_s2['D' + str(8 + i + s2_cnt)])
            excellstyle.exel_font_set(combine_s2['E' + str(8 + i + s2_cnt)])
            excellstyle.exel_font_set(combine_s2['F' + str(8 + i + s2_cnt)])
            excellstyle.exel_font_set(combine_s2['G' + str(8 + i + s2_cnt)])
            excellstyle.exel_font_set(combine_s2['H' + str(8 + i + s2_cnt)])
            excellstyle.exel_font_set(combine_s2['I' + str(8 + i + s2_cnt)])
            excellstyle.exel_font_set(combine_s2['J' + str(8 + i + s2_cnt)])


        s2_cnt = s2_cnt + oper_len

        ##s3항목
        combine_s3 = combine_wb['30.S-3']
        s3 = wb_data['30.S-3']
        s3_len = 0
        while True:
            if s3['A' + str(8 + s3_len)].value == None:
                break
            s3_len = s3_len + 1
        for i in range(0, s3_len):
            combine_s3['A' + str(8 + i + s3_cnt)].value = api_name
            combine_s3['B' + str(8 + i + s3_cnt)].value = s3['A' + str(8 + i)].value
            combine_s3['C' + str(8 + i + s3_cnt)].value = s3['B' + str(8 + i)].value
            combine_s3['D' + str(8 + i + s3_cnt)].value = s3['C' + str(8 + i)].value
            combine_s3['E' + str(8 + i + s3_cnt)].value = s3['D' + str(8 + i)].value
            combine_s3['F' + str(8 + i + s3_cnt)].value = s3['E' + str(8 + i)].value
            combine_s3['G' + str(8 + i + s3_cnt)].value = s3['F' + str(8 + i)].value
            combine_s3['H' + str(8 + i + s3_cnt)].value = s3['G' + str(8 + i)].value
            combine_s3['I' + str(8 + i + s3_cnt)].value = s3['H' + str(8 + i)].value
            combine_s3['J' + str(8 + i + s3_cnt)].value = s3['I' + str(8 + i)].value
            combine_s3['K' + str(8 + i + s3_cnt)].value = s3['J' + str(8 + i)].value
            combine_s3['L' + str(8 + i + s3_cnt)].value = s3['K' + str(8 + i)].value
            combine_s3['M' + str(8 + i + s3_cnt)].value = s3['L' + str(8 + i)].value
            combine_s3['N' + str(8 + i + s3_cnt)].value = s3['M' + str(8 + i)].value
            excellstyle.exel_font_set(combine_s3['A' + str(8 + i + s3_cnt)])
            excellstyle.exel_font_set(combine_s3['B' + str(8 + i + s3_cnt)])
            excellstyle.exel_font_set(combine_s3['C' + str(8 + i + s3_cnt)])
            excellstyle.exel_font_set(combine_s3['D' + str(8 + i + s3_cnt)])
            excellstyle.exel_font_set(combine_s3['E' + str(8 + i + s3_cnt)])
            excellstyle.exel_font_set(combine_s3['F' + str(8 + i + s3_cnt)])
            excellstyle.exel_font_set(combine_s3['G' + str(8 + i + s3_cnt)])
            excellstyle.exel_font_set(combine_s3['H' + str(8 + i + s3_cnt)])
            excellstyle.exel_font_set(combine_s3['I' + str(8 + i + s3_cnt)])
            excellstyle.exel_font_set(combine_s3['J' + str(8 + i + s3_cnt)])
            excellstyle.exel_font_set(combine_s3['K' + str(8 + i + s3_cnt)])
            excellstyle.exel_font_set(combine_s3['L' + str(8 + i + s3_cnt)])
            excellstyle.exel_font_set(combine_s3['M' + str(8 + i + s3_cnt)])
            excellstyle.exel_font_set(combine_s3['N' + str(8 + i + s3_cnt)])



        s3_cnt = s3_cnt + s3_len

        ##s4항목
        combine_s4 = combine_wb['40.S-4']
        s4 = wb_data['40.S-4']
        opernamelist = []
        s4_len = 0
        while True:
            if s4['B' + str(8 + s4_len)].value == None:
                break
            opernamelist.append(s4['B' + str(8 + s4_len)].value)
            s4_len = s4_len + 1

        new_opernamelist = []
        for v in opernamelist:
            if v not in new_opernamelist:
                new_opernamelist.append(v)



        opercnt1 = 8 + s4_cnt
        opercnt2 = 0
        opercnt3 = 8
        opercnt4 = 1
        for opername in new_opernamelist:
            print(api_name)
            opercnt2 = opernamelist.count(opername) + opercnt1 - 1
            combine_s4.merge_cells("E" + str(opercnt1) + ":" + 'E' + str(opercnt2))
            combine_s4.merge_cells("F" + str(opercnt1) + ":" + 'F' + str(
                opercnt2))
            combine_s4.merge_cells("G" + str(opercnt1) + ":" + 'G' + str(
                opercnt2))
            combine_s4.merge_cells("H" + str(opercnt1) + ":" + 'H' + str(
                opercnt2))

            excellstyle.style_range(combine_s4, "E" + str(opercnt1) + ":" + 'E' + str(opercnt2),border=Border(left=Side(style='thin'), right=Side(style='thin'),
                                                      top=Side(style='thin'),
                                                      bottom=Side(style='thin')))
            excellstyle.style_range(combine_s4, "F" + str(opercnt1) + ":" + 'F' + str(opercnt2),
                                    border=Border(left=Side(style='thin'), right=Side(style='thin'),
                                                  top=Side(style='thin'),
                                                  bottom=Side(style='thin')))
            excellstyle.style_range(combine_s4, "G" + str(opercnt1) + ":" + 'G' + str(opercnt2),
                                    border=Border(left=Side(style='thin'), right=Side(style='thin'),
                                                  top=Side(style='thin'),
                                                  bottom=Side(style='thin')))
            excellstyle.style_range(combine_s4, "H" + str(opercnt1) + ":" + 'H' + str(opercnt2),
                                    border=Border(left=Side(style='thin'), right=Side(style='thin'),
                                                  top=Side(style='thin'),
                                                  bottom=Side(style='thin')))

            combine_s4["E" + str(opercnt1)].value = opercnt4
            combine_s4["F" + str(opercnt1)].value = s4['E' + str(opercnt3)].value
            combine_s4["G" + str(opercnt1)].value = s4['F' + str(opercnt3)].value
            combine_s4["H" + str(opercnt1)].value = s4['G' + str(opercnt3)].value
            excellstyle.exel_font_set(combine_s4["E" + str(opercnt1)])
            excellstyle.exel_font_set(combine_s4["F" + str(opercnt1)])
            excellstyle.exel_font_set(combine_s4["G" + str(opercnt1)])
            excellstyle.exel_font_set(combine_s4["H" + str(opercnt1)])




            opercnt3 = opercnt3 + opernamelist.count(opername)
            opercnt1 = opercnt2 + 1
            opercnt4 = opercnt4 + 1

        for i in range(0, s4_len):
            combine_s4['A' + str(8 + i + s4_cnt)].value = api_name
            combine_s4['B' + str(8 + i + s4_cnt)].value = str(i+1)
            combine_s4['C' + str(8 + i + s4_cnt)].value = s4['B' + str(8 + i)].value
            combine_s4['D' + str(8 + i + s4_cnt)].value = s4['C' + str(8 + i)].value
            excellstyle.exel_font_set(combine_s4['A' + str(8 + i + s4_cnt)])
            excellstyle.exel_font_set(combine_s4['B' + str(8 + i + s4_cnt)])
            excellstyle.exel_font_set(combine_s4['C' + str(8 + i + s4_cnt)])
            excellstyle.exel_font_set(combine_s4['D' + str(8 + i + s4_cnt)])
        s4_cnt = s4_cnt + s4_len


        ##s5항목
        combine_s5 = combine_wb['50.S-5']
        s5 = wb_data['50.S-5']
        opernamelist5 = []
        s5_len = 0
        while True:
            if s5['B' + str(8 + s5_len)].value == None:
                break
            opernamelist5.append(s5['B' + str(8 + s5_len)].value)
            s5_len = s5_len + 1

        new_opernamelist5 = []
        for v in opernamelist5:
            if v not in new_opernamelist5:
                new_opernamelist5.append(v)

        opercnt1 = 8 + s5_cnt
        opercnt2 = 0
        opercnt3 = 8
        opercnt4 = 1
        for opername in new_opernamelist5:

            opercnt2 = opernamelist5.count(opername) + opercnt1 - 1

            combine_s5.merge_cells("E" + str(opercnt1) + ":" + 'E' + str(
                opercnt2))
            combine_s5.merge_cells("F" + str(opercnt1) + ":" + 'F' + str(
                opercnt2))
            combine_s5.merge_cells("G" + str(opercnt1) + ":" + 'G' + str(
                opercnt2))
            combine_s5.merge_cells("H" + str(opercnt1) + ":" + 'H' + str(
                opercnt2))

            excellstyle.style_range(combine_s5, "E" + str(opercnt1) + ":" + 'E' + str(opercnt2),
                                    border=Border(left=Side(style='thin'), right=Side(style='thin'),
                                                  top=Side(style='thin'),
                                                  bottom=Side(style='thin')))
            excellstyle.style_range(combine_s5, "F" + str(opercnt1) + ":" + 'F' + str(opercnt2),
                                    border=Border(left=Side(style='thin'), right=Side(style='thin'),
                                                  top=Side(style='thin'),
                                                  bottom=Side(style='thin')))
            excellstyle.style_range(combine_s5, "G" + str(opercnt1) + ":" + 'G' + str(opercnt2),
                                    border=Border(left=Side(style='thin'), right=Side(style='thin'),
                                                  top=Side(style='thin'),
                                                  bottom=Side(style='thin')))
            excellstyle.style_range(combine_s5, "H" + str(opercnt1) + ":" + 'H' + str(opercnt2),
                                    border=Border(left=Side(style='thin'), right=Side(style='thin'),
                                                  top=Side(style='thin'),
                                                  bottom=Side(style='thin')))

            combine_s5["E" + str(opercnt1)].value = opercnt4
            combine_s5["F" + str(opercnt1)].value = s5['E' + str(opercnt3)].value
            combine_s5["G" + str(opercnt1)].value = s5['F' + str(opercnt3)].value
            combine_s5["H" + str(opercnt1)].value = s5['G' + str(opercnt3)].value
            excellstyle.exel_font_set(combine_s5["E" + str(opercnt1)])
            excellstyle.exel_font_set(combine_s5["F" + str(opercnt1)])
            excellstyle.exel_font_set(combine_s5["G" + str(opercnt1)])
            excellstyle.exel_font_set(combine_s5["H" + str(opercnt1)])


            opercnt3 = opercnt3 + opernamelist5.count(opername)
            opercnt1 = opercnt2 + 1
            opercnt4 = opercnt4 + 1

        for i in range(0, s5_len):
            combine_s5['A' + str(8 + i + s5_cnt)].value = api_name
            combine_s5['B' + str(8 + i + s5_cnt)].value = str(i+1)
            combine_s5['C' + str(8 + i + s5_cnt)].value = s5['B' + str(8 + i)].value
            combine_s5['D' + str(8 + i + s5_cnt)].value = s5['C' + str(8 + i)].value
            excellstyle.exel_font_set(combine_s5['A' + str(8 + i + s5_cnt)])
            excellstyle.exel_font_set(combine_s5['B' + str(8 + i + s5_cnt)])
            excellstyle.exel_font_set(combine_s5['C' + str(8 + i + s5_cnt)])
            excellstyle.exel_font_set(combine_s5['D' + str(8 + i + s5_cnt)])
        s5_cnt = s5_cnt + s5_len


        ##s6항목
        combine_s6 = combine_wb['60.S-6']
        s6 = wb_data['60.S-6']
        s6_len = 0
        while True:
            if s6['A' + str(8 + s6_len)].value == None:
                break
            s6_len = s6_len + 1
        for i in range(0, s6_len):
            combine_s6['A' + str(8 + i + s6_cnt)].value = api_name
            combine_s6['B' + str(8 + i + s6_cnt)].value = s6['A' + str(8 + i)].value
            combine_s6['C' + str(8 + i + s6_cnt)].value = s6['B' + str(8 + i)].value
            combine_s6['D' + str(8 + i + s6_cnt)].value = s6['C' + str(8 + i)].value
            combine_s6['E' + str(8 + i + s6_cnt)].value = s6['D' + str(8 + i)].value
            combine_s6['F' + str(8 + i + s6_cnt)].value = s6['E' + str(8 + i)].value
            combine_s6['G' + str(8 + i + s6_cnt)].value = s6['F' + str(8 + i)].value
            combine_s6['H' + str(8 + i + s6_cnt)].value = s6['G' + str(8 + i)].value
            combine_s6['I' + str(8 + i + s6_cnt)].value = s6['H' + str(8 + i)].value
            combine_s6['J' + str(8 + i + s6_cnt)].value = s6['I' + str(8 + i)].value
            combine_s6['K' + str(8 + i + s6_cnt)].value = s6['J' + str(8 + i)].value
            combine_s6['L' + str(8 + i + s6_cnt)].value = s6['K' + str(8 + i)].value
            combine_s6['M' + str(8 + i + s6_cnt)].value = s6['L' + str(8 + i)].value
            combine_s6['N' + str(8 + i + s6_cnt)].value = s6['M' + str(8 + i)].value
            excellstyle.exel_font_set(combine_s6['A' + str(8 + i + s6_cnt)])
            excellstyle.exel_font_set(combine_s6['B' + str(8 + i + s6_cnt)])
            excellstyle.exel_font_set(combine_s6['C' + str(8 + i + s6_cnt)])
            excellstyle.exel_font_set(combine_s6['D' + str(8 + i + s6_cnt)])
            excellstyle.exel_font_set(combine_s6['E' + str(8 + i + s6_cnt)])
            excellstyle.exel_font_set(combine_s6['F' + str(8 + i + s6_cnt)])
            excellstyle.exel_font_set(combine_s6['G' + str(8 + i + s6_cnt)])
            excellstyle.exel_font_set(combine_s6['H' + str(8 + i + s6_cnt)])
            excellstyle.exel_font_set(combine_s6['I' + str(8 + i + s6_cnt)])
            excellstyle.exel_font_set(combine_s6['J' + str(8 + i + s6_cnt)])
            excellstyle.exel_font_set(combine_s6['K' + str(8 + i + s6_cnt)])
            excellstyle.exel_font_set(combine_s6['L' + str(8 + i + s6_cnt)])
            excellstyle.exel_font_set(combine_s6['M' + str(8 + i + s6_cnt)])
            excellstyle.exel_font_set(combine_s6['N' + str(8 + i + s6_cnt)])

















        s6_cnt = s6_cnt + s6_len

        report_cnt = report_cnt + 1
    combine_wb.save(result)


