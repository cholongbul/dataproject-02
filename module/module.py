import cx_Oracle
import pandas as pd
import traceback
from urllib import parse, request
from urllib.error import HTTPError
from bs4 import BeautifulSoup
import ssl
import requests
import excellstyle
from difflib import SequenceMatcher
from openpyxl.styles import Font, Border, Side, Alignment
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
import os
import openpyxl
import re

def dbcursor_module():
    dsn = cx_Oracle.makedsn('localhost', 1521, 'orcl')
    db = cx_Oracle.connect('DQM_01', 'DQM_01', dsn)
    cursor = db.cursor()

    return cursor

def opernm_list(cursor, apiid):
    cursor.execute("""SELECT DISTINCT APICD, OPERNM FROM OPERATION_REQ WHERE APICD = """ + "'" + apiid + "'")
    row = cursor.fetchall()
    colname = cursor.description
    col = []
    for i in colname:
        col.append(i[0])

    apilist_df = pd.DataFrame(row, columns=col)
    opernm_list = apilist_df['OPERNM'].tolist()
    return opernm_list

def apilist_df_module(cursor, apiid):
    cursor.execute("""select * from api where apicd = """ + "'" + apiid + "'")
    row = cursor.fetchall()
    colname = cursor.description
    col = []
    for i in colname:
        col.append(i[0])

    apilist_df = pd.DataFrame(row, columns=col)
    return apilist_df

def oper_req_df_A(cursor, apiid):
    cursor.execute(
        """select * from operation_REQ where apicd = """ + "'" + str(int(apiid)) + "'")
    row = cursor.fetchall()
    colname = cursor.description
    col = []
    for xxx in colname:
        col.append(xxx[0])
    oper_req_df = pd.DataFrame(row, columns=col)

    return oper_req_df

def oper_req_df_A_B(cursor, apiid):
    cursor.execute(
        """select distinct opercd from operation_REQ where apicd = """ + "'" + str(int(apiid)) + "'")
    row = cursor.fetchall()
    colname = cursor.description
    col = []
    for xxx in colname:
        col.append(xxx[0])
    oper_req_df = pd.DataFrame(row, columns=col)

    return oper_req_df['OPERCD'].tolist()

def oper_req_df_B(cursor, apiid, opercode):
    cursor.execute(
        """select * from operation_REQ where apicd = """ + "'" + str(int(apiid)) + "' and opercd = '" + str(
            int(opercode)) + "'")
    row = cursor.fetchall()
    colname = cursor.description
    col = []
    for xxx in colname:
        col.append(xxx[0])

    oper_req_df = pd.DataFrame(row, columns=col)

    return oper_req_df

def oper_res_df_module(cursor, apiid, opercode):
    cursor.execute(
        """select * from operation_REs where apicd = """ + "'" + str(int(apiid)) + "' and opercd = '" + str(
            int(opercode)) + "'")
    row = cursor.fetchall()
    colname = cursor.description
    col = []
    for xxx in colname:
        col.append(xxx[0])
    oper_res_df = pd.DataFrame(row, columns=col)
    return oper_res_df

def oper_code_name_setter(df_data,opername_list,oper_req_df):
    code_name = []
    if str(type(df_data.iloc[0, 0])) == "<class 'str'>":
        opername = df_data.iloc[0, 0]
        opername_list.append(opername)

    else:
        opercode = int(df_data.iloc[0, 0])
        opername = ''

    try:
        if opername == '':
            opername = oper_req_df[oper_req_df['OPERCD'] == str(int(opercode))].iloc[0]['OPERNM']
            opername_list.append(opername)
        else:
            opercode = oper_req_df[oper_req_df['OPERNM'] == opername].iloc[0]['OPERCD']
        is_opername = oper_req_df['OPERNM'] == opername
        same_oper_df = oper_req_df[is_opername]
        if len(same_oper_df['OPERNM'].tolist()) < 1:
            print(opername)
            exit()
        code_name.append(opercode)
        code_name.append(opername)
    except:
        print(opername)
        traceback.print_exc()
        exit()

    return code_name

def mkApilink(df_data,servicekey_tail):
    try:
        api_req_link = str(df_data.iloc[0, 12]).replace(" ", '').replace("\n", "").replace("\t", "").replace("\r", "")
        api_req_link_head = api_req_link.split('?')[0]
        api_req_link_param_all = api_req_link.split('?')[1]
        api_req_link_param_list = api_req_link_param_all.split('&')
        changed_param = ''
        for api_req_param in api_req_link_param_list:
            if api_req_param.lower().startswith('servicekey'):
                servicekey_head = api_req_param.split('=')[0]
                api_req_param = servicekey_head + '=' + servicekey_tail
            if api_req_param.lower().startswith('authapikey'):
                servicekey_head = api_req_param.split('=')[0]
                api_req_param = servicekey_head + '=' + servicekey_tail
            if changed_param == '':
                changed_param = changed_param + api_req_param
            else:
                changed_param = changed_param + '&' + api_req_param

        api_req_link = api_req_link_head + '?' + changed_param
        api_req_link_head = api_req_link.split('?')[0]
    except:
        if str(df_data.iloc[0, 12]) == 'nan':
            api_req_link = '요청메세지 없음'
        elif not 'http' in str(df_data.iloc[0, 12]):
            api_req_link = '요청메세지 없음'
        else:
            api_req_link = str(df_data.iloc[0, 12]).replace(" ", '').replace("\n", "").replace("\t", "").replace("\r", "")

    api_req_link = parse.urlparse(api_req_link)
    query = parse.parse_qs(api_req_link.query)
    api_req_link = api_req_link_head + '?' + parse.urlencode(query, doseq=True)

    return api_req_link

def req_html_module(api_req):
    context = ssl._create_unverified_context()

    try:
        req_html = request.urlopen(api_req, data=None, context=context)
    except HTTPError as e:
        req_html = e
    req_html = req_html.read().decode("utf-8")
    req_html = BeautifulSoup(req_html, 'html.parser')
    req_html = str(req_html)

    return req_html

def json_xml(api_req_link,oper_req_df, req_html):
    req_html_list = []
    if req_html == '이미지':
        req_html2 = '이미지'
        req_html3 = '이미지'
    else:
        jsonurl_main1 = str(api_req_link).split('?')[0]
        jsonurl_main2 = str(api_req_link).split('?')[0]
        try:
            jsonurl_pram_list = str(api_req_link).split('?')[1].split('&')
        except IndexError:
            jsonurl_pram_list = []
        returnparam_list = ['_returnType', '_type', 'act', 'apiType', 'contentType', 'dataFormat', 'dataFormat',
                            'dataType', 'dateType'
            , 'format', 'output', 'resultType', 'retunType', 'returnType', 'service_Type', 'type', 'viewType',
                            'ViewType']
        selected_returnparam = ''

        for returnparam in returnparam_list:
            if returnparam in oper_req_df['COLNMEN'].tolist():
                selected_returnparam = returnparam
                break

        for jsonurl_pram_i in range(0, len(jsonurl_pram_list)):
            if selected_returnparam == '':
                if jsonurl_pram_i == 0:
                    jsonurl_main1 = jsonurl_main1 + '?' + jsonurl_pram_list[jsonurl_pram_i]
                    jsonurl_main2 = jsonurl_main2 + '?' + jsonurl_pram_list[jsonurl_pram_i]
                elif jsonurl_pram_i == len(jsonurl_pram_list) - 1:
                    jsonurl_main1 = jsonurl_main1 + '&' + jsonurl_pram_list[
                        jsonurl_pram_i] + '&resultType=xml'
                    jsonurl_main2 = jsonurl_main2 + '&' + jsonurl_pram_list[
                        jsonurl_pram_i] + '&resultType=json'
                else:
                    jsonurl_main1 = jsonurl_main1 + '&' + jsonurl_pram_list[jsonurl_pram_i]
                    jsonurl_main2 = jsonurl_main2 + '&' + jsonurl_pram_list[jsonurl_pram_i]

            elif selected_returnparam in str(api_req_link):
                if jsonurl_pram_list[jsonurl_pram_i].startswith(selected_returnparam):
                    if jsonurl_pram_i == 0:
                        jsonurl_main1 = jsonurl_main1 + '?' + selected_returnparam + '=xml'
                        jsonurl_main2 = jsonurl_main2 + '?' + selected_returnparam + '=json'
                    else:
                        jsonurl_main1 = jsonurl_main1 + '&' + selected_returnparam + '=xml'
                        jsonurl_main2 = jsonurl_main2 + '&' + selected_returnparam + '=json'
                else:
                    if jsonurl_pram_i == 0:
                        jsonurl_main1 = jsonurl_main1 + '?' + jsonurl_pram_list[jsonurl_pram_i]
                        jsonurl_main2 = jsonurl_main2 + '?' + jsonurl_pram_list[jsonurl_pram_i]
                    else:
                        jsonurl_main1 = jsonurl_main1 + '&' + jsonurl_pram_list[jsonurl_pram_i]
                        jsonurl_main2 = jsonurl_main2 + '&' + jsonurl_pram_list[jsonurl_pram_i]
            else:
                if jsonurl_pram_i == 0:
                    jsonurl_main1 = jsonurl_main1 + '?' + jsonurl_pram_list[jsonurl_pram_i]
                    jsonurl_main2 = jsonurl_main2 + '?' + jsonurl_pram_list[jsonurl_pram_i]

                elif jsonurl_pram_i == len(jsonurl_pram_list) - 1:
                    jsonurl_main1 = jsonurl_main1 + '&' + jsonurl_pram_list[
                        jsonurl_pram_i] + '&' + selected_returnparam + '=xml'
                    jsonurl_main2 = jsonurl_main2 + '&' + jsonurl_pram_list[
                        jsonurl_pram_i] + '&' + selected_returnparam + '=json'

                else:
                    jsonurl_main1 = jsonurl_main1 + '&' + jsonurl_pram_list[jsonurl_pram_i]
                    jsonurl_main2 = jsonurl_main2 + '&' + jsonurl_pram_list[jsonurl_pram_i]
        jsonurl_main1 = jsonurl_main1.replace(' ', '').replace('\n', '').replace('\t', '').replace('\r', '')
        jsonurl_main2 = jsonurl_main2.replace(' ', '').replace('\n', '').replace('\t', '').replace('\r', '')
        try:
            req_html2 = requests.get(jsonurl_main1).text.strip()


        except:
            req_html2 = "응답없음"

        try:
            req_html3 = requests.get(jsonurl_main2).text.strip()


        except:
            req_html3 = "응답없음"

    req_html_list.append(req_html2)
    req_html_list.append(req_html3)

    return req_html_list

def s3m(s3, oper_req_df, df_data, opername, result_list1, num1_cnt1):
    num1htmllist = oper_req_df['COLNMEN'].tolist()
    num1reqlist = df_data['항목명(영문)'].values.tolist()
    resultlist1_temp = []
    num1reqlist_temp = []
    num1htmllist_temp = []
    for j in num1reqlist:  ##리소스에 적혀있는 요청변수만큼 반복한다
        if str(j) == 'nan':  ##만약 리소스에 적혀 있는 요청변수가 비어있다면 반복을 멈춘다
            break
        if len(num1htmllist) == 0:  ##만약 포털이 비어있다면 멈춘다
            break
        if num1htmllist[0] == '검색 결과가 없습니다.':
            break
        for jk in num1htmllist:

            if str(j).lower().strip().replace(" ", "") == str(jk).lower().strip().replace(" ", ""):
                num1reqlist_temp.append(j)
                num1htmllist_temp.append(jk)

                num1_cnt1 = num1_cnt1 + 1
                s3['A' + str(8 + num1_cnt1 - 1)].value = str(num1_cnt1)
                excellstyle.exel_font_set(s3['A' + str(8 + num1_cnt1 - 1)])

                s3['B' + str(8 + num1_cnt1 - 1)].value = opername
                excellstyle.exel_font_set(s3['B' + str(8 + num1_cnt1 - 1)])

                s3['C' + str(8 + num1_cnt1 - 1)].value = j
                excellstyle.exel_font_set(s3['C' + str(8 + num1_cnt1 - 1)])
                try:
                    s3['D' + str(8 + num1_cnt1 - 1)].value = \
                        df_data[df_data['항목명(영문)'] == j]['항목명(국문)'].values[0]
                    excellstyle.exel_font_set(s3['D' + str(8 + num1_cnt1 - 1)])
                except IndexError:
                    s3['D' + str(8 + num1_cnt1 - 1)].value = ''
                    excellstyle.exel_font_set(s3['D' + str(8 + num1_cnt1 - 1)])
                try:
                    s3['E' + str(8 + num1_cnt1 - 1)].value = \
                        df_data[df_data['항목명(영문)'] == j]['항목크기'].values[0]
                    excellstyle.exel_font_set(s3['E' + str(8 + num1_cnt1 - 1)])
                except IndexError:
                    s3['E' + str(8 + num1_cnt1 - 1)].value = ''
                    excellstyle.exel_font_set(s3['E' + str(8 + num1_cnt1 - 1)])
                try:
                    s3['F' + str(8 + num1_cnt1 - 1)].value = \
                        df_data[df_data['항목명(영문)'] == j]['항목구분'].values[0]
                    excellstyle.exel_font_set(s3['F' + str(8 + num1_cnt1 - 1)])
                except IndexError:
                    s3['F' + str(8 + num1_cnt1 - 1)].value = ''
                    excellstyle.exel_font_set(s3['F' + str(8 + num1_cnt1 - 1)])
                s3['G' + str(8 + num1_cnt1 - 1)].value = str(num1_cnt1)
                excellstyle.exel_font_set(s3['G' + str(8 + num1_cnt1 - 1)])
                s3['H' + str(8 + num1_cnt1 - 1)].value = opername
                excellstyle.exel_font_set(s3['H' + str(8 + num1_cnt1 - 1)])
                s3['I' + str(8 + num1_cnt1 - 1)].value = jk
                excellstyle.exel_font_set(s3['I' + str(8 + num1_cnt1 - 1)])
                s3['J' + str(8 + num1_cnt1 - 1)].value = \
                    oper_req_df[oper_req_df['COLNMEN'] == jk]['COLNMKR'].values[0]
                excellstyle.exel_font_set(s3['J' + str(8 + num1_cnt1 - 1)])
                s3['K' + str(8 + num1_cnt1 - 1)].value = \
                    oper_req_df[oper_req_df['COLNMEN'] == jk]['COLSIZE'].values[0]
                excellstyle.exel_font_set(s3['K' + str(8 + num1_cnt1 - 1)])
                s3['L' + str(8 + num1_cnt1 - 1)].value = \
                    oper_req_df[oper_req_df['COLNMEN'] == jk]['COLTYPE'].values[0]
                excellstyle.exel_font_set(s3['L' + str(8 + num1_cnt1 - 1)])
                s3['M' + str(8 + num1_cnt1 - 1)].value = '정상'
                resultlist1_temp.append(s3['M' + str(8 + num1_cnt1 - 1)].value)
                excellstyle.exel_font_set(s3['M' + str(8 + num1_cnt1 - 1)])

    num1reqlist = [x for x in num1reqlist if x not in num1reqlist_temp]
    num1htmllist = [x for x in num1htmllist if x not in num1htmllist_temp]

    if len(num1reqlist) > 0:
        for j in num1reqlist:
            if str(j) == 'nan':
                continue
            num1_cnt1 = num1_cnt1 + 1
            s3['A' + str(8 + num1_cnt1 - 1)].value = str(num1_cnt1)
            excellstyle.exel_font_set(s3['A' + str(8 + num1_cnt1 - 1)])
            s3['G' + str(8 + num1_cnt1 - 1)].value = str(num1_cnt1)
            excellstyle.exel_font_set(s3['G' + str(8 + num1_cnt1 - 1)])
            s3['B' + str(8 + num1_cnt1 - 1)].value = opername
            excellstyle.exel_font_set(s3['B' + str(8 + num1_cnt1 - 1)])
            s3['H' + str(8 + num1_cnt1 - 1)].value = opername
            excellstyle.exel_font_set(s3['H' + str(8 + num1_cnt1 - 1)])
            s3['C' + str(8 + num1_cnt1 - 1)].value = j
            excellstyle.exel_font_set(s3['C' + str(8 + num1_cnt1 - 1)])
            s3['D' + str(8 + num1_cnt1 - 1)].value = \
                df_data[df_data['항목명(영문)'] == j]['항목명(국문)'].values[0]
            excellstyle.exel_font_set(s3['D' + str(8 + num1_cnt1 - 1)])
            s3['E' + str(8 + num1_cnt1 - 1)].value = \
                df_data[df_data['항목명(영문)'] == j]['항목크기'].values[0]
            excellstyle.exel_font_set(s3['E' + str(8 + num1_cnt1 - 1)])
            s3['F' + str(8 + num1_cnt1 - 1)].value = \
                df_data[df_data['항목명(영문)'] == j]['항목구분'].values[0]
            excellstyle.exel_font_set(s3['F' + str(8 + num1_cnt1 - 1)])
            s3['I' + str(8 + num1_cnt1 - 1)].value = '동일항목없음'
            excellstyle.exel_font_set(s3['I' + str(8 + num1_cnt1 - 1)])
            s3['J' + str(8 + num1_cnt1 - 1)].value = '동일항목없음'
            excellstyle.exel_font_set(s3['J' + str(8 + num1_cnt1 - 1)])
            excellstyle.exel_font_set(s3['K' + str(8 + num1_cnt1 - 1)])
            excellstyle.exel_font_set(s3['L' + str(8 + num1_cnt1 - 1)])
            s3['M' + str(8 + num1_cnt1 - 1)].value = '오류'
            resultlist1_temp.append(s3['M' + str(8 + num1_cnt1 - 1)].value)
            excellstyle.exel_font_set(s3['M' + str(8 + num1_cnt1 - 1)])

    if len(num1htmllist) > 0:
        for j in num1htmllist:
            if str(j) == 'nan':
                continue
            elif str(j) == '검색 결과가 없습니다.':
                continue
            num1_cnt1 = num1_cnt1 + 1
            s3['G' + str(8 + num1_cnt1 - 1)].value = str(num1_cnt1 - 1 + 1)
            excellstyle.exel_font_set(s3['G' + str(8 + num1_cnt1 - 1)])
            s3['A' + str(8 + num1_cnt1 - 1)].value = str(num1_cnt1 - 1 + 1)
            excellstyle.exel_font_set(s3['A' + str(8 + num1_cnt1 - 1)])
            s3['H' + str(8 + num1_cnt1 - 1)].value = opername
            excellstyle.exel_font_set(s3['H' + str(8 + num1_cnt1 - 1)])
            s3['B' + str(8 + num1_cnt1 - 1)].value = opername
            excellstyle.exel_font_set(s3['B' + str(8 + num1_cnt1 - 1)])
            s3['I' + str(8 + num1_cnt1 - 1)].value = j
            excellstyle.exel_font_set(s3['I' + str(8 + num1_cnt1 - 1)])
            s3['J' + str(8 + num1_cnt1 - 1)].value = \
                oper_req_df[oper_req_df['COLNMEN'] == j]['COLNMKR'].values[0]
            excellstyle.exel_font_set(s3['J' + str(8 + num1_cnt1 - 1)])
            s3['K' + str(8 + num1_cnt1 - 1)].value = \
                oper_req_df[oper_req_df['COLNMEN'] == j]['COLSIZE'].values[0]
            excellstyle.exel_font_set(s3['K' + str(8 + num1_cnt1 - 1)])
            s3['L' + str(8 + num1_cnt1 - 1)].value = \
                oper_req_df[oper_req_df['COLNMEN'] == j]['COLTYPE'].values[0]
            excellstyle.exel_font_set(s3['L' + str(8 + num1_cnt1 - 1)])
            s3['M' + str(8 + num1_cnt1 - 1)].value = '오류'
            resultlist1_temp.append(s3['M' + str(8 + num1_cnt1 - 1)].value)
            excellstyle.exel_font_set(s3['M' + str(8 + num1_cnt1 - 1)])
            excellstyle.exel_font_set(s3['B' + str(8 + num1_cnt1 - 1)])
            s3['C' + str(8 + num1_cnt1 - 1)].value = '동일항목없음'
            s3['D' + str(8 + num1_cnt1 - 1)].value = '동일항목없음'
            excellstyle.exel_font_set(s3['C' + str(8 + num1_cnt1 - 1)])
            excellstyle.exel_font_set(s3['D' + str(8 + num1_cnt1 - 1)])
            excellstyle.exel_font_set(s3['E' + str(8 + num1_cnt1 - 1)])
            excellstyle.exel_font_set(s3['F' + str(8 + num1_cnt1 - 1)])

    if '오류' in resultlist1_temp:
        result_list1.append('오류')
    else:
        result_list1.append('정상')

    return_list = [s3,result_list1,num1_cnt1]
    return return_list


def s4m(s4, i, opername, api_req_link, df_data, result_list2, req_html):
    s4['A' + str(8 + i)].value = str(i + 1)
    excellstyle.exel_font_set(s4['A' + str(8 + i)])
    s4['B' + str(8 + i)].value = opername
    excellstyle.exel_font_set(s4['B' + str(8 + i)])
    s4['C' + str(8 + i)].value = api_req_link
    excellstyle.exel_font_set(s4['C' + str(8 + i)])
    s4['D' + str(8 + i)].value = df_data.iloc[0, 13]
    excellstyle.exel_font_set(s4['D' + str(8 + i)])
    s4['E' + str(8 + i)].value = str(i + 1)
    excellstyle.exel_font_set(s4['E' + str(8 + i)])
    s4['F' + str(8 + i)].value = opername
    excellstyle.exel_font_set(s4['F' + str(8 + i)])
    s4['G' + str(8 + i)].value = api_req_link
    excellstyle.exel_font_set(s4['G' + str(8 + i)])
    s4['H' + str(8 + i)].value = req_html[:1000]
    excellstyle.exel_font_set(s4['H' + str(8 + i)])
    matchraio = int(
        SequenceMatcher(None, str(req_html.replace(
            'This XML file does not appear to have any style information associated with it. The document tree is shown below.',
            '')).strip().replace('\n', '').replace('\t', '').replace(' ', '')[:200],
                        str(df_data.iloc[0, 13]).strip().replace('\n', '').replace('\t',
                                                                                   '').replace(' ',
                                                                                               '')[
                        :200]).ratio() * 100)
    if 'normal service' in req_html.lower():
        s4['I' + str(8 + i)].value = "정상"
    elif '<result_msg>ok</result_msg>' in req_html.replace(' ', '').replace('\n', '').lower():
        s4['I' + str(8 + i)].value = "정상"
    elif 'normal_service' in req_html.lower():
        s4['I' + str(8 + i)].value = "정상"
    elif 'non_error' in req_html.lower():
        s4['I' + str(8 + i)].value = "정상"
    elif req_html == '응답없음':
        s4['I' + str(8 + i)].value = "오류"
    elif 'success' in req_html.lower():
        s4['I' + str(8 + i)].value = "정상"
    elif '정상' in req_html:
        s4['I' + str(8 + i)].value = "정상"
    elif '성공' in req_html:
        s4['I' + str(8 + i)].value = "정상"
    elif '이미지' in req_html:
        s4['I' + str(8 + i)].value = "정상"
    elif 'success_info' in req_html.lower():
        s4['I' + str(8 + i)].value = "정상"
    elif matchraio > 70:
        s4['I' + str(8 + i)].value = "정상"
    elif '<successYN>N</successYN>' in req_html.lower():
        s4['I' + str(8 + i)].value = "오류"
    elif 'normal' in req_html.lower():
        s4['I' + str(8 + i)].value = '정상'
    elif 'nomal' in req_html.lower():
        s4['I' + str(8 + i)].value = '정상'
    elif '<successYN>Y</successYN>' in req_html.lower():
        s4['I' + str(8 + i)].value = '정상'
    elif '<resultcode>00' in req_html.lower():
        s4['I' + str(8 + i)].value = '정상'
    elif 'wfs' in req_html.lower():
        s4['I' + str(8 + i)].value = '정상'
    elif 'soapenv' in req_html.lower():
        s4['I' + str(8 + i)].value = "오류"
    elif 'errorcode' in req_html.lower():
        s4['I' + str(8 + i)].value = "오류"
    elif 'service error' in req_html.lower():
        s4['I' + str(8 + i)].value = "오류"
    elif 'SERVICE KEY IS NOT' in req_html.upper():
        s4['I' + str(8 + i)].value = "오류"
    elif 'APPLICATION ERROR' in req_html.upper():
        s4['I' + str(8 + i)].value = "오류"
    elif 'DB ERROR' in req_html.upper():
        s4['I' + str(8 + i)].value = "오류"
    elif 'NODATA ERROR' in req_html.upper():
        s4['I' + str(8 + i)].value = "오류"
    elif 'HTTP ERROR' in req_html.upper():
        s4['I' + str(8 + i)].value = "오류"
    elif 'SERVICETIMEOUT ERROR' in req_html.upper():
        s4['I' + str(8 + i)].value = "오류"
    elif 'INVALID_REQUEST PARAMETER_ERROR' in req_html.upper():
        s4['I' + str(8 + i)].value = "오류"
    elif 'NO MANDATORY REQUEST PARAMETERS ERROR' in req_html.upper():
        s4['I' + str(8 + i)].value = "오류"
    elif 'NO OPENAPI SERVICE ERROR' in req_html.upper():
        s4['I' + str(8 + i)].value = "오류"
    elif 'SERVICE ACCESS DENIED ERROR' in req_html.upper():
        s4['I' + str(8 + i)].value = "오류"
    elif 'LIMITED NUMBER OF SERVICE REQUESTS EXCEEDS ERROR' in req_html.upper():
        s4['I' + str(8 + i)].value = "오류"
    elif 'SERVICE KEY IS NOT REGISTERED ERROR' in req_html.upper():
        s4['I' + str(8 + i)].value = "오류"
    elif 'DEADLINE HAS EXPIRED ERROR' in req_html.upper():
        s4['I' + str(8 + i)].value = "오류"
    elif 'UNREGISTERED IP ERROR' in req_html.upper():
        s4['I' + str(8 + i)].value = "오류"
    elif 'INVALID REQUEST PARAMETER ERROR' in req_html.upper():
        s4['I' + str(8 + i)].value = "오류"
    else:
        s4['I' + str(8 + i)].value = "오류"

    excellstyle.exel_font_set(s4['I' + str(8 + i)])
    result_list2.append(s4['I' + str(8 + i)].value)
    return_list = [s4,result_list2]
    return return_list


def s5m(s5, num1_cnt1, s3, result_list3):
    result3_temp = []
    for j in range(0, num1_cnt1):
        s5['A' + str(8 + j)].value = s3['A' + str(8 + j)].value
        excellstyle.exel_font_set(s5['A' + str(8 + j)])
        s5['B' + str(8 + j)].value = s3['B' + str(8 + j)].value
        excellstyle.exel_font_set(s5['B' + str(8 + j)])
        s5['C' + str(8 + j)].value = s3['C' + str(8 + j)].value
        excellstyle.exel_font_set(s5['C' + str(8 + j)])
        s5['D' + str(8 + j)].value = s3['D' + str(8 + j)].value
        excellstyle.exel_font_set(s5['D' + str(8 + j)])
        s5['E' + str(8 + j)].value = s3['E' + str(8 + j)].value
        excellstyle.exel_font_set(s5['E' + str(8 + j)])
        s5['F' + str(8 + j)].value = s3['F' + str(8 + j)].value
        excellstyle.exel_font_set(s5['F' + str(8 + j)])
        s5['G' + str(8 + j)].value = s3['G' + str(8 + j)].value
        excellstyle.exel_font_set(s5['G' + str(8 + j)])
        s5['H' + str(8 + j)].value = s3['H' + str(8 + j)].value
        excellstyle.exel_font_set(s5['H' + str(8 + j)])
        s5['I' + str(8 + j)].value = s3['I' + str(8 + j)].value
        excellstyle.exel_font_set(s5['I' + str(8 + j)])
        s5['J' + str(8 + j)].value = s3['J' + str(8 + j)].value
        excellstyle.exel_font_set(s5['J' + str(8 + j)])
        s5['K' + str(8 + j)].value = s3['K' + str(8 + j)].value
        excellstyle.exel_font_set(s5['K' + str(8 + j)])
        s5['L' + str(8 + j)].value = s3['L' + str(8 + j)].value
        excellstyle.exel_font_set(s5['L' + str(8 + j)])
        if s5['M' + str(8 + j)].value == None:
            if str(s5['I' + str(8 + j)].value) == '동일항목없음' or str(s5['C' + str(8 + j)].value) == '동일항목없음':
                s5['M' + str(8 + j)].value = '오류'
                result3_temp.append("오류")
            elif str(s5['F' + str(8 + j)].value).startswith('0') and str(
                    s5['L' + str(8 + j)].value).startswith('옵'):
                s5['M' + str(8 + j)].value = '정상'
            elif str(s5['F' + str(8 + j)].value).startswith('0') and str(
                    s5['L' + str(8 + j)].value) == '':
                s5['M' + str(8 + j)].value = '정상'
            elif str(s5['F' + str(8 + j)].value) == '선택' and str(
                    s5['L' + str(8 + j)].value).startswith('옵'):
                s5['M' + str(8 + j)].value = '정상'
            elif '0' in str(s5['F' + str(8 + j)].value) and str(
                    s5['L' + str(8 + j)].value) == '':
                s5['M' + str(8 + j)].value = '정상'
            elif str(s5['F' + str(8 + j)].value) == '선택' and str(
                    s5['L' + str(8 + j)].value) == '':
                s5['M' + str(8 + j)].value = '정상'
            elif '0' in str(s5['F' + str(8 + j)].value) and str(
                    s5['L' + str(8 + j)].value) == 'nan':
                s5['M' + str(8 + j)].value = '정상'
            elif str(s5['F' + str(8 + j)].value) == 'nan' and str(
                    s5['L' + str(8 + j)].value).startswith('옵'):
                s5['M' + str(8 + j)].value = '정상'
            elif str(s5['F' + str(8 + j)].value).startswith('옵') and str(
                    s5['L' + str(8 + j)].value).startswith('옵'):
                s5['M' + str(8 + j)].value = '정상'
            elif str(s5['F' + str(8 + j)].value).startswith('필') and str(
                    s5['L' + str(8 + j)].value).startswith('필'):
                s5['M' + str(8 + j)].value = '정상'
            elif str(s5['F' + str(8 + j)].value) == '선택' and str(
                    s5['L' + str(8 + j)].value) == 'nan':
                s5['M' + str(8 + j)].value = '정상'
            elif str(s5['F' + str(8 + j)].value) == str(
                    s5['L' + str(8 + j)].value):
                s5['M' + str(8 + j)].value = '정상'
            elif str(s5['F' + str(8 + j)].value).startswith('1') and str(
                    s5['L' + str(8 + j)].value).startswith('1'):
                s5['M' + str(8 + j)].value = '정상'
            elif str(s5['F' + str(8 + j)].value).startswith('0') and str(
                    s5['L' + str(8 + j)].value).startswith('0'):
                s5['M' + str(8 + j)].value = '정상'
            elif str(s5['F' + str(8 + j)].value).startswith('0') and str(
                    s5['L' + str(8 + j)].value).startswith('필'):
                s5['M' + str(8 + j)].value = '오류'
                result3_temp.append("오류")
            elif str(s5['F' + str(8 + j)].value).startswith('0') and str(
                    s5['L' + str(8 + j)].value) == 'nan':
                s5['M' + str(8 + j)].value = '정상'
            elif str(s5['F' + str(8 + j)].value).startswith('1') and str(
                    s5['L' + str(8 + j)].value).startswith('필'):
                s5['M' + str(8 + j)].value = '정상'
            else:
                s5['M' + str(8 + j)].value = '오류'
                result3_temp.append("오류")
            excellstyle.exel_font_set(s5['M' + str(8 + j)])

    if '오류' in result3_temp:
        result_list3.append('오류')

    return_list = [s5,result_list3]
    return return_list

def s6m(s6, res_name_index, df_data, s6cnt, opername, req_html, res_eng_name_list, real_tagnames, result_list4, i):
    if len(res_name_index) == 0 or str(df_data.iloc[0, 2]) == 'nan':
        s6['A' + str(8 + s6cnt)].value = str(1 + s6cnt)
        excellstyle.exel_font_set(s6['A' + str(8 + s6cnt)])
        s6['B' + str(8 + s6cnt)].value = opername
        excellstyle.exel_font_set(s6['B' + str(8 + s6cnt)])
        s6['C' + str(8 + s6cnt)].value = '항목없음'
        excellstyle.exel_font_set(s6['C' + str(8 + s6cnt)])
        s6['D' + str(8 + s6cnt)].value = str(8 + s6cnt - 7)
        excellstyle.exel_font_set(s6['D' + str(8 + s6cnt)])
        s6['E' + str(8 + s6cnt)].value = opername
        excellstyle.exel_font_set(s6['E' + str(8 + s6cnt)])
        s6['F' + str(8 + s6cnt)].value = req_html[:1000]
        if req_html == "응답없음":
            s6['G' + str(8 + s6cnt)].value = '오류'
        elif '<result_msg>ok</result_msg>' in req_html.replace(' ', '').replace('\n', '').lower():
            s6['G' + str(8 + s6cnt)].value = '정상'
        elif '이미지' in req_html:
            s6['G' + str(8 + s6cnt)].value = '정상'
        elif 'normal' in req_html.lower():
            s6['G' + str(8 + s6cnt)].value = '정상'
        elif 'non_error' in req_html.lower():
            s6['G' + str(8 + s6cnt)].value = '정상'
        elif 'nomal' in req_html.lower():
            s6['G' + str(8 + s6cnt)].value = '정상'
        elif '<successyn>y</successyn>' in req_html.lower():
            s6['G' + str(8 + s6cnt)].value = '정상'
        elif '정상' in req_html.lower():
            s6['G' + str(8 + s6cnt)].value = '정상'
        elif '성공' in req_html.lower():
            s6['G' + str(8 + s6cnt)].value = '정상'
        elif 'success' in req_html.lower():
            s6['G' + str(8 + s6cnt)].value = '정상'
        elif '>ok<' in req_html.lower():
            s6['G' + str(8 + s6cnt)].value = '정상'
        elif 'wfs' in req_html.lower():
            s6['G' + str(8 + s6cnt)].value = '정상'
        elif '<resultcode>00' in req_html.lower():
            s6['G' + str(8 + s6cnt)].value = '정상'
        elif 'soapenv' in req_html.lower():
            s6['G' + str(8 + s6cnt)].value = "오류"
        elif '<successyn>n</successyn>' in req_html.lower():
            s6['G' + str(8 + s6cnt)].value = "오류"
        elif 'errorcode' in req_html.lower():
            s6['G' + str(8 + s6cnt)].value = "오류"
        elif 'service error' in req_html.lower():
            s6['G' + str(8 + s6cnt)].value = "오류"
        elif 'SERVICE KEY IS NOT' in req_html.upper():
            s6['G' + str(8 + s6cnt)].value = "오류"
        elif 'APPLICATION ERROR' in req_html.upper():
            s6['G' + str(8 + s6cnt)].value = "오류"
        elif 'DB ERROR' in req_html.upper():
            s6['G' + str(8 + s6cnt)].value = "오류"
        elif 'NODATA ERROR' in req_html.upper():
            s6['G' + str(8 + s6cnt)].value = "오류"
        elif 'HTTP ERROR' in req_html.upper():
            s6['G' + str(8 + s6cnt)].value = "오류"
        elif 'SERVICETIMEOUT ERROR' in req_html.upper():
            s6['G' + str(8 + s6cnt)].value = "오류"
        elif 'INVALID_REQUEST PARAMETER_ERROR' in req_html.upper():
            s6['G' + str(8 + s6cnt)].value = "오류"
        elif 'NO MANDATORY REQUEST PARAMETERS ERROR' in req_html.upper():
            s6['G' + str(8 + s6cnt)].value = "오류"
        elif 'NO OPENAPI SERVICE ERROR' in req_html.upper():
            s6['G' + str(8 + s6cnt)].value = "오류"
        elif 'SERVICE ACCESS DENIED ERROR' in req_html.upper():
            s6['G' + str(8 + s6cnt)].value = "오류"
        elif 'LIMITED NUMBER OF SERVICE REQUESTS EXCEEDS ERROR' in req_html.upper():
            s6['G' + str(8 + s6cnt)].value = "오류"
        elif 'SERVICE KEY IS NOT REGISTERED ERROR' in req_html.upper():
            s6['G' + str(8 + s6cnt)].value = "오류"
        elif 'DEADLINE HAS EXPIRED ERROR' in req_html.upper():
            s6['G' + str(8 + s6cnt)].value = "오류"
        elif 'UNREGISTERED IP ERROR' in req_html.upper():
            s6['G' + str(8 + s6cnt)].value = "오류"
        elif 'INVALID REQUEST PARAMETER ERROR' in req_html.upper():
            s6['G' + str(8 + s6cnt)].value = "오류"
        elif len(set(res_eng_name_list)) == 0:
            pass
        elif len(set(res_eng_name_list) - set(real_tagnames)) / len(set(res_eng_name_list)) < 0.7:
            s6['G' + str(8 + s6cnt)].value = '정상'
        else:
            s6['G' + str(8 + s6cnt)].value = '오류'
        excellstyle.exel_font_set(s6['F' + str(8 + s6cnt)])
        excellstyle.exel_font_set(s6['G' + str(8 + s6cnt)])
        result_list4.append(s6['G' + str(8 + s6cnt)].value)
        s6cnt = s6cnt + 1
    else:
        s6.merge_cells("D" + str(8 + s6cnt) + ":" + 'D' + str(
            8 + s6cnt + len(res_name_index) - 1))
        s6.merge_cells("E" + str(8 + s6cnt) + ":" + 'E' + str(
            8 + len(res_name_index) + s6cnt - 1))
        s6.merge_cells("F" + str(8 + s6cnt) + ":" + 'F' + str(
            8 + len(res_name_index) + s6cnt - 1))
        s6.merge_cells("G" + str(8 + s6cnt) + ":" + 'G' + str(
            8 + len(res_name_index) + s6cnt - 1))
        excellstyle.style_range(s6, "D" + str(8 + s6cnt) + ":" + 'D' + str(
            8 + s6cnt + len(res_name_index) - 1),
                                border=Border(left=Side(style='thin'), right=Side(style='thin'),
                                              top=Side(style='thin'),
                                              bottom=Side(style='thin')))
        excellstyle.style_range(s6, "E" + str(8 + s6cnt) + ":" + 'E' + str(
            8 + s6cnt + len(res_name_index) - 1),
                                border=Border(left=Side(style='thin'), right=Side(style='thin'),
                                              top=Side(style='thin'),
                                              bottom=Side(style='thin')))
        excellstyle.style_range(s6, "F" + str(8 + s6cnt) + ":" + 'F' + str(
            8 + s6cnt + len(res_name_index) - 1),
                                border=Border(left=Side(style='thin'), right=Side(style='thin'),
                                              top=Side(style='thin'),
                                              bottom=Side(style='thin')))
        excellstyle.style_range(s6, "G" + str(8 + s6cnt) + ":" + 'G' + str(
            8 + s6cnt + len(res_name_index) - 1),
                                border=Border(left=Side(style='thin'), right=Side(style='thin'),
                                              top=Side(style='thin'),
                                              bottom=Side(style='thin')))
        s6['D' + str(8 + s6cnt)].value = str(i + 1)
        excellstyle.exel_font_set(s6['D' + str(8 + s6cnt)])
        s6['E' + str(8 + s6cnt)].value = opername
        excellstyle.exel_font_set(s6['E' + str(8 + s6cnt)])
        s6['F' + str(8 + s6cnt)].value = req_html[:1000]
        excellstyle.exel_font_set(s6['F' + str(8 + s6cnt)])
        res_eng_name_list = []
        s6cnt2 = s6cnt
        for j in range(0, len(res_name_index)):
            s6['A' + str(8 + s6cnt)].value = str(1 + s6cnt)
            excellstyle.exel_font_set(s6['A' + str(8 + s6cnt)])
            s6['B' + str(8 + s6cnt)].value = opername
            excellstyle.exel_font_set(s6['B' + str(8 + s6cnt)])
            s6['C' + str(8 + s6cnt)].value = df_data.iloc[j, 9]
            excellstyle.exel_font_set(s6['C' + str(8 + s6cnt)])
            res_eng_name_list.append(str(df_data.iloc[j, 9]))
            s6cnt = s6cnt + 1

        if req_html == "응답없음":
            s6['G' + str(8 + s6cnt2)].value = '오류'
        elif '<result_msg>ok</result_msg>' in req_html.replace(' ', '').replace('\n', '').lower():
            s6['G' + str(8 + s6cnt2)].value = '정상'
        elif '이미지' in req_html:
            s6['G' + str(8 + s6cnt2)].value = '정상'
        elif 'non_error' in req_html.lower():
            s6['G' + str(8 + s6cnt2)].value = '정상'
        elif 'normal' in req_html.lower():
            s6['G' + str(8 + s6cnt2)].value = '정상'
        elif '정상' in req_html.lower():
            s6['G' + str(8 + s6cnt2)].value = '정상'
        elif '성공' in req_html.lower():
            s6['G' + str(8 + s6cnt2)].value = '정상'
        elif 'success' in req_html.lower():
            s6['G' + str(8 + s6cnt2)].value = '정상'
        elif '>ok<' in req_html.lower():
            s6['G' + str(8 + s6cnt2)].value = '정상'
        elif 'nomal' in req_html.lower():
            s6['G' + str(8 + s6cnt2)].value = '정상'
        elif '<successYN>Y</successYN>' in req_html.lower():
            s6['G' + str(8 + s6cnt2)].value = '정상'
        elif '<resultcode>00' in req_html.lower():
            s6['G' + str(8 + s6cnt2)].value = '정상'
        elif 'wfs' in req_html.lower():
            s6['G' + str(8 + s6cnt2)].value = '정상'
        elif 'soapenv' in req_html.lower():
            s6['G' + str(8 + s6cnt2)].value = "오류"
        elif '<successYN>N</successYN>' in req_html.lower():
            s6['G' + str(8 + s6cnt2)].value = "오류"
        elif 'errorcode' in req_html.lower():
            s6['G' + str(8 + s6cnt2)].value = "오류"
        elif 'service error' in req_html.lower():
            s6['G' + str(8 + s6cnt2)].value = "오류"
        elif 'SERVICE KEY IS NOT' in req_html.upper():
            s6['G' + str(8 + s6cnt2)].value = "오류"
        elif 'APPLICATION ERROR' in req_html.upper():
            s6['G' + str(8 + s6cnt2)].value = "오류"
        elif 'DB ERROR' in req_html.upper():
            s6['G' + str(8 + s6cnt2)].value = "오류"
        elif 'NODATA ERROR' in req_html.upper():
            s6['G' + str(8 + s6cnt2)].value = "오류"
        elif 'HTTP ERROR' in req_html.upper():
            s6['G' + str(8 + s6cnt2)].value = "오류"
        elif 'SERVICETIMEOUT ERROR' in req_html.upper():
            s6['G' + str(8 + s6cnt2)].value = "오류"
        elif 'INVALID_REQUEST PARAMETER_ERROR' in req_html.upper():
            s6['G' + str(8 + s6cnt2)].value = "오류"
        elif 'NO MANDATORY REQUEST PARAMETERS ERROR' in req_html.upper():
            s6['G' + str(8 + s6cnt2)].value = "오류"
        elif 'NO OPENAPI SERVICE ERROR' in req_html.upper():
            s6['G' + str(8 + s6cnt2)].value = "오류"
        elif 'SERVICE ACCESS DENIED ERROR' in req_html.upper():
            s6['G' + str(8 + s6cnt2)].value = "오류"
        elif 'LIMITED NUMBER OF SERVICE REQUESTS EXCEEDS ERROR' in req_html.upper():
            s6['G' + str(8 + s6cnt2)].value = "오류"
        elif 'SERVICE KEY IS NOT REGISTERED ERROR' in req_html.upper():
            s6['G' + str(8 + s6cnt2)].value = "오류"
        elif 'DEADLINE HAS EXPIRED ERROR' in req_html.upper():
            s6['G' + str(8 + s6cnt2)].value = "오류"
        elif 'UNREGISTERED IP ERROR' in req_html.upper():
            s6['G' + str(8 + s6cnt2)].value = "오류"
        elif 'INVALID REQUEST PARAMETER ERROR' in req_html.upper():
            s6['G' + str(8 + s6cnt2)].value = "오류"
        elif len(set(res_eng_name_list)) == 0:
            pass
        elif len(set(res_eng_name_list) - set(real_tagnames)) / len(set(res_eng_name_list)) < 0.7:
            s6['G' + str(8 + s6cnt2)].value = '정상'
        else:
            s6['G' + str(8 + s6cnt2)].value = '오류'
        excellstyle.exel_font_set(s6['G' + str(8 + s6cnt2)])
        result_list4.append(s6['G' + str(8 + s6cnt2)].value)

    return_list = [s6, result_list4, s6cnt]
    return return_list

def s7m(s7, oper_res_df, s7cnt, opername,req_html,real_tagnames,portal_res_name_index,result_list5, i):
    if len(oper_res_df['APICD'].tolist()) == 0 or oper_res_df['COLNMEN'].tolist()[0] == 'nan':
        s7['A' + str(8 + s7cnt)].value = str(8 + s7cnt - 7)
        excellstyle.exel_font_set(s7['A' + str(8 + s7cnt)])
        s7['B' + str(8 + s7cnt)].value = opername
        excellstyle.exel_font_set(s7['B' + str(8 + s7cnt)])
        s7['C' + str(8 + s7cnt)].value = '항목없음'
        excellstyle.exel_font_set(s7['C' + str(8 + s7cnt)])
        s7['D' + str(8 + s7cnt)].value = str(8 + s7cnt - 7)
        excellstyle.exel_font_set(s7['D' + str(8 + s7cnt)])
        s7['E' + str(8 + s7cnt)].value = opername
        excellstyle.exel_font_set(s7['E' + str(8 + s7cnt)])
        if req_html == '이미지':
            s7['F' + str(8 + s7cnt)].value = '이미지'
        else:
            s7['F' + str(8 + s7cnt)].value = ''

        excellstyle.exel_font_set(s7['F' + str(8 + s7cnt)])
        portal_engname_list = []

        for portal_engname in oper_res_df['COLNMEN'].tolist():
            portal_engname_list.append(portal_engname)
        if req_html == "응답없음":
            s7['G' + str(8 + s7cnt)].value = '오류'
        elif '이미지' in req_html:
            s7['G' + str(8 + s7cnt)].value = '정상'
        elif 'normal' in req_html.lower():
            s7['G' + str(8 + s7cnt)].value = '정상'
        elif 'non_error' in req_html.lower():
            s7['G' + str(8 + s7cnt)].value = '정상'
        elif '정상' in req_html.lower():
            s7['G' + str(8 + s7cnt)].value = '정상'
        elif '성공' in req_html.lower():
            s7['G' + str(8 + s7cnt)].value = '정상'
        elif 'success' in req_html.lower():
            s7['G' + str(8 + s7cnt)].value = '정상'
        elif '>ok<' in req_html.lower():
            s7['G' + str(8 + s7cnt)].value = '정상'
        elif 'nomal' in req_html.lower():
            s7['G' + str(8 + s7cnt)].value = '정상'
        elif '<successyn>y</successyn>' in req_html.lower():
            s7['G' + str(8 + s7cnt)].value = '정상'
        elif '<resultcode>00' in req_html.lower():
            s7['G' + str(8 + s7cnt)].value = '정상'
        elif 'wfs' in req_html.lower():
            s7['G' + str(8 + s7cnt)].value = '정상'
        elif 'soapenv' in req_html.lower():
            s7['G' + str(8 + s7cnt)].value = "오류"
        elif 'errorcode' in req_html.lower():
            s7['G' + str(8 + s7cnt)].value = "오류"
        elif 'service error' in req_html.lower():
            s7['G' + str(8 + s7cnt)].value = "오류"
        elif 'SERVICE KEY IS NOT' in req_html.upper():
            s7['G' + str(8 + s7cnt)].value = "오류"
        elif '<successyn>n</successyn>' in req_html.lower():
            s7['G' + str(8 + s7cnt)].value = "오류"
        elif 'APPLICATION ERROR' in req_html.upper():
            s7['G' + str(8 + s7cnt)].value = "오류"
        elif 'DB ERROR' in req_html.upper():
            s7['G' + str(8 + s7cnt)].value = "오류"
        elif 'NODATA ERROR' in req_html.upper():
            s7['G' + str(8 + s7cnt)].value = "오류"
        elif 'HTTP ERROR' in req_html.upper():
            s7['G' + str(8 + s7cnt)].value = "오류"
        elif 'SERVICETIMEOUT ERROR' in req_html.upper():
            s7['G' + str(8 + s7cnt)].value = "오류"
        elif 'INVALID_REQUEST PARAMETER_ERROR' in req_html.upper():
            s7['G' + str(8 + s7cnt)].value = "오류"
        elif 'NO MANDATORY REQUEST PARAMETERS ERROR' in req_html.upper():
            s7['G' + str(8 + s7cnt)].value = "오류"
        elif 'NO OPENAPI SERVICE ERROR' in req_html.upper():
            s7['G' + str(8 + s7cnt)].value = "오류"
        elif 'SERVICE ACCESS DENIED ERROR' in req_html.upper():
            s7['G' + str(8 + s7cnt)].value = "오류"
        elif 'LIMITED NUMBER OF SERVICE REQUESTS EXCEEDS ERROR' in req_html.upper():
            s7['G' + str(8 + s7cnt)].value = "오류"
        elif 'SERVICE KEY IS NOT REGISTERED ERROR' in req_html.upper():
            s7['G' + str(8 + s7cnt)].value = "오류"
        elif 'DEADLINE HAS EXPIRED ERROR' in req_html.upper():
            s7['G' + str(8 + s7cnt)].value = "오류"
        elif 'UNREGISTERED IP ERROR' in req_html.upper():
            s7['G' + str(8 + s7cnt)].value = "오류"
        elif 'INVALID REQUEST PARAMETER ERROR' in req_html.upper():
            s7['G' + str(8 + s7cnt)].value = "오류"
        elif len(set(portal_engname_list)) == 0:
            pass
        elif len(set(portal_engname_list) - set(real_tagnames)) / len(
                set(portal_engname_list)) < 0.7:
            s7['G' + str(8 + s7cnt)].value = '정상'
        else:
            s7['G' + str(8 + s7cnt)].value = '오류'
        excellstyle.exel_font_set(s7['G' + str(8 + s7cnt)])
        result_list5.append(s7['G' + str(8 + s7cnt)].value)
        s7cnt = s7cnt + 1
    else:
        s7.merge_cells("D" + str(8 + s7cnt) + ":" + 'D' + str(
            8 + len(portal_res_name_index) + s7cnt - 1))
        s7.merge_cells("E" + str(8 + s7cnt) + ":" + 'E' + str(
            8 + len(portal_res_name_index) + s7cnt - 1))
        s7.merge_cells("F" + str(8 + s7cnt) + ":" + 'F' + str(
            8 + len(portal_res_name_index) + s7cnt - 1))
        s7.merge_cells("G" + str(8 + s7cnt) + ":" + 'G' + str(
            8 + len(portal_res_name_index) + s7cnt - 1))
        excellstyle.style_range(s7, "D" + str(8 + s7cnt) + ":" + 'D' + str(
            8 + len(portal_res_name_index) + s7cnt - 1),
                                border=Border(left=Side(style='thin'), right=Side(style='thin'),
                                              top=Side(style='thin'),
                                              bottom=Side(style='thin')))
        excellstyle.style_range(s7, "E" + str(8 + s7cnt) + ":" + 'E' + str(
            8 + len(portal_res_name_index) + s7cnt - 1),
                                border=Border(left=Side(style='thin'), right=Side(style='thin'),
                                              top=Side(style='thin'),
                                              bottom=Side(style='thin')))
        excellstyle.style_range(s7, "F" + str(8 + s7cnt) + ":" + 'F' + str(
            8 + len(portal_res_name_index) + s7cnt - 1),
                                border=Border(left=Side(style='thin'), right=Side(style='thin'),
                                              top=Side(style='thin'),
                                              bottom=Side(style='thin')))
        excellstyle.style_range(s7, "G" + str(8 + s7cnt) + ":" + 'G' + str(
            8 + len(portal_res_name_index) + s7cnt - 1),
                                border=Border(left=Side(style='thin'), right=Side(style='thin'),
                                              top=Side(style='thin'),
                                              bottom=Side(style='thin')))
        s7cnt2 = s7cnt
        for j in range(0, len(oper_res_df['APICD'].tolist())):
            s7['A' + str(8 + s7cnt)].value = str(
                + s7cnt + 1)
            excellstyle.exel_font_set(s7['A' + str(8 + s7cnt)])
            s7['B' + str(8 + s7cnt)].value = opername
            excellstyle.exel_font_set(s7['B' + str(8 + s7cnt)])
            s7['C' + str(8 + s7cnt)].value = oper_res_df['COLNMEN'].tolist()[j]
            excellstyle.exel_font_set(s7['C' + str(8 + s7cnt)])
            s7cnt = s7cnt + 1

        s7['D' + str(8 + s7cnt2)].value = str(i + 1)
        excellstyle.exel_font_set(s7['D' + str(8 + s7cnt2)])
        s7['E' + str(8 + s7cnt2)].value = opername
        excellstyle.exel_font_set(s7['E' + str(8 + s7cnt2)])
        print(req_html)
        s7['F' + str(8 + s7cnt2)].value = ILLEGAL_CHARACTERS_RE.sub(r'', req_html[:1000].replace('\\n', ''))
        excellstyle.exel_font_set(s7['F' + str(8 + s7cnt2)])
        ##포털 항목명 영문 리스트 작성
        portal_engname_list = []

        for portal_engname in oper_res_df['COLNMEN']:
            portal_engname_list.append(portal_engname)
        if req_html == "응답없음":
            s7['G' + str(8 + s7cnt2)].value = '오류'
        elif '이미지' in req_html:
            s7['G' + str(8 + s7cnt2)].value = '정상'
        elif 'normal' in req_html.lower():
            s7['G' + str(8 + s7cnt2)].value = '정상'
        elif 'non_error' in req_html.lower():
            s7['G' + str(8 + s7cnt2)].value = '정상'
        elif '정상' in req_html.lower():
            s7['G' + str(8 + s7cnt2)].value = '정상'
        elif '성공' in req_html.lower():
            s7['G' + str(8 + s7cnt2)].value = '정상'
        elif 'success' in req_html.lower():
            s7['G' + str(8 + s7cnt2)].value = '정상'
        elif '>ok<' in req_html.lower():
            s7['G' + str(8 + s7cnt2)].value = '정상'
        elif 'nomal' in req_html.lower():
            s7['G' + str(8 + s7cnt2)].value = '정상'
        elif '<successyn>y</successyn>' in req_html.lower():
            s7['G' + str(8 + s7cnt2)].value = '정상'
        elif '<resultcode>00' in req_html.lower():
            s7['G' + str(8 + s7cnt2)].value = '정상'
        elif 'wfs' in req_html.lower():
            s7['G' + str(8 + s7cnt2)].value = '정상'
        elif 'soapenv' in req_html.lower():
            s7['G' + str(8 + s7cnt2)].value = "오류"
        elif 'errorcode' in req_html.lower():
            s7['G' + str(8 + s7cnt2)].value = "오류"
        elif 'service error' in req_html.lower():
            s7['G' + str(8 + s7cnt2)].value = "오류"
        elif '<successyn>n</successyn>' in req_html.lower():
            s7['G' + str(8 + s7cnt2)].value = "오류"
        elif 'SERVICE KEY IS NOT' in req_html.upper():
            s7['G' + str(8 + s7cnt2)].value = "오류"
        elif 'APPLICATION ERROR' in req_html.upper():
            s7['G' + str(8 + s7cnt2)].value = "오류"
        elif 'DB ERROR' in req_html.upper():
            s7['G' + str(8 + s7cnt2)].value = "오류"
        elif 'NODATA ERROR' in req_html.upper():
            s7['G' + str(8 + s7cnt2)].value = "오류"
        elif 'HTTP ERROR' in req_html.upper():
            s7['G' + str(8 + s7cnt2)].value = "오류"
        elif 'SERVICETIMEOUT ERROR' in req_html.upper():
            s7['G' + str(8 + s7cnt2)].value = "오류"
        elif 'INVALID_REQUEST PARAMETER_ERROR' in req_html.upper():
            s7['G' + str(8 + s7cnt2)].value = "오류"
        elif 'NO MANDATORY REQUEST PARAMETERS ERROR' in req_html.upper():
            s7['G' + str(8 + s7cnt2)].value = "오류"
        elif 'NO OPENAPI SERVICE ERROR' in req_html.upper():
            s7['G' + str(8 + s7cnt2)].value = "오류"
        elif 'SERVICE ACCESS DENIED ERROR' in req_html.upper():
            s7['G' + str(8 + s7cnt2)].value = "오류"
        elif 'LIMITED NUMBER OF SERVICE REQUESTS EXCEEDS ERROR' in req_html.upper():
            s7['G' + str(8 + s7cnt2)].value = "오류"
        elif 'SERVICE KEY IS NOT REGISTERED ERROR' in req_html.upper():
            s7['G' + str(8 + s7cnt2)].value = "오류"
        elif 'DEADLINE HAS EXPIRED ERROR' in req_html.upper():
            s7['G' + str(8 + s7cnt2)].value = "오류"
        elif 'UNREGISTERED IP ERROR' in req_html.upper():
            s7['G' + str(8 + s7cnt2)].value = "오류"
        elif 'INVALID REQUEST PARAMETER ERROR' in req_html.upper():
            s7['G' + str(8 + s7cnt2)].value = "오류"
        elif len(set(portal_engname_list)) == 0:
            pass
        elif len(set(portal_engname_list) - set(real_tagnames)) / len(
                set(portal_engname_list)) < 0.7:
            s7['G' + str(8 + s7cnt2)].value = '정상'
        else:
            s7['G' + str(8 + s7cnt2)].value = '오류'
        excellstyle.exel_font_set(s7['G' + str(8 + s7cnt2)])
        result_list5.append(s7['G' + str(8 + s7cnt2)].value)

    return_list = [s7, result_list5, s7cnt]

    return return_list




def s8m(df_data, req_name_index,code_df, opername, s8, result_list6, code_cnt):
    print(str(df_data['항목명(영문)'][0]))
    if str(df_data['항목명(영문)'][0]) != 'nan':
        df_data_sorted = df_data.iloc[df_data['항목명(영문)'].str.lower().argsort()]
    else:
        df_data_sorted = df_data
    for j in req_name_index:
        is_sameoper = code_df['오퍼명'] == opername
        is_samename = code_df['항목명(영문)'] == df_data_sorted.iloc[j, 2]
        samename = code_df[is_sameoper & is_samename]
        if len(samename['항목명(영문)']) > 0:
            code_cnt = code_cnt + 1
            s8['A' + str(8 + code_cnt - 1)].value = str(code_cnt)
            excellstyle.exel_font_set(s8['A' + str(8 + code_cnt - 1)])
            s8['B' + str(8 + code_cnt - 1)].value = opername
            excellstyle.exel_font_set(s8['B' + str(8 + code_cnt - 1)])
            s8['C' + str(8 + code_cnt - 1)].value = df_data_sorted.iloc[j, 2]
            excellstyle.exel_font_set(s8['C' + str(8 + code_cnt - 1)])
            s8['D' + str(8 + code_cnt - 1)].value = df_data_sorted.iloc[j, 3]
            excellstyle.exel_font_set(s8['D' + str(8 + code_cnt - 1)])
            s8['E' + str(8 + code_cnt - 1)].value = df_data_sorted.iloc[j, 7]
            excellstyle.exel_font_set(s8['E' + str(8 + code_cnt - 1)])
            s8['F' + str(8 + code_cnt - 1)].value = df_data_sorted.iloc[j, 6]
            excellstyle.exel_font_set(s8['F' + str(8 + code_cnt - 1)])

            s8['G' + str(8 + code_cnt - 1)].value = str(code_cnt)
            excellstyle.exel_font_set(s8['G' + str(8 + code_cnt - 1)])
            s8['H' + str(8 + code_cnt - 1)].value = samename['오퍼명'].values[0]
            excellstyle.exel_font_set(s8['H' + str(8 + code_cnt - 1)])
            s8['I' + str(8 + code_cnt - 1)].value = samename['항목명(영문)'].values[0]
            excellstyle.exel_font_set(s8['I' + str(8 + code_cnt - 1)])
            s8['J' + str(8 + code_cnt - 1)].value = samename['항목명'].values[0]
            excellstyle.exel_font_set(s8['J' + str(8 + code_cnt - 1)])
            s8['K' + str(8 + code_cnt - 1)].value = samename['항목설명'].values[0]
            excellstyle.exel_font_set(s8['K' + str(8 + code_cnt - 1)])
            s8['L' + str(8 + code_cnt - 1)].value = samename['샘플데이터'].values[0]
            excellstyle.exel_font_set(s8['L' + str(8 + code_cnt - 1)])
            if samename['코드표제공여부'].values[0] == True:
                s8['M' + str(8 + code_cnt - 1)].value = '제공'
            else:
                s8['M' + str(8 + code_cnt - 1)].value = '미제공'
            result_list6.append(s8['M' + str(8 + code_cnt - 1)].value)
            excellstyle.exel_font_set(s8['M' + str(8 + code_cnt - 1)])
    returnlist = [s8, result_list6]
    return returnlist

def s9m(dataformmat, i, s9, opername, result_list7, req_html, req_html2, req_html3):
    if dataformmat == 'XML' or dataformmat == 'JSON':
        s9['A' + str(9 + i)].value = str(i + 1)
        excellstyle.exel_font_set(s9['A' + str(9 + i)])
        s9['B' + str(9 + i)].value = opername
        excellstyle.exel_font_set(s9['B' + str(9 + i)])
        s9['C' + str(9 + i)].value = dataformmat
        excellstyle.exel_font_set(s9['C' + str(9 + i)])
        s9['D' + str(9 + i)].value = str(i + 1)
        excellstyle.exel_font_set(s9['D' + str(9 + i)])
        s9['E' + str(9 + i)].value = ILLEGAL_CHARACTERS_RE.sub(r'', req_html[:1000])
        excellstyle.exel_font_set(s9['E' + str(9 + i)])
        if '이미지' in req_html:
            s9['F' + str(9 + i)].value = '정상'
            excellstyle.exel_font_set(s9['F' + str(9 + i)])
            result_list7.append('정상')
        elif dataformmat == 'XML' and '</' in req_html[:1000]:
            s9['F' + str(9 + i)].value = '정상'
            excellstyle.exel_font_set(s9['F' + str(9 + i)])
            result_list7.append('정상')
        elif dataformmat == 'JSON' and '{' in req_html[:1000]:
            s9['F' + str(9 + i)].value = '정상'
            excellstyle.exel_font_set(s9['F' + str(9 + i)])
            result_list7.append('정상')
        else:
            s9['F' + str(9 + i)].value = '오류'
            excellstyle.exel_font_set(s9['F' + str(9 + i)])
            result_list7.append('오류')
    else:
        s9['A' + str(9 + i * 2)].value = str(i * 2 + 1)
        excellstyle.exel_font_set(s9['A' + str(9 + i * 2)])
        s9['A' + str(9 + i * 2 + 1)].value = str(i * 2 + 1 + 1)
        excellstyle.exel_font_set(s9['A' + str(9 + i * 2 + 1)])
        s9['B' + str(9 + 2 * i)].value = opername
        excellstyle.exel_font_set(s9['B' + str(9 + i)])
        s9['B' + str(9 + 2 * i + 1)].value = opername
        excellstyle.exel_font_set(s9['B' + str(9 + 2 * i + 1)])
        s9['C' + str(9 + 2 * i)].value = 'XML'
        excellstyle.exel_font_set(s9['C' + str(9 + 2 * i)])
        s9['C' + str(9 + 2 * i + 1)].value = 'JSON'
        excellstyle.exel_font_set(s9['C' + str(9 + 2 * i + 1)])
        s9['D' + str(9 + 2 * i)].value = str(2 * i + 1)
        excellstyle.exel_font_set(s9['D' + str(9 + 2 * i)])
        s9['D' + str(9 + 2 * i + 1)].value = str(2 * i + 1 + 1)
        excellstyle.exel_font_set(s9['D' + str(9 + 2 * i + 1)])
        s9['E' + str(9 + 2 * i)].value = ILLEGAL_CHARACTERS_RE.sub(r'', req_html2[:1000])
        excellstyle.exel_font_set(s9['E' + str(9 + 2 * i)])
        s9['E' + str(9 + 2 * i + 1)].value = ILLEGAL_CHARACTERS_RE.sub(r'', req_html3[:1000])
        excellstyle.exel_font_set(s9['E' + str(9 + 2 * i + 1)])
        if '이미지' in req_html2:
            s9['F' + str(9 + 2 * i)].value = '정상'
            result_list7.append('정상')
        elif '</' in req_html2[:1000]:
            s9['F' + str(9 + 2 * i)].value = '정상'
            result_list7.append('정상')
        else:
            s9['F' + str(9 + 2 * i)].value = '오류'
            result_list7.append('오류')
        if '이미지' in req_html3:
            s9['F' + str(9 + 2 * i + 1)].value = '정상'
            result_list7.append('정상')
        elif '{' in req_html3[:1000]:
            s9['F' + str(9 + 2 * i + 1)].value = '정상'
            result_list7.append('정상')
        else:
            s9['F' + str(9 + 2 * i + 1)].value = '오류'
            result_list7.append('오류')

        excellstyle.exel_font_set(s9['F' + str(9 + 2 * i + 1)])
        excellstyle.exel_font_set(s9['F' + str(9 + 2 * i)])

    return_list = [s9, result_list7]
    return return_list

def s10m(s10, operation_len, opername_list, result_list1, result_list2, result_list4, result_list8):
    s10cnt = 0
    print(operation_len)
    for chek in range(0, operation_len):
        print(chek)
        print(opername_list[chek])
        s10cnt = s10cnt + 1
        s10['A' + str(9 + chek)].value = s10cnt
        s10['B' + str(9 + chek)].value = opername_list[chek]
        s10['C' + str(9 + chek)].value = result_list1[chek]
        s10['D' + str(9 + chek)].value = result_list2[chek]
        s10['E' + str(9 + chek)].value = result_list4[chek]
        if s10['C' + str(9 + chek)].value == '오류' or s10['D' + str(9 + chek)].value == '오류' or s10[
            'E' + str(9 + chek)].value == '오류':
            s10['F' + str(9 + chek)].value = '오류'
        else:
            s10['F' + str(9 + chek)].value = '정상'

        excellstyle.exel_font_set(s10['A' + str(9 + chek)])
        excellstyle.exel_font_set(s10['B' + str(9 + chek)])
        excellstyle.exel_font_set(s10['C' + str(9 + chek)])
        excellstyle.exel_font_set(s10['D' + str(9 + chek)])
        excellstyle.exel_font_set(s10['E' + str(9 + chek)])
        excellstyle.exel_font_set(s10['F' + str(9 + chek)])
        result_list8.append(s10['F' + str(9 + chek)].value)

    return result_list8

def s1m(s1, organ_name,api_name,api_type,dataformmat,operation_len,docxname,api_portal_url,result_list1,result_list2,result_list3,result_list4,result_list5,result_list6):
    s1['C5'].value = organ_name
    s1['C6'].value = api_name
    s1['C7'].value = api_type
    s1['C8'].value = dataformmat
    s1['C9'].value = operation_len
    s1['C10'].value = docxname
    s1['C11'].value = api_portal_url
    s1['D14'].value = operation_len
    s1['D15'].value = operation_len
    s1['D16'].value = operation_len
    s1['D17'].value = operation_len
    s1['D18'].value = operation_len

    s1['E14'].value = operation_len - result_list1.count('오류')
    s1['F15'].value = operation_len - result_list2.count('오류')
    s1['E16'].value = operation_len - result_list3.count('오류')
    s1['F17'].value = operation_len - result_list4.count('오류')
    s1['E18'].value = operation_len
    s1['E18'].value = operation_len - result_list5.count('오류')

    if len(result_list6) != 0:
        s1['D19'].value = operation_len
        s1['E19'].value = operation_len - result_list5.count('미제공')

def s1m_s5(s1, organ_name,api_name,api_type,dataformmat,operation_len,docxname,api_portal_url,result_list5):
    s1['C5'].value = organ_name
    s1['C6'].value = api_name
    s1['C7'].value = api_type
    s1['C8'].value = dataformmat
    s1['C9'].value = operation_len
    s1['C10'].value = docxname
    s1['C11'].value = api_portal_url

    s1['E18'].value = operation_len
    s1['E18'].value = operation_len - result_list5.count('오류')


def createFolder(directory):
    try:
        if not os.path.exists(directory):
            os.makedirs(directory)
    except OSError:
        print('Error: Creating directory. ' + directory)

def s2m(s2, organ_name,api_name,api_type,dataformmat,operation_len,docxname,api_portal_url,result_list7,result_list8):
    s2['C5'].value = organ_name
    s2['C6'].value = api_name
    s2['C7'].value = api_type
    s2['C8'].value = dataformmat
    s2['C9'].value = operation_len
    s2['C10'].value = docxname
    s2['C11'].value = api_portal_url

    if dataformmat == 'XML' or dataformmat == 'JSON':
        s2['E14'].value = operation_len
        s2['F14'].value = operation_len - result_list7.count('오류')
    else:
        s2['E14'].value = operation_len
        minus_cnt = 0
        flag = True
        for i in range(0, operation_len * 2):
            if i % 2 == 0:
                if result_list7[i] == '오류':
                    minus_cnt = minus_cnt + 1
                    flag = False
                    continue
            elif result_list7[i] == '오류' and flag:
                minus_cnt = minus_cnt + 1
            flag = True
        s2['F14'].value = operation_len - minus_cnt

    s2['D15'].value = operation_len
    s2['F15'].value = operation_len - result_list8.count('오류')

def style_range(ws, cell_range, border=Border(), fill=None, font=None,
                alignment=None):
    """
    Apply styles to a range of cells as if they were a single cell.

    :param ws:  Excel worksheet instance
    :param range: An excel range to style (e.g. A1:F20)
    :param border: An openpyxl Border
    :param fill: An openpyxl PatternFill or GradientFill
    :param font: An openpyxl Font object
    """

    top = Border(top=border.top)
    left = Border(left=border.left)
    right = Border(right=border.right)
    bottom = Border(bottom=border.bottom)

    first_cell = ws[cell_range.split(":")[0]]
    if alignment:
        ws.merge_cells(cell_range)
        first_cell.alignment = alignment

    rows = ws[cell_range]
    if font:
        first_cell.font = font

    for cell in rows[0]:
        cell.border = cell.border + top
    for cell in rows[-1]:
        cell.border = cell.border + bottom

    for row in rows:
        l = row[0]
        r = row[-1]
        l.border = l.border + left
        r.border = r.border + right
        if fill:
            for c in row:
                c.fill = fill


def exel_font_set(sheet):
    sheet.font = Font(name="맑은 고딕")
    sheet.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    sheet.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'),
                          bottom=Side(style='thin'))


def filechk(self, fname_list, fname_list_before):
    if len(fname_list) == 0:
        self.parent().label2.setText("리소스 폴더가 비어있습니다")
        pass
    for fn in fname_list_before:
        if ".xlsx" in fn and not fn.startswith('~'):
            fname_list.append(fn)

    return fname_list

def mergechk(self, sheet, ws_data, result_path, result_file):
    if len(ws_data.merged_cells.ranges) > 0:
        self.parent().label2.setText("보고서 생성 에러: 리소스 " + str(sheet) + " 시트에 병합된 셀이 있습니다.")
        self.parent().label2.repaint()
        if os.path.isfile(result_path + result_file):
            os.remove(result_path + result_file)

def mergechange(path, resource):
    wb_data = openpyxl.load_workbook(path + resource)
    sheetnames = wb_data.sheetnames
    for sheetname in sheetnames:
        ws_data = wb_data[sheetname]
        if len(ws_data.merged_cells.ranges) > 0:
            mergedcell = str(ws_data.merged_cells.ranges)
            celllist = mergedcell.split(',')
            str_col_list = []
            str_row_list = []
            ed_col_list = []
            ed_row_list = []
            gdata_list = []
            hdata_list = []
            kdata_list = []
            for cell in celllist:
                x = cell.strip().split(' ')[1]
                merge_range = x.rstrip(']')
                merge_range = merge_range.rstrip('>')
                print(merge_range)
                ws_data.unmerge_cells(merge_range)
                merge_range_str = merge_range.split(':')[0]
                merge_range_ed = merge_range.split(':')[1]
                str_col = list(merge_range_str)[0]
                str_row = re.sub(r'[^0-9]', '', merge_range_str)
                str_col_list.append(str_col)
                str_row_list.append(str_row)
                ed_col = list(merge_range_ed)[0]
                ed_row = re.sub(r'[^0-9]', '', merge_range_ed)
                ed_col_list.append(ed_col)
                ed_row_list.append(ed_row)
                gdata = str(ws_data['G' + str(int(str_row))].value)
                gdata_list.append(gdata)
                hdata = str(ws_data['H' + str(int(str_row))].value)
                hdata_list.append(hdata)
                kdata = str(ws_data['K' + str(int(str_row))].value)
                kdata_list.append(kdata)
            print(str_col_list)
            for ch_i in range(0,len(str_col_list)):
                if str_col_list[ch_i] == 'C':
                    for i in range(0, int(ed_row_list[ch_i])-int(str_row_list[ch_i])):
                        if str(ws_data['G' + str(int(str_row_list[ch_i]) + i+1)].value) != 'None':
                            gdata_list[ch_i] = str(gdata_list[ch_i]) +" " + str(ws_data['G' + str(int(str_row_list[ch_i]) + i+1)].value)
                        if str(ws_data['H' + str(int(str_row_list[ch_i]) + i+1)].value) != 'None':
                            hdata_list[ch_i] = str(hdata_list[ch_i]) +" " + str(ws_data['H' + str(int(str_row_list[ch_i]) + i+1)].value)
                        ws_data['G' + str(int(str_row_list[ch_i]))].value = gdata_list[ch_i]
                        print('H' + str(int(str_row_list[ch_i]) + i+1))
                        ws_data['G' + str(int(str_row_list[ch_i]) + i+1)].value = None
                        ws_data['H' + str(int(str_row_list[ch_i]))].value = hdata_list[ch_i]
                        ws_data['H' + str(int(str_row_list[ch_i]) + i+1)].value = None
                elif str_col_list[ch_i] == 'J':
                    for i in range(0, int(ed_row_list[ch_i])-int(str_row_list[ch_i])):
                        print(int(ed_row_list[ch_i]))
                        print(int(str_row_list[ch_i]))
                        print(int(ed_row_list[ch_i]) - int(str_row_list[ch_i]))
                        if str(ws_data['K' + str(int(str_row_list[ch_i]) + i + 1)].value) != 'None':
                            kdata_list[ch_i] = str(kdata_list[ch_i]) + " " + str(ws_data['K' + str(int(str_row_list[ch_i]) + i + 1)].value)
                        ws_data['K' + str(int(str_row_list[ch_i]))].value = kdata_list[ch_i]
                        ws_data['K' + str(int(str_row_list[ch_i]) + i+1)].value = None



    wb_data.save(path + resource)